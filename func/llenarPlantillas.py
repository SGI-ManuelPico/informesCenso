import pandas as pd
import re
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from util.descargas import descargarImagenDrive, parseFileId
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D as PositiveSize2D

def insertarImagenConOffset(
    ws,
    drive_service,
    file_id,
    start_row=1,
    start_col=1,
    max_width_px=300,
    max_height_px=200,
    rotate_degrees=0,
    offset_x_px=0,
    offset_y_px=0
):
    """
    Inserta una imagen en la hoja 'ws', partiendo de un file_id de Drive.
    - Mantiene la relación de aspecto (aspect ratio).
    - La imagen no excederá max_width_px ni max_height_px.
    """
    # 1) Descarga la imagen
    img_bytes_original = descargarImagenDrive(drive_service, file_id)

    # 2) Procesar con Pillow
    with PILImage.open(img_bytes_original) as pil_img:
        # Rotar si se indica
        if rotate_degrees:
            pil_img = pil_img.rotate(rotate_degrees, expand=True)

        # Calcular la escala para no sobrepasar max_width_px y max_height_px
        orig_width, orig_height = pil_img.size

        # Razón de escalado en cada eje
        ratio_w = max_width_px / orig_width
        ratio_h = max_height_px / orig_height

        # Tomar la escala mínima para que quepa en ambas dimensiones
        scale = min(ratio_w, ratio_h, 1.0)  
        # el "1.0" es para no agrandar la imagen si es más pequeña

        new_width  = int(orig_width * scale)
        new_height = int(orig_height * scale)

        pil_img = pil_img.resize((new_width, new_height), PILImage.LANCZOS)

        # Guardar en BytesIO
        img_bytes_transformada = io.BytesIO()
        pil_img.save(img_bytes_transformada, format="JPEG")
        img_bytes_transformada.seek(0)

    # 3) Crear imagen openpyxl
    image_para_excel = OpenpyxlImage(img_bytes_transformada)

    # 4) Ancla "OneCellAnchor"
    row0 = start_row - 1
    col0 = start_col - 1

    # Offsets en EMUs
    colOff = int(offset_x_px * 9525)
    rowOff = int(offset_y_px * 9525)

    # Tamaño en EMUs según new_width/new_height
    cx = int(new_width * 9525)
    cy = int(new_height * 9525)

    marker_from = AnchorMarker(col=col0, row=row0, colOff=colOff, rowOff=rowOff)
    size = PositiveSize2D(cx=cx, cy=cy)

    one_cell_anchor = OneCellAnchor(
        _from=marker_from,
        ext=size
    )
    image_para_excel.anchor = one_cell_anchor

    # 5) Insertar en la hoja
    ws.add_image(image_para_excel)



def safe_str(value) -> str:
    """
    Retorna '' si value es None, de lo contrario el valor original.
    """
    return '' if value is None else str(value)
def marcarXdict(ws, dictionary, df_value) -> None:
    """
    Busca en 'dictionary' la celda a usar según 'df_value'.
    Si encuentra una celda válida (por ejemplo "A10"), escribe 'X'.
    Si no la encuentra, no hace nada y evita errores de OpenPyXL.
    """
    if df_value is None:
        return  # No hay valor => no escribir
    cell_addr = dictionary.get(df_value, None)
    if cell_addr:
        ws[cell_addr] = 'X'
def marcarXdict_multiple(ws, dictionary, df_value) -> None:
    """
    Similar a marcarXdict, pero maneja varios valores separados por comas.
    Ej: df_value = "gas,leña" => Se buscan en el diccionario las celdas
    y se marca 'X' en cada una.
    """
    if not df_value:  # Maneja None o cadena vacía
        return
    for part in df_value.split(','):
        part = part.strip()
        cell_addr = dictionary.get(part, None)
        if cell_addr:
            ws[cell_addr] = 'X'
def obtener_url_directa(url) -> str:
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if m:
        file_id = m.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    return url
def valorCol(base_name, index, df_fila) -> str:
    if index == 0:
        return df_fila.get(base_name, '')
    else:
        return df_fila.get(f'{base_name}.{index}', '')
def valSeguro(val):
    """Retorna val si no es nulo/NaN, de lo contrario retorna ''. """
    return val if (pd.notna(val) and val is not None) else ''


def llenarInforme1(ws, df_fila) -> None:
    """
    Llena la plantilla de identificación con los valores de una fila de la base de datos.

    Args:
        ws: Plantilla en la que se va a llenar el informe.
        df_fila: Fila de la base de datos que contiene los valores a llenar en el informe.
    """
    ws['Y1'] = df_fila['data-info_general-num_encuesta']
    if pd.notna(df_fila['data-info_general-fecha']):
        fecha_valor = pd.to_datetime(df_fila['data-info_general-fecha'])
        ws['X2'] = str(fecha_valor.day)
        ws['Z2'] = str(fecha_valor.month)
        ws['AD2'] = str(fecha_valor.year)
    else:
        print("Campo de fecha vacío o inválido.")

    ws['W3'] = df_fila['data-start_act_economica-encuestador']
    ws['A8'] = df_fila['data-info_general-departamento']
    ws['N8'] = df_fila['data-info_general-municipio']
    ws['U8'] = df_fila['data-info_general-vereda']
    ws['A10'] = df_fila['data-start_act_economica-coordenadas']

    if df_fila['data-start_act_economica-permite_entrevista'] == 'yes':
        ws['W10'] = 'X'

        ws['G14'] = df_fila['data-start_act_economica-nombre_establecimiento']
        ws['D15'] = df_fila['data-start_act_economica-direccion']
        ws['U15'] = df_fila['data-start_act_economica-telefono_contacto']
        ws['G16'] = df_fila['data-start_act_economica-actividad_economica']
        ws['W16'] = df_fila['data-start_act_economica-inicio_actividad']
        ws['D17'] = df_fila['data-start_act_economica-propietario']
        ws['Q17'] = df_fila['data-start_act_economica-procedencia_propietario']
        ws['Y17'] = df_fila['data-start_act_economica-lugar_residencia']
        ws['D18'] = df_fila['data-start_act_economica-administrador']
        ws['Q18'] = df_fila['data-start_act_economica-procedencia_administrador']
        ws['Z18'] = df_fila['data-start_act_economica-lugar_residencia_admin']

        actividad = df_fila['data-start_act_economica-actividad_como']
        mapeo_actividad = {
            'natural': 'G23',
            'sociedad_hecho': 'N23',
            'unipersonal': 'G24',
            'sociedad_comercial': 'N24',
            'cooperativa': 'G25',
            'predio': 'G26',
            'other': 'N25'
        }
        if actividad in mapeo_actividad:
            ws[mapeo_actividad[actividad]] = 'X'
            if actividad == 'other':
                ws['K26'] = df_fila['data-start_act_economica-tenencia_propiedad_other']

        actividad2 = df_fila['data-start_act_economica-tipo_actividad']
        mapeo_tipo_actividad = {
            'agricola': 'G30',
            'pecuaria': 'G31',
            'agroindustrial': 'G32',
            'servicios': 'G33',
            'comercial': 'M31',
            'manufactura': 'M32',
            'transporte': 'M33'
        }
        if actividad2 in mapeo_tipo_actividad:
            ws[mapeo_tipo_actividad[actividad2]] = 'X'

        ws['B37'] = df_fila['data-start_act_economica-producto_principal']

        tenencia = df_fila['data-start_act_economica-tenencia_propiedad']
        mapeo_tenencia = {
            'propia': 'E40',
            'administrada': 'E41',
            'arrendada': 'E42',
            'other': 'E44'
        }
        if tenencia in mapeo_tenencia:
            ws[mapeo_tenencia[tenencia]] = 'X'
            if tenencia == 'arrendada':
                ws['J41'] = df_fila['data-start_act_economica-canon_arrendamiento']
            if tenencia == 'other':
                ws['C45'] = df_fila['data-start_act_economica-tenencia_propiedad_other']

        ws['A48'] = df_fila['data-start_act_economica-actividad_ingresos']

        ingresos = df_fila['data-start_act_economica-ingresos']
        mapeo_ingresos = {
            'inferior_smlv': 'Y23',
            'igual_smlv' : 'Y24',
            'entre_1y2_smlv' : 'Y25',
            'superior_2' : 'Y26',
            'other' : 'AD24'
        }
        if ingresos in mapeo_ingresos:
            ws[mapeo_ingresos[ingresos]] = 'X'
            if ingresos == 'other':
                ws['AA25'] = df_fila['data-start_act_economica-ingresos_other']
        
        ws['P30'] = str(df_fila['data-start_act_economica-horario_inicio']) + ', ' + str(df_fila['data-start_act_economica-horario_fin'])

        if df_fila['data-start_act_economica-tiene_registro'] == 'yes':
            ws['T33'] = 'X'
        else:
            ws['W33'] = 'X'

        lugares_comercializa = df_fila['data-start_act_economica-lugares_comercializa']
        mapeo_lugares_comercializa = {
            'sitio': 'X39',
            'empresa': 'X40',
            'mercado': 'X41',
            'acopio': 'X43',
            'other': 'X44'
        }
        if lugares_comercializa in mapeo_lugares_comercializa:
            ws[mapeo_lugares_comercializa[lugares_comercializa]] = 'X'
            if lugares_comercializa == 'other':
                ws['K50'] = df_fila['data-start_act_economica-lugares_comercializa_other']


        frecuencia = df_fila['data-start_act_economica-frecuencia_ingresos']
        mapeo_frecuencia = {
            'diario': 'E53',
            'semanal': 'E54',
            'quincenal': 'E55',
            'mensual': 'M53',
            'semestral': 'M54',
            'other': 'M55'
        }
        if frecuencia in mapeo_frecuencia:
            ws[mapeo_frecuencia[frecuencia]] = 'X'
            if frecuencia == 'other':
                ws['K56'] = df_fila['data-start_act_economica-ingresos_other']


        if df_fila['data-start_act_economica-compra_vereda'] == 'yes':
            ws['T48'] = 'X'
        else:
            ws['W48'] = 'X'


        if df_fila['data-start_act_economica-comercializa_otra_vereda'] == 'yes':
            ws['U53'] = 'X'
            ws['U55'] = df_fila['data-start_act_economica-donde_comercializa']
        else:
            ws['Y53'] = 'X'

        estrato = int(df_fila['data-start_act_economica-estrato'])
        mapeo_estrato = {
            1: 'R59',
            2: 'T59',
            3: 'V59'
        }
        if estrato in mapeo_estrato:
            ws[mapeo_estrato[estrato]] = 'X'

        ws['Y59'] = df_fila['data-start_act_economica-servicios_publicos']

    elif df_fila['data-start_act_economica-permite_entrevista'] == 'no':
        ws['Z10'] = 'X'


def llenarFichaPredial(ws, df1_fila, df_pob_fila, drive_service):

    # 1. DATOS GENERALES DE LA FICHA PREDIAL
    
    ws['U8'] = safe_str(df1_fila['data-info_general-num_encuesta'])
    ws['C8'] = safe_str(df1_fila['data-info_general-fecha'])
    
    ws['C10'] = safe_str(df1_fila['data-info_general-proyecto'])
    ws['S10'] = safe_str(df1_fila['data-info_general-vereda'])
    ws['C12'] = safe_str(df1_fila['data-info_general-municipio'])
    ws['L12'] = safe_str(df1_fila['data-info_general-departamento'])
    ws['T12'] = safe_str(df1_fila['data-info_general-centro_poblado'])
    ws['D14'] = safe_str(df1_fila['data-info_general-nombre_predio'])
    
    mapeo_tenencia = {
        'propia': 'O14',
        'arriendo': 'R14'
    }
    marcarXdict(ws, mapeo_tenencia, df1_fila['data-info_general-tenencia_pert'])
    
    mapeo_uso_para = {
        'trabajo': 'U14',
        'familiar': 'X14'
    }
    marcarXdict(ws, mapeo_uso_para, df1_fila['data-info_general-tenencia_para'])
    
    ws['E16'] = safe_str(df1_fila['data-info_general-nombre_propietario'])
    ws['K16'] = safe_str(df1_fila['data-info_general-telefono_propietario'])
    
    map_vive_predio = {
        'yes': 'R16',
        'no': 'T16'
    }
    marcarXdict(ws, map_vive_predio, df1_fila['data-info_general-vive_en_predio'])
    
    mapa_escriturada = {
        'yes': 'D18',
        'no': 'G18',
        'no_sabe': 'J18' 
    }
    marcarXdict(ws, mapa_escriturada, df1_fila['data-info_general-escriturada'])
    
    ws['Q18'] = safe_str(df1_fila['data-info_general-registro_escritura'])
    ws['H20'] = safe_str(df1_fila['data-info_general-nombre_administrador'])
    ws['T20'] = safe_str(df1_fila['data-info_general-telefono_administrador'])

    # 2. INFORMACIÓN ESPECÍFICA DEL PREDIO
    ws['F24'] = safe_str(df1_fila['data-info_especifica-limite_norte'])
    ws['F26'] = safe_str(df1_fila['data-info_especifica-limite_sur'])
    ws['F28'] = safe_str(df1_fila['data-info_especifica-limite_este'])
    ws['F30'] = safe_str(df1_fila['data-info_especifica-limite_oeste'])

    map_cuenta_con_vivienda = {
        'yes': 'K32',
        'no': 'M32'
    }
    # Ejemplo: si 'data-info_especifica-coord_captacion_este' != None => marcamos 'yes'
    if df1_fila['data-info_especifica-coord_captacion_este'] is not None:
        llave = 'yes'
    else:
        llave = 'no'
    marcarXdict(ws, map_cuenta_con_vivienda, llave)

    ws['S32'] = safe_str(df1_fila['data-info_especifica-punto_gps'])
    ws['L34'] = safe_str(df1_fila['data-info_especifica-coord_captacion_este'])
    ws['S34'] = safe_str(df1_fila['data-info_especifica-coord_captacion_norte'])

    mapa_via_pavimentada = {
        'yes': 'I39',
        'no': 'K39'
    }
    marcarXdict(ws, mapa_via_pavimentada, df1_fila['data-info_especifica-via_municipal'])

    ws['M39'] = safe_str(df1_fila['data-info_especifica-via_municipal_km'])
    
    mapa_estad_via_pavimentada = {
        'B': 'R39',
        'R': 'T39',
        'M': 'W39',
    }
    marcarXdict(ws, mapa_estad_via_pavimentada, df1_fila['data-info_especifica-via_municipal_estado'])

    mapa_trocha = {
        'yes': 'I40',
        'no': 'K40'
    }
    marcarXdict(ws, mapa_trocha, df1_fila['data-info_especifica-trocha'])
    ws['M40'] = safe_str(df1_fila['data-info_especifica-trocha_km'])

    mapa_estad_trocha = {
        'B': 'R40',
        'R': 'T40',
        'M': 'W40',
    }
    marcarXdict(ws, mapa_estad_trocha, df1_fila['data-info_especifica-trocha_estado'])

    mapa_camino_herradura = {
        'yes': 'I41',
        'no': 'K41'
    }
    marcarXdict(ws, mapa_camino_herradura, df1_fila['data-info_especifica-camino_herradura'])
    ws['M41'] = safe_str(df1_fila['data-info_especifica-camino_herradura_km'])

    mapa_estad_camino_herradura = {
        'B': 'R41',
        'R': 'T41',
        'M': 'W41',
    }
    marcarXdict(ws, mapa_estad_camino_herradura, df1_fila['data-info_especifica-camino_herradura_estado'])

    mapa_via_fluvial = {
        'yes': 'I42',
        'no': 'K42'
    }
    marcarXdict(ws, mapa_via_fluvial, df1_fila['data-info_especifica-via_fluvial'])
    ws['M42'] = safe_str(df1_fila['data-info_especifica-via_fluvial_km'])

    mapa_estad_via_fluvial = {
        'B': 'R42',
        'R': 'T42',
        'M': 'W42',
    }
    marcarXdict(ws, mapa_estad_via_fluvial, df1_fila['data-info_especifica-via_fluvial_estado'])

    if df1_fila['data-info_especifica-utilizable_tiempo'] == 'yes':
        ws['I43'] = 'Sí'
    else:
        ws['I43'] = 'No'
    ws['K45'] = safe_str(df1_fila['data-info_especifica-cabecera_cercana'])
    ws['K46'] = safe_str(df1_fila['data-info_especifica-distancia_cabecera'])

    # 3. SERVICIOS PÚBLICOS
    map_energia = {
        'yes': 'E50',
        'no': 'G50'
    }
    marcarXdict(ws, map_energia, df1_fila['data-start_servicios_publicos-energia'])

    map_contador_energia = {
        'yes': 'O50',
        'no': 'R50',
        'other': 'U50'
    }
    val_contador = df1_fila['data-start_servicios_publicos-contador_energia']
    if val_contador == 'other':
        ws['U50'] = safe_str(df1_fila['data-start_servicios_publicos-contador_energia_other'])
    else:
        marcarXdict(ws, map_contador_energia, val_contador)

    map_cocina = {
        'gas': 'E52',
        'leña': 'G52',
        'velas': 'I52',
        'gasolina': 'M52',
        'other': 'R52'
    }
    val_cocina = df1_fila['data-start_servicios_publicos-tipo_cocina']
    if val_cocina == 'other':
        ws['R52'] = safe_str(df1_fila['data-start_servicios_publicos-tipo_cocina_other'])
    else:
        marcarXdict(ws, map_cocina, val_cocina)

    map_acueducto = {
        'yes': 'E54',
        'no': 'G54'
    }
    marcarXdict(ws, map_acueducto, df1_fila['data-start_servicios_publicos-acueducto'])

    # SUMINISTRO DE AGUA
    map_fuente_agua = {
        'pozo_aljibe': 'N54',
        'rio_quebrada': 'N56',
        'recolec_lluvia': 'V54',
        'conex_domici': 'V56',
    }
    marcarXdict(ws, map_fuente_agua, df1_fila['data-start_servicios_publicos-suministro_agua'])

    # ALCANTARILLADO
    map_alcantarillado = {
        'yes': 'E58',
        'no': 'G58'
    }
    marcarXdict(ws, map_alcantarillado, df1_fila['data-start_servicios_publicos-alcantarillado'])

    # DISPOSICIÓN DE AGUAS RESIDUALES
    map_disposicion_aguas = {
        'inodoro': 'N58',
        'letrina': 'N60',
        'campo_abierto': 'V58',
        'pozo_septico': 'V60',
    }
    val_aguas = df1_fila['data-start_servicios_publicos-disposicion_aguas']
    if val_aguas:
        for key in val_aguas.split(','):
            key = key.strip()
            cell = map_disposicion_aguas.get(key, None)
            if cell:
                ws[cell] = 'X'

    # Manejo basuras
    mapa_manejo_basuras = {
        'recoleccion': 'H62',
        'arroja': 'H64',
        'quema': 'N62',
        'arroja_agua': 'N64',
        'entierro': 'V62',
        'otro': 'T64'
    }
    val_basura = df1_fila['data-start_servicios_publicos-manejo_basura']
    if val_basura:
        for key in val_basura.split(','):
            key = key.strip()
            if key == 'other':
                ws['T64'] = safe_str(df1_fila['data-start_servicios_publicos-manejo_basura_other'])
            else:
                cell = mapa_manejo_basuras.get(key, None)
                if cell:
                    ws[cell] = 'X'

    # TELECOMUNICACIONES
    map_telecom = {
        'redes_tel': 'H66',
        'cabina_tel': 'N66',
        'internet': 'V66',
        'other': 'G68'
    }
    val_telecom = df1_fila['data-start_servicios_publicos-telecomunicaciones']
    if val_telecom:
        for key in val_telecom.split(','):
            key = key.strip()
            if key == 'other':
                ws['G68'] = safe_str(df1_fila['data-start_servicios_publicos-telecomunicaciones_other'])
            else:
                cell = map_telecom.get(key, None)
                if cell:
                    ws[cell] = 'X'

    ws['B72'] = safe_str(df1_fila['data-start_servicios_publicos-observaciones_servicios_pub'])

    # 4. SERVICIOS SOCIALES
    mapa_regimen_salud_dueños = {
        'subsidiado': 'P78',
        'contributivo': 'U78',
    }
    marcarXdict(ws, mapa_regimen_salud_dueños, df1_fila['data-start_servicios_sociales-regimen_salud_duenos'])

    mapa_regimen_salud_habitantes = {
        'subsidiado': 'P80',
        'contributivo': 'U80',
    }
    marcarXdict(ws, mapa_regimen_salud_habitantes, df1_fila['data-start_servicios_sociales-regimen_salud_habitantes'])

    # Ejemplo de concatenar 2 columnas
    if (df1_fila['data-start_servicios_sociales-entidad_prestadora_duenos'] is not None and
        df1_fila['data-start_servicios_sociales-entidad_prestadora_habitantes'] is not None):
        ws['J82'] = (safe_str(df1_fila['data-start_servicios_sociales-entidad_prestadora_duenos']) 
                     + ', ' +
                     safe_str(df1_fila['data-start_servicios_sociales-entidad_prestadora_habitantes']))
    else:
        ws['J82'] = ''

    ws['J84'] = safe_str(df1_fila['data-start_servicios_sociales-hospital_mas_cercano'])
    ws['T84'] = safe_str(df1_fila['data-start_servicios_sociales-hospital_localizado'])

    ws['J86'] = safe_str(df1_fila['data-start_servicios_sociales-material_paredes'])
    ws['O86'] = safe_str(df1_fila['data-start_servicios_sociales-material_techo'])
    ws['T86'] = safe_str(df1_fila['data-start_servicios_sociales-material_pisos'])

    ws['J88'] = safe_str(df1_fila['data-start_servicios_sociales-num_habitaciones'])
    if safe_str(df1_fila['data-start_servicios_sociales-distribucion_sala']) == 'yes':
        ws['O88'] = 'Sí'
    else:
        ws['O88'] = 'No'
    if safe_str(df1_fila['data-start_servicios_sociales-distribucion_comedor']) == 'yes':
        ws['S88'] = 'Sí'
    else:
        ws['S88'] = 'No'
    if safe_str(df1_fila['data-start_servicios_sociales-distribucion_cocina']) == 'yes':
        ws['W88'] = 'Sí'
    else:
        ws['W88'] = 'No'
    ws['I90'] = safe_str(df1_fila['data-start_servicios_sociales-area_vivienda'])
    ws['R90'] = safe_str(df1_fila['data-start_carac_poblacion-num_personas'])
    ws['F92'] = safe_str(df1_fila['data-start_servicios_sociales-estado_vivienda'])
    ws['B96'] = safe_str(df1_fila['data-start_servicios_sociales-observaciones_servicios'])

    # La fila 102 es donde comienza la tabla de características de la población, acá iteramos sobre cada fila del df_pob_fila

    df_pob_fila = df_pob_fila.reset_index(drop=True) 

    fila_tabla = 102

    for i, row_pob in df_pob_fila.iterrows():
        fila_excel = fila_tabla + i
        print("Escribiendo en fila:", fila_excel)

        # 1) Nombre
        ws[f"B{fila_excel}"] = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-nombre']

        # 2) Edad
        ws[f"I{fila_excel}"] = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-edad']

        # 3) Género
        genero = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-genero']
        if genero == 'M':
            ws[f"L{fila_excel}"] = 'X'
        elif genero == 'F':
            ws[f"K{fila_excel}"] = 'X'

        # 4) Escolaridad
        ws[f"M{fila_excel}"] = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-escolaridad']

        # 5) Relación con el encargado
        ws[f"P{fila_excel}"] = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-relacion_encargado']

        # 6) Actividad
        ws[f"T{fila_excel}"] = row_pob['data-start_carac_poblacion-caracteristicas_poblacion-actividad']

    map_participacion_com = {
        'junta_padres': 'F114',
        'junta_accion_comunal': 'K114',
        'asociacion_empleados': 'Q114',
        'other': 'U114'
    }

    val_part_com = df1_fila.get('data-start_carac_poblacion-participacion_comunal', '')
    if val_part_com == 'other':
        ws['U114'] = safe_str(df1_fila['data-start_carac_poblacion-participacion_comunal_other'])
    else:
        marcarXdict(ws, map_participacion_com, val_part_com)

    ws['E116'] = safe_str(df1_fila['data-start_carac_poblacion-presencia_institucional'])

    # 6. USOS DEL SUELO
    ws['E120'] = safe_str(df1_fila['data-start_usos_suelo-area_total_predio'])
    ws['K120'] = 'X'  # Hectáreas por defecto
    ws['V120'] = safe_str(df1_fila['data-start_usos_suelo-estrato'])

    map_uso_suelo = {
        'ganaderia': 'F122',
        'pastizales': 'F123',
        'agricultura': 'F122',
        'pancoger': 'F123',
        'other': 'F124'
    }
    val_uso_suelo = df1_fila.get('data-start_usos_suelo-uso_suelo', None)
    if val_uso_suelo:
        for key in val_uso_suelo.split(','):
            key = key.strip()
            if key == 'other':
                ws['F124'] = safe_str(df1_fila['data-start_usos_suelo-uso_suelo_other'])
            else:
                cell_uso = map_uso_suelo.get(key, None)
                if cell_uso:
                    ws[cell_uso] = 'X'
    
    ws['P126'] = safe_str(df1_fila['data-start_usos_suelo-actividades_complementarias'])
    ws['B131'] = safe_str(df1_fila['data-start_usos_suelo-actividades_culturales'])
    ws['B135'] = safe_str(df1_fila['data-start_usos_suelo-alternativas_reasentamiento'])
    ws['B143'] = safe_str(df1_fila['data-start_usos_suelo-expectativas_familia_proyecto'])
    ws['B151'] = safe_str(df1_fila['data-start_usos_suelo-observaciones'])

    # CEDULA RESPONSABLE

    ws['G159'] = safe_str(df1_fila['data-start_usos_suelo-cc_responsable'])
    
    # FIRMA RESPONSABLE
    url_imagen_firma = df1_fila.get('data-start_usos_suelo-firma_responsable')
    if url_imagen_firma:
        file_id = parseFileId(url_imagen_firma)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=157,      # Fila donde anclar
                start_col=20,       # 'T' es la 20
                max_width_px=150,       # Ancho deseado en píxeles
                max_height_px=60,      # Alto deseado en píxeles
                rotate_degrees=0,   # Sin rotar
                offset_x_px=0,     
                offset_y_px=0
            )
        else:
            ws['G157'] = 'No se pudo extraer file_id del link'
    else:
        ws['G157'] = 'No se encontró firma del responsable'

    # 7. FOTOGRAFÍA DE LA VIVIENDA #Fila 163, Columna B
    url_imagen_vivienda = df1_fila.get('data-foto_vivienda')
    if url_imagen_vivienda:
        file_id = parseFileId(url_imagen_vivienda)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=164,      # Fila donde anclar
                start_col=5,        # 'B' es la 2
                max_width_px=800,       # Ancho deseado en píxeles
                max_height_px=480,      # Alto deseado en píxeles
                rotate_degrees=-90, # Rotar 90 grados
                offset_x_px=0,      
                offset_y_px=10000   # Desplazar 1000 píxeles hacia abajo
            )
        else:
            ws['B163'] = 'No se pudo extraer file_id del link'
    else:
        ws['B163'] = 'No se encontró fotografía de la vivienda'

    


    # 8. ACTIVIDAD ECONÓMICA
    mapa_genera_actividad = {
        'yes': 'K182',
        'no': 'O182'
    }
    marcarXdict(ws, mapa_genera_actividad, df1_fila['data-genera_actividad'])

    ws['K183'] = safe_str(df1_fila['data-info_general-num_encuesta'])

    # 9. USOS Y USUARIOS (BIENES Y SERVICIOS AMBIENTALES)
    mapa_capta_aguas_sup = {
        'yes': 'I188',
        'no': 'L188'
    }
    marcarXdict(ws, mapa_capta_aguas_sup, df1_fila['data-capta_fuentes_superf'])

    ws['S188'] = safe_str(df1_fila['data-info_general-num_encuesta'])

    mapa_capta_aguas_sub = {
        'yes': 'I189',
        'no': 'L189'
    }
    marcarXdict(ws, mapa_capta_aguas_sub, df1_fila['data-capta_fuentes_subt'])

    ws['S189'] = safe_str(df1_fila['data-id_funias'])

def llenarUsosUsuarios(ws, df1_fila, df_usos, drive_service):
    # Reemplazamos "valSeguro" por "safe_str"
    # --------------------------------------------------
    ws['H4']  = safe_str(df1_fila.get('data-info_general-num_encuesta'))
    ws['H6']  = safe_str(df1_fila.get('data-start_bienes_serv-profesional'))
    ws['H8']  = safe_str(df1_fila.get('data-info_general-fecha'))
    ws['P4']  = safe_str(df1_fila.get('data-info_general-departamento'))
    ws['P6']  = safe_str(df1_fila.get('data-info_general-municipio'))
    ws['P8']  = safe_str(df1_fila.get('data-info_general-vereda'))
    ws['P10'] = safe_str(df1_fila.get('data-start_bienes_serv-predio'))
    ws['P12'] = safe_str(df1_fila.get('data-info_general-nombre_propietario'))

    map_tipo_fuente_sup = {
        'rio':       'B19',
        'quebrada':  'C19',
        'arroyo':    'D19',
        'caño':      'E19',
        'canal':     'F19',
        'lago':      'G19',
        'laguna':    'H19',
        'cienaga':   'I19',
        'pantano':   'K19',
        'embalse':   'L19',
        'estero':    'M19',
        'jagüey':    'O19',
        'humedal':   'P19',
        'manantial': 'R19',
    }

    fuente_val = df1_fila['data-start_bienes_serv-tipo_fuente_superficial']
    if fuente_val and fuente_val != 'other':
        marcarXdict(ws, map_tipo_fuente_sup, fuente_val)

    ws['B21'] = safe_str(df1_fila.get('data-start_bienes_serv-nombre_corriente_carto'))
    ws['L21'] = safe_str(df1_fila.get('data-start_bienes_serv-nombre_corriente_local'))

    def get_val(df, idx, col):
        val = df.loc[idx, col]
        return '' if pd.isna(val) else val

    # Mapeos para armar las celdas (columna + fila)
    # --------------------------------------------------
    map_tipo_uso = {
        'principal': 'B',
        'secundario': 'F',
        'terciario': 'J',
        'otro': 'N'
    }

    map_columnas_coordenadas = {
        'principal':  ('C','D','E'),
        'secundario': ('G','H','I'),
        'terciario':  ('K','L','M'),
        'otro':       ('O','P','Q')
    }

    map_filas_usos = {
        'consumo_humano':          27,
        'necesidades_domesticas':  28,
        'agropecuarios_comunitarios': 29,
        'agropecuarios_individuales': 30,
        'energia_hidroelectrica':  31,
        'industriales':            32,
        'mineros':                 33,
        'recreativos_comunitarios':34,
        'recreativos_individuales':35,
        'vertimientos':            36
    }

    for idx in df_usos.index:
        uso_actual = df_usos.loc[idx, 'data-start_bienes_serv-informacion_usos-uso_actual']
        uso_tipo   = df_usos.loc[idx, 'data-start_bienes_serv-informacion_usos-tipo_uso_agua']

        fila_excel = map_filas_usos.get(uso_actual, None)
        col_marca  = map_tipo_uso.get(uso_tipo, None)

        if fila_excel and col_marca:
            # Marcamos la “X” en la celda adecuada (col + fila)
            ws[f'{col_marca}{fila_excel}'] = 'X'

            # Tomamos las 3 columnas de coordenadas
            col_este, col_norte, col_cota = map_columnas_coordenadas[uso_tipo]

            # Obtenemos valores
            val_este  = get_val(df_usos, idx, 'data-start_bienes_serv-informacion_usos-coord_este')
            val_norte = get_val(df_usos, idx, 'data-start_bienes_serv-informacion_usos-coord_norte')
            val_cota  = get_val(df_usos, idx, 'data-start_bienes_serv-informacion_usos-cota_msnm')

            # Asignamos a esas celdas
            ws[f'{col_este}{fila_excel}']  = val_este
            ws[f'{col_norte}{fila_excel}'] = val_norte
            ws[f'{col_cota}{fila_excel}']  = val_cota
 
    ws['B39'] = safe_str(df1_fila.get('data-start_bienes_serv-descripcion'))
    ws['B48'] = safe_str(df1_fila.get('data-start_bienes_serv-observacion'))


    # **FOTOS

    # Foto Principal 1
    url_img1 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_principal-foto1_principal')
    if url_img1:
        file_id = parseFileId(url_img1)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=62,      # Fila donde anclar
                start_col=2,       # Columna B
                max_width_px=360,      # Ancho en píxeles
                max_height_px=200,     # Alto en píxeles
                rotate_degrees=-90,
                offset_x_px=20,    # Desplazamiento dentro de la celda (px)
                offset_y_px=50
            )
        else:
            ws['B62'] = 'N/A'
    else:
        ws['B62'] = 'No se encontró foto 1'


    # Foto Principal 2
    url_img2 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_principal-foto2_principal')
    if url_img2:
        file_id = parseFileId(url_img2)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=65,
                start_col=2, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['B65'] = 'N/A'
    else:
        ws['B65'] = 'No se encontró foto 2'


    # Foto Principal 3
    url_img3 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_principal-foto3_principal')
    if url_img3:
        file_id = parseFileId(url_img3)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=55,
                start_col=2, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['B68'] = 'N/A'
    else:
        ws['B68'] = 'No se encontró foto 3'


    # Foto secundaria 1
    url_img4 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_secundario-foto1_secundario')
    if url_img4:
        file_id = parseFileId(url_img4)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=56,
                start_col=7, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['G62'] = 'N/A'
    else:
        ws['G62'] = 'No se encontró foto 1'


    # Foto secundaria 2
    url_img5 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_secundario-foto2_secundario')
    if url_img5:
        file_id = parseFileId(url_img5)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=57,
                start_col=7, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=500
            )
        else:
            ws['G65'] = 'N/A'
    else:
        ws['G65'] = 'No se encontró foto 2'


    # Foto secundaria 3
    url_img6 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_secundario-foto3_secundario')
    if url_img6:
        file_id = parseFileId(url_img6)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=58,
                start_col=7, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['G68'] = 'N/A'
    else:
        ws['G68'] = 'No se encontró foto 3'


    # Foto terciaria 1
    url_img7 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_terciario-foto1_terciario')
    if url_img7:
        file_id = parseFileId(url_img7)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=59,
                start_col=12, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['L62'] = 'N/A'
    else:
        ws['L62'] = 'No se encontró foto 1'


    # Foto terciaria 2
    url_img8 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_terciario-foto2_terciario')
    if url_img8:
        file_id = parseFileId(url_img8)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=60,
                start_col=12, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['L65'] = 'N/A'
    else:
        ws['L65'] = 'No se encontró foto 2'


    # Foto terciaria 3
    url_img9 = df1_fila.get('data-start_bienes_serv-registro_fotos-uso_terciario-foto3_terciario')
    if url_img9:
        file_id = parseFileId(url_img9)
        if file_id:
            insertarImagenConOffset(
                ws=ws,
                drive_service=drive_service,
                file_id=file_id,
                start_row=61,
                start_col=12, 
                max_width_px=360,
                max_height_px=200,
                rotate_degrees=-90,
                offset_x_px=20,
                offset_y_px=50
            )
        else:
            ws['L68'] = 'N/A'
    else:
        ws['L68'] = 'No se encontró foto 3'


        

def llenarFormatoAgropecuario(ws, df_fila, df_info_comercial, df_explot_avicola, df_info_laboral, df_explot_agricola, df_explot_porcina, df_detalle_jornal):

    ws['AO1'] = df_fila['data-datos_encuesta-num_encuesta']

    if pd.notna(df_fila['data-datos_encuesta-fecha']):
        fecha_str = str(df_fila['data-datos_encuesta-fecha'])
        if '/' in fecha_str:
            ws['AM2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
            ws['AP2'] = fecha_str.split('/')[1]
            ws['AU2'] = fecha_str.split('/')[0]
        elif '-' in fecha_str:
            ws['AM2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
            ws['AP2'] = fecha_str.split('-')[1]
            ws['AU2'] = fecha_str.split('-')[0]
        else:
            print(f'Formato de fecha inesperado: {fecha_str}')
    else:
        print('Campo de fecha vacío')   
    
    ws['AL3'] = df_fila['data-datos_encuesta-encuestador']
    ws['E7'] = df_fila['data-ident_entrevistado-nombre']
    ws['AC7'] = df_fila['data-ident_entrevistado-empresa']
    ws['AQ7'] = df_fila['data-ident_entrevistado-cargo']

    map_pertenece_asoc = {  

        'yes': 'M8',
        'no': 'AO8'
    }

    ws[map_pertenece_asoc.get(df_fila['data-ident_entrevistado-pertenencia_asociacion'], '')] = 'X'
    ws['W8'] = safe_str(df_fila['data-ident_entrevistado-asociacion_cual'])



    # B. INFORMACIÓN E IDENTIFICACIÓN GENERAL DEL PREDIO

    # Área Total
    ws['G11'] = safe_str(df_fila['data-info_general_predio-area_total'])
    # Tipo de uso
    ws['G14'] = safe_str(df_fila['data-info_general_predio-tipo_uso_cultivos'])
    ws['G15'] = safe_str(df_fila['data-info_general_predio-tipo_uso_pastos'])
    ws['G16'] = safe_str(df_fila['data-info_general_predio-tipo_uso_bosque_natural'])
    ws['G17'] = safe_str(df_fila['data-info_general_predio-tipo_uso_rastrojo'])
    ws['G18'] = safe_str(df_fila['data-info_general_predio-tipo_uso_bosque_plantado'])
    ws['AB14'] = safe_str(df_fila['data-info_general_predio-tipo_uso_tierras_erosionadas'])
    ws['AB15'] = safe_str(df_fila['data-info_general_predio-tipo_uso_lagos_lagunas'])
    ws['AB16'] = safe_str(df_fila['data-info_general_predio-tipo_uso_reservorios'])
    ws['AB17'] = safe_str(df_fila['data-info_general_predio-tipo_uso_construcciones'])
    ws['Y18'] = safe_str(df_fila['data-info_general_predio-tipo_uso_otros_area'])

    # Precio de arrendamiento por Ha
    ws['AN13'] = valSeguro(df_fila['data-info_general_predio-precio_arrendamiento'])

    # Observaciones
    ws['AM16'] = df_fila['data-info_general_predio-observaciones']

    # ACTIVIDADES AGROPECUARIAS

        # Explotación agrícola
    
    map_filas = {
        'data-begin_agricola-cultivo-producto': 23,          
        'data-begin_agricola-cultivo-area_cultivada': 24,     
        'data-begin_agricola-cultivo-cosechas_anuales': 25,  
        'data-begin_agricola-cultivo-costos_establecimiento': 26,
        'data-begin_agricola-cultivo-costos_mantenimiento': 27,
        'data-begin_agricola-cultivo-costos_cosecha': 28,
        'data-begin_agricola-cultivo-volumen_produccion':29,
        'data-begin_agricola-cultivo-precio_venta': 30,
        'data-begin_agricola-cultivo-autoconsumo': 31,
        'data-begin_agricola-cultivo-destino_producto': 32
    }

    # Las columnas, una por cada cultivo
    cultivo_col_start = ['V', 'AF', 'AN']   

    if df_fila['data-explota_agricola'] == 'yes':
        df_explot_agricola = df_explot_agricola.reset_index(drop=True).head(3)
        for i, row_cultivo in df_explot_agricola.iterrows():
            col_cultivo = cultivo_col_start[i]
            # Para cada campo en map_filas, escribimos en su fila correspondiente
            for key_campo, fila_destino in map_filas.items():
                valor = row_cultivo.get(key_campo, "")
                celda = f"{col_cultivo}{fila_destino}"
                ws[celda] = valor


        # ws['L32'] = df_fila['¿Destino final del producto?']
        

        continuidad = df_fila['data-begin_agricola-begin_sobre_act-continuar_actividad']
        if pd.notna(continuidad):
            if continuidad == 'yes':
                ws['G36'] = 'X'
                ws['AP36'] = 'X'
            elif continuidad == 'no':
                ws['I36'] = 'X'
                ws['AN36'] = 'X'  
        else:
            print('Campo vacío') 

        produccion = df_fila['data-begin_agricola-begin_sobre_act-ampliar_produccion']
        if pd.notna(produccion):
            if produccion == 'yes':
                ws['Q36'] = 'X'
                ws['AF36'] = 'X'               
            elif produccion == 'no':
                ws['S36'] = 'X'
                ws['AD36'] = 'X'   
        else:
            print('Campo vacío') 

        ws['X33'] = df_fila['data-begin_agricola-begin_sobre_act-porque_continuar'] + ', ' + 'Ampliar:' + df_fila['data-begin_agricola-begin_sobre_act-porque_ampliar']

    if df_fila['data-explota_pecuarios'] == 'yes': ## Hoja data-begin_explot_avicola

        # -------------------------------
        # SECCIÓN LECHE (Leche o Cría)
        # -------------------------------

        razas_leche = [
            ('data-parametros_pecuarios-leche_o_cria-raza_leche_1',
            'data-parametros_pecuarios-leche_o_cria-numero_cabezas_leche_1'),
            ('data-parametros_pecuarios-leche_o_cria-raza_leche_2',
            'data-parametros_pecuarios-leche_o_cria-numero_cabezas_leche_2'),
            ('data-parametros_pecuarios-leche_o_cria-raza_leche_3',
            'data-parametros_pecuarios-leche_o_cria-numero_cabezas_leche_3'),
        ]

        for i, (col_raza, col_cabezas) in enumerate(razas_leche):
            row_f = 41 + i
            # Asignar la raza en N(row_f)
            raza_val = df_fila.get(col_raza, "")
            ws[f'N{row_f}'] = raza_val
            
            # Asignar el # de cabezas en AB(row_f)
            cabezas_val = df_fila.get(col_cabezas, "")
            ws[f'AB{row_f}'] = cabezas_val

        # 17. Número de reses en producción 
        ws['Q44'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-reses_produccion_leche', "")
        # 18. Número de terneros
        ws['Q45'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-numero_terneros_leche', "")
        # 19. Número de novillos
        ws['Q46'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-numero_novillos_leche', "")
        # 20. Número de novillas
        ws['Q47'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-numero_novillas_leche', "")
        # 21. Número de reproductores
        ws['Q48'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-numero_reproductores_leche', "")
        # 22. Número de partos al año
        ws['Q49'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-partos_anio_leche', "")
        # 23. Número de vacas para ordeño
        ws['Q50'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-vacas_ordeño_leche', "")
        # 24. Tiempo de venta después destetado
        ws['Q51'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-tiempo_venta_leche', "")
        # 25. Peso promedio de venta
        ws['Q52'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-peso_promedio_venta_leche', "")
        # 26. Precio promedio ($)
        #  Asegúrate de que encaje con tu plantilla actual)
        ws['N53'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-precio_promedio_leche', "")

        # 27. ¿Cada cuánto produce? (Diario, Semanal, etc.)
        frecuencia_leche = df_fila.get('data-parametros_pecuarios-leche_o_cria-frecuencia_produccion_leche')
        if pd.notna(frecuencia_leche):
            # Ejemplo: si es "Diario" => ws['P54'] = 'X'
            if frecuencia_leche == 'diario':
                ws['P54'] = 'X'
            elif frecuencia_leche == 'semanal':
                ws['W54'] = 'X'
            elif frecuencia_leche == 'otro' or 'other':
                ws['AD54'] = 'X'
            # etc. Ajusta columnas según tu plantilla

        # 28. Promedio de producción (Litros, Botellas)
        ws['M55'] = df_fila.get('data-parametros_pecuarios-leche_o_cria-promedio_produccion_leche', "")
        # Si manejas varios tipos (Litro/Botella), ajústalo a tu layout actual


        # ------------------------------------------------------------------
        # SECCIÓN CARNE
        # ------------------------------------------------------------------
        if df_fila.get('data-explota_carne') == 'yes':
            # Razas carne (similar a leche)
            razas_carne = [
                ('data-carne-raza_carne_1', 'data-carne-numero_cabezas_carne_1'),
                ('data-carne-raza_carne_2', 'data-carne-numero_cabezas_carne_2'),
                ('data-carne-raza_carne_3', 'data-carne-numero_cabezas_carne_3'),
            ]
            for i, (col_raza, col_cabezas) in enumerate(razas_carne):
                row_f = 41 + i
                ws[f'AJ{row_f}'] = df_fila.get(col_raza, "")
                ws[f'AT{row_f}'] = df_fila.get(col_cabezas, "")

            # Números de reses en producción, terneros, novillos, etc.
            ws['AK44'] = df_fila.get('data-carne-reses_produccion_carne', "")
            ws['AK45'] = df_fila.get('data-carne-numero_terneros_carne', "")
            ws['AK46'] = df_fila.get('data-carne-numero_novillos_carne', "")
            ws['AK47'] = df_fila.get('data-carne-numero_novillas_carne', "")
            ws['AK48'] = df_fila.get('data-carne-numero_reproductores_carne', "")
            ws['AK49'] = df_fila.get('data-carne-partos_anio_carne', "")
            ws['AK50'] = df_fila.get('data-carne-vacas_ordeño_carne', "")
            ws['AK51'] = df_fila.get('data-carne-tiempo_venta_carne', "")
            ws['AK52'] = df_fila.get('data-carne-peso_promedio_venta_carne', "")

            # Precio promedio
            ws['AQ53'] = df_fila.get('data-carne-precio_promedio_carne', "")

            # Frecuencia (¿Cada cuánto produce? Mensual, Trimestral, etc.)
            frecuencia_carne = df_fila.get('data-carne-frecuencia_produccion_carne')
            if pd.notna(frecuencia_carne):
                if frecuencia_carne == 'mensual':
                    ws['AJ54'] = 'X'
                elif frecuencia_carne == 'trimestral':
                    ws['AO54'] = 'X'
                elif frecuencia_carne == 'otro' or 'other':
                    ws['AU54'] = 'X'

            # Promedio de producción (Kg, Cabezas)
            ws['AG55'] = df_fila.get('data-carne-promedio_produccion_carne', "")

            # Continuidad y Ampliación
        if df_fila['data-carne-destino_final_carne']:
            ws['L56'] = df_fila['data-carne-destino_final_carne'] 

        continuidad2 = df_fila['data-carne-begin_sobre_act3-continuar_actividad3']
        if pd.notna(continuidad2):
            if continuidad2 == 'yes':
                ws['G60'] = 'X'
                ws['AP60'] = 'X'
            elif continuidad2 == 'no':
                ws['I60'] = 'X'
                ws['AN60'] = 'X'  
        else:
            print('Campo vacío') 

        produccion2 = df_fila['data-carne-begin_sobre_act3-ampliar_produccion3']
        if pd.notna(produccion2):
            if produccion2 == 'yes':
                ws['Q60'] = 'X'
                ws['AF60'] = 'X'               
            elif produccion2 == 'no':
                ws['S60'] = 'X'
                ws['AD60'] = 'X'   
        else:
            print('Campo vacío') 

        ws['AS77'] = df_fila['data-carne-begin_sobre_act3-porque_ampliar3']

    #Raza

    if df_explot_porcina is not None:

        # 1) Mapeo para decidir qué columna de Excel usar según la 'categoría'
        map_categoria_a_columna = {
            'cria': 'T',
            'levante': 'AF',
            'ceba': 'AO'
        }

        # 2) Mapeo de cada campo a la fila donde se escribe.
        map_campo_a_fila = {
            # 31. Raza
            'data-explotacion_porcina-categoria_porquina-raza': 64,
            # 32. Número de Hembras
            'data-explotacion_porcina-categoria_porquina-numero_hembras': 65,
            # 33. Número de machos
            'data-explotacion_porcina-categoria_porquina-numero_machos': 66,
            # 34. ¿Tiene marranos para la venta?
            'data-explotacion_porcina-categoria_porquina-tiene_marranos_venta': 67,
            # 35. Peso promedio para la venta
            'data-explotacion_porcina-categoria_porquina-peso_promedio_venta': 68,
            # 36. Nº promedio de animales vendidos por año
            'data-explotacion_porcina-categoria_porquina-promedio_vendidos_anio': 69,
            # 37. Cantidad empleada para autoconsumo
            'data-explotacion_porcina-categoria_porquina-autoconsumo': 70,
            # 38. Precio unitario de venta
            'data-explotacion_porcina-categoria_porquina-precio_unitario': 71,
            # 39. Costo aproximado de producción
            'data-explotacion_porcina-categoria_porquina-costo_produccion': 72,
            # 40. ¿Destino final del producto?
            'data-explotacion_porcina-categoria_porquina-destino_producto': 73,
        }

        # 3) Iterar sobre cada fila de df_explot_porcina
        for idx, fila_porcina in df_explot_porcina.iterrows():
            # Tomar la categoría (cria/levante/ceba) para decidir columna
            categoria = str(fila_porcina.get(
                'data-explotacion_porcina-categoria_porquina-tipo_categoria',
                ''
            )).lower()

            # Buscar la columna en map_categoria_a_columna
            col_excel = map_categoria_a_columna.get(categoria)
            if not col_excel:
                # Si la categoría no coincide (o está vacía), omitir
                continue

            # 4) Asignar cada campo a su fila
            for campo, fila_excel in map_campo_a_fila.items():
                valor = fila_porcina.get(campo, "")
                celda = f"{col_excel}{fila_excel}"
                ws[celda] = valor


    continuidad3 = df_fila['data-explotacion_porcina-begin_sobre_act4-continuar_actividad4']
    if pd.notna(continuidad3):
        if continuidad3 == 'Continuar con la actividad':
            ws['G77'] = 'X'
            ws['AP77'] = 'X'
        elif continuidad3 == 'Finalizar la actividad':
            ws['I77'] = 'X'
            ws['AN77'] = 'X'  
    else:
        print('Campo vacío') 

    produccion3 = df_fila['data-explotacion_porcina-begin_sobre_act4-ampliar_produccion4']
    if pd.notna(produccion3):
        if produccion3 == 'Ampliar la producción':
            ws['Q77'] = 'X'
            ws['AF77'] = 'X'               
        elif produccion3 == 'Permanecer con la misma producción':
            ws['S77'] = 'X'
            ws['AD77'] = 'X'   
    else:
        print('Campo vacío') 

    ws['X74'] = df_fila['data-explotacion_porcina-begin_sobre_act4-porque_continuar4']

    # Explotación Avícola

    if df_explot_avicola is not None:

    # 1) Definir el mapeo de categoría => columnas “normales”
        map_categoria_a_col_normal = {
            'Engorde': 'N',
            'Ponedoras': 'AB',
            'Gallina campesina': 'AM'
        }
        # 2) Definir el mapeo de categoría => columnas “costo por animal”
        map_categoria_a_col_costo = {
            'Engorde': 'O',
            'Ponedoras': 'AC',
            'Gallina campesina': 'AN'
        }

        # 3) Definir mapeo de campo => fila (ajusta según tu plantilla)
        map_campo_a_fila = {
            # 43. Número de animales (fila 82)
            'data-explotacion_avicola-tipo_explotacion-numero_animales': 82,
            # 44. Producción mensual (fila 83)
            'data-explotacion_avicola-tipo_explotacion-produccion_mensual': 83,
            # 45. Unidades vendidas al mes (fila 84)
            'data-explotacion_avicola-tipo_explotacion-unidades_vendidas_mes': 84,
            # 46. Valor unitario de venta (fila 85)
            'data-explotacion_avicola-tipo_explotacion-valor_unitario_venta': 85,
            # 47. Costo por animal (fila 86) -> va a columnas O/AC/AN
            'data-explotacion_avicola-tipo_explotacion-costo_por_animal': 86,
            # 48. Cantidad empleada para autoconsumo/mes (fila 87)
            'data-explotacion_avicola-tipo_explotacion-autoconsumo_mes': 87,
            # 49. ¿Destino final del producto? (fila 88)
            'data-explotacion_avicola-tipo_explotacion-destino_producto': 88
        }

        # 4) Recorrer cada fila (categoría) del DF
        for idx, fila_avicola in df_explot_avicola.iterrows():
            # Tomamos la categoría
            categoria = str(
                fila_avicola.get('data-explotacion_avicola-tipo_explotacion-categoria_avicola', '')
            )
            # Solo procesamos si es Engorde / Ponedoras / Gallina campesina
            if categoria not in map_categoria_a_col_normal:
                # 'Cría' (u otra) se ignora
                continue
            col_normal = map_categoria_a_col_normal[categoria]
            # Columna “costo” (O/AC/AN)
            col_costo = map_categoria_a_col_costo[categoria]

            # 5) Marcar con “X” el tipo de explotación en la fila 80
            if categoria == 'Engorde':
                ws['N80'] = 'X'   # o donde se marque Engorde
            elif categoria == 'Ponedoras':
                ws['AB80'] = 'X'  # …
            elif categoria == 'Gallina campesina':
                ws['AM80'] = 'X'

            # 6) Rellenar valores en filas 82..88
            for campo_df, fila_excel in map_campo_a_fila.items():
                valor = fila_avicola.get(campo_df, "")

                if campo_df.endswith('-costo_por_animal'):
                    # Caso especial => usar col_costo (O/AC/AN)
                    celda = f"{col_costo}{fila_excel}"
                else:
                    # Caso normal => usar col_normal (N/AB/AM)
                    celda = f"{col_normal}{fila_excel}"

                ws[celda] = valor
                

    continuidad3 = df_fila['data-explotacion_avicola-sobre_la_actividad5-continuar_actividad5']
    if pd.notna(continuidad3):
        if continuidad3 == 'Continuar con la actividad':
            ws['G92'] = 'X'
            ws['AP92'] = 'X'
        elif continuidad3 == 'Finalizar la actividad':
            ws['I92'] = 'X'
            ws['AN92'] = 'X'  
    else:
        print('Campo vacío') 

    produccion3 = df_fila['data-explotacion_avicola-sobre_la_actividad5-ampliar_produccion5']
    if pd.notna(produccion3):
        if produccion3 == 'Ampliar la producción':
            ws['Q92'] = 'X'
            ws['AF92'] = 'X'               
        elif produccion3 == 'Permanecer con la misma producción':
            ws['S92'] = 'X'
            ws['AD92'] = 'X'   
    else:
        print('Campo vacío') 

    ws['Y89'] = df_fila['data-explotacion_avicola-sobre_la_actividad5-razon_ampliar5']

    # INFORMACIÓN COMERCIAL
    # Asegurarte de no pasarte de 13 filas (o las que permita tu plantilla).
    df_info_comercial = df_info_comercial.head(13).reset_index(drop=True)

    for i, fila_insumo in df_info_comercial.iterrows():
        # La fila en Excel donde irá este insumo
        excel_row = 99 + i

        # 1) Nombre del insumo => columna F
        nombre_insumo = fila_insumo.get('data-informacion_comercial-insumos_actividad-nombre_insumo', "")
        ws[f'F{excel_row}'] = nombre_insumo

        # 2) Unidad de medida => marcar X en M, Q, S, U, W dependiendo
        unidad_medida = fila_insumo.get('data-informacion_comercial-insumos_actividad-unidad_medida', "")
        if pd.notna(unidad_medida):
            if unidad_medida == 'Kg':
                ws[f'M{excel_row}'] = 'X'
            elif unidad_medida == 'Bulto':
                ws[f'Q{excel_row}'] = 'X'
            elif unidad_medida == '@':
                ws[f'S{excel_row}'] = 'X'
            elif unidad_medida == 'Lt':
                ws[f'U{excel_row}'] = 'X'
            elif unidad_medida == 'Gn':
                ws[f'W{excel_row}'] = 'X'
        # Si hubiera un "unidad_medida_other", podrías ponerlo aparte

        # 3) Cantidad de insumo => columna X
        cantidad_insumo = fila_insumo.get('data-informacion_comercial-insumos_actividad-cantidad_insumo', "")
        ws[f'X{excel_row}'] = cantidad_insumo

        # 4) Frecuencia de compra => marcar X en AC, AE, AG, AJ, AL, AN
        frecuencia = fila_insumo.get('data-informacion_comercial-insumos_actividad-frecuencia_compra', "")
        if pd.notna(frecuencia):
            if frecuencia == 'unica':
                ws[f'AC{excel_row}'] = 'X'
            elif frecuencia == 'diario':
                ws[f'AE{excel_row}'] = 'X'
            elif frecuencia == 'semanal':
                ws[f'AG{excel_row}'] = 'X'
            elif frecuencia == 'quincenal':
                ws[f'AJ{excel_row}'] = 'X'
            elif frecuencia == 'mensual':
                ws[f'AL{excel_row}'] = 'X'
            elif frecuencia == 'trimestral':
                ws[f'AN{excel_row}'] = 'X'

        # 5) Precio de compra/unidad => columna AQ
        precio_compra = fila_insumo.get('data-informacion_comercial-insumos_actividad-precio_compra', "")
        ws[f'AQ{excel_row}'] = precio_compra

        # 6) Lugar de compra => columna AT
        lugar_compra = fila_insumo.get('data-informacion_comercial-insumos_actividad-lugar_compra', "")
        ws[f'AT{excel_row}'] = lugar_compra
    
    ws['L112'] = df_fila['data-informacion_comercial-almacenamiento_productos']

    ws['L113'] = df_fila['data-informacion_comercial-procedencia_compradores']

    agua_fuente = df_fila['data-informacion_comercial-recurso_hidrico']
    if pd.notna(agua_fuente):
        if agua_fuente == 'aljibe':
            ws['P114'] = 'X'
        elif agua_fuente == 'acueducto':
            ws['AC114'] = 'X'
        elif agua_fuente == 'otro':
            ws['AJ114'] = 'X'
            ws['AP114'] = df_fila['data-informacion_comercial-recurso_hidrico_otro']    

    # ws['T115'] = df_fila['Forma de extracción']

    # ws['AN115'] = df_fila['Cantidad estimada (m3)']

    alcantarillado = df_fila['data-informacion_comercial-cuenta_alcantarillado']
    if pd.notna(alcantarillado):
        if alcantarillado == 'yes':
            ws['U116'] = 'X'
            ws['AM116'] = df_fila['data-informacion_comercial-cual_alcantarillado']  
        elif alcantarillado == 'no':
            ws['AC116'] = 'X'

    mapa_energia = {
        'electrica': 'T117',
        'solar': 'AI117',
        'otro': 'AP117'
    }
    energia = df_fila['data-informacion_comercial-energia_utiliza'].split(', ')
    for tipo_energia in energia:
        if tipo_energia.strip() in mapa_energia and tipo_energia != 'otro':
            ws[mapa_energia.get(tipo_energia, '')] = 'X'
        else:
            ws['AP117'] = df_fila['data-informacion_comercial-energia_otro']


    energia_coccion = df_fila['data-informacion_comercial-energia_coccion']
    if energia_coccion == 'electrica':
        ws['AB118'] = 'X'
    elif energia_coccion == 'lenia':
        ws['AG118'] = 'X'
    elif energia_coccion == 'gas':
        ws['AK118'] = 'X'
    elif energia_coccion == 'other':       
        ws['AQ118'] = df_fila['data-informacion_comercial-energia_coccion_other']

    ws['L119'] = df_fila['data-informacion_comercial-manejo_aguas']

    mano_obra = df_fila['data-informacion_comercial-contrata_mano_obra']
    if mano_obra == 'yes':
        ws['U120'] = 'X'
    elif mano_obra == 'no':
        ws['AC120'] = 'X'  

    # E. Información Laboral 

    df_info_laboral = df_info_laboral.head(10).reset_index(drop=True)

    for i, fila_lab in df_info_laboral.iterrows():
        excel_row = 124 + i  # Filas 124..133

        # 1) Cargo => Columna E
        cargo = fila_lab.get('data-informacion_laboral-cargo', "")
        ws[f'E{excel_row}'] = cargo

        # 2) Edad => Columna J
        edad = fila_lab.get('data-informacion_laboral-edad', "")
        ws[f'J{excel_row}'] = edad

        # 3) Duración jornada => Columna K
        dur_jornada = fila_lab.get('data-informacion_laboral-duracion_jornada', "")
        ws[f'K{excel_row}'] = dur_jornada

        # 4) Tipo de mano de obra => “familiar” => B, “contratado” => D
        tipo_mano_obra = fila_lab.get('data-informacion_laboral-tipo_mano_obra', "")
        if pd.notna(tipo_mano_obra):
            if tipo_mano_obra.lower() == 'familiar':
                ws[f'B{excel_row}'] = 'X'
            elif tipo_mano_obra.lower() == 'contratado':
                ws[f'D{excel_row}'] = 'X'

        # 5) Género => “M” => I, “F” => G
        genero = fila_lab.get('data-informacion_laboral-genero', "")
        if pd.notna(genero):
            if genero.upper() == 'M':
                ws[f'I{excel_row}'] = 'X'
            elif genero.upper() == 'F':
                ws[f'G{excel_row}'] = 'X'

        # 6) Escolaridad => “primaria” => M, “bachillerato” => O, “tecnico” => Q, 
        #    “pregrado” => S, “posgrado” => U
        esc = fila_lab.get('data-informacion_laboral-escolaridad', "")
        if pd.notna(esc):
            esc_lower = esc.lower()
            if esc_lower == 'primaria':
                ws[f'M{excel_row}'] = 'X'
            elif esc_lower == 'bachillerato':
                ws[f'O{excel_row}'] = 'X'
            elif esc_lower == 'tecnico':
                ws[f'Q{excel_row}'] = 'X'
            elif esc_lower == 'pregrado':
                ws[f'S{excel_row}'] = 'X'
            elif esc_lower == 'posgrado':
                ws[f'U{excel_row}'] = 'X'

        # 7) Contrato => “temporal” => AA, “fijo” => AC
        contrato = fila_lab.get('data-informacion_laboral-contrato', "")
        if pd.notna(contrato):
            contrato_lower = contrato.lower()
            if contrato_lower == 'temporal':
                # En tu código anterior parecía que escribías el valor, 
                # pero si sólo quieres marcar X, haz:
                ws[f'AA{excel_row}'] = 'X'
            elif contrato_lower == 'fijo':
                ws[f'AC{excel_row}'] = 'X'

        # 8) Pago de seguridad social => “yes” => AE, “no” => AG
        pago_seg = fila_lab.get('data-informacion_laboral-pago_seguridad_social', "")
        if pd.notna(pago_seg):
            pago_seg_lower = pago_seg.lower()
            if pago_seg_lower == 'yes':
                ws[f'AE{excel_row}'] = 'X'
            elif pago_seg_lower == 'no':
                ws[f'AG{excel_row}'] = 'X'

        # 9) Remuneración => “inferior_smlv” => AS, “igual_smlv” => AT, 
        #    “entre_1y2_smlv” => AU, “superior_2” => AV
        remuneracion = fila_lab.get('data-informacion_laboral-remuneracion', "")
        if pd.notna(remuneracion):
            if remuneracion == 'inferior_smlv':
                ws[f'AS{excel_row}'] = 'X'
            elif remuneracion == 'igual_smlv':
                ws[f'AT{excel_row}'] = 'X'
            elif remuneracion == 'entre_1y2_smlv':
                ws[f'AU{excel_row}'] = 'X'
            elif remuneracion == 'superior_2':
                ws[f'AV{excel_row}'] = 'X'

        # 10) Otros campos
        procedencia = fila_lab.get('data-informacion_laboral-procedencia', "")
        ws[f'AH{excel_row}'] = procedencia

        residencia = fila_lab.get('data-informacion_laboral-residencia', "")
        ws[f'AJ{excel_row}'] = residencia

        tiempo_trab = fila_lab.get('data-informacion_laboral-tiempo_trabajado', "")
        ws[f'AL{excel_row}'] = tiempo_trab

        nucleo_familiar = fila_lab.get('data-informacion_laboral-nucleo_familiar', "")
        ws[f'AM{excel_row}'] = nucleo_familiar

        personas_a_cargo = fila_lab.get('data-informacion_laboral-personas_a_cargo', "")
        ws[f'AN{excel_row}'] = personas_a_cargo

        lugar_res_fam = fila_lab.get('data-informacion_laboral-lugar_residencia_familiar', "")
        ws[f'AO{excel_row}'] = lugar_res_fam

        # ¿Contrata persona por jornal? 
    df_detalle_jornal = df_detalle_jornal.head(3).reset_index(drop=True)

    for i, fila_jornal in df_detalle_jornal.iterrows():
        # Fila en Excel donde pondremos la info
        excel_row = 138 + i

        # 1) Tipo de obra o labor => columna A
        tipo_obra = fila_jornal.get('data-detalle_jornal-jornal_detalle-tipo_obra_labor', "")
        ws[f'A{excel_row}'] = tipo_obra

        # 2) Frecuencia de contratación/año => columna G
        frecuencia = fila_jornal.get('data-detalle_jornal-jornal_detalle-frecuencia_contratacion', "")
        ws[f'G{excel_row}'] = frecuencia

        # 3) Duración en jornales del contrato => columna Q
        duracion = fila_jornal.get('data-detalle_jornal-jornal_detalle-duracion_jornales', "")
        ws[f'Q{excel_row}'] = duracion

        # 4) Valor del jornal => columna Y
        valor_jornal = fila_jornal.get('data-detalle_jornal-jornal_detalle-valor_jornal', "")
        ws[f'Y{excel_row}'] = valor_jornal

        # 5) Cantidad de jornaleros empleados => columna AG
        cant_jornaleros = fila_jornal.get('data-detalle_jornal-jornal_detalle-cantidad_jornaleros', "")
        ws[f'AG{excel_row}'] = cant_jornaleros

        # 6) Residencia de los jornaleros => columna AN
        residencia = fila_jornal.get('data-detalle_jornal-jornal_detalle-residencia_jornaleros', "")
        ws[f'AN{excel_row}'] = residencia

def llenarFormatoComercial(ws, df_fila):
        ws['AI1'] = df_fila["Encuesta No."]

        fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
        if '/' in fecha_str:
            ws['AK2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
            ws['AN2'] = fecha_str.split('/')[1]
            ws['AS2'] = fecha_str.split('/')[0]
        elif '-' in fecha_str:
            ws['AK2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
            ws['AN2'] = fecha_str.split('-')[1]
            ws['AS2'] = fecha_str.split('-')[0] ####### LLENAR FECHA EN ESPACIOS VACÍOS Y NO SOBRE EL SÍMBOLO.
        else:
            print(f'Formato de fecha inesperado: {fecha_str}')

        ws['AK3'] = df_fila["Encuestador"]
        ws['F6'] = df_fila["Nombre"]
        ws['AC6'] = df_fila["Empresa"]
        ws['AP6'] = df_fila["Cargo"]

        if df_fila["¿Pertenece a alguna asociación?"] == 'Si':
            ws['S7'] = "X"
            ws['AF7'] = df_fila["Otro, ¿Cuál?"]

        elif df_fila["¿Pertenece a alguna asociación?"] == 'No':
            ws['Y7'] = "X"

        actividad = df_fila['¿Qué tipo de productos comercializa?']
        if actividad == 'Agrícola':
            ws['L11'] = 'X'
        elif actividad == 'Pecuario':
            ws['L12'] = 'X'
        elif actividad == 'Víveres':
            ws['L13'] = 'X'
        elif actividad == 'Agua en botella/bolsa':
            ws['L14'] = 'X'
        elif actividad == 'Licores':
            ws['U11'] = 'X'
        elif actividad == 'Productos Naturales: Animal':
            ws['U13'] = 'X'
        elif actividad == 'Productos Naturales: Vegetal':
            ws['Y13'] = 'X'
        elif actividad == "Otros":
            ws['T14'] = df_fila['Otro, ¿Cuáles?']

        ws['C16'] = df_fila["¿Cuál es el principal producto que comercializa?"]

        actividad2 = df_fila['¿Con qué frecuencia compra los productos que comercializa?']
        if actividad2 == 'Diario':
            ws['L20'] = 'X'
        elif actividad2 == 'Semanal':
            ws['L21'] = 'X'
        elif actividad2 == 'Quincenal':
            ws['L22'] = 'X'
        elif actividad2 == 'Mensual':
            ws['L23'] = 'X'
        elif actividad2 == 'Semestral':
            ws['V22'] = actividad2
        elif actividad2 == 'Otra':
            ws['V22'] = df_fila["Otro"]

        ws['C28'] = df_fila["Producto"]
        ws['I28'] = df_fila["Cantidad"]
        ws['M28'] = df_fila["Unidad de medida"]
        ws['S28'] = df_fila["Valor"]
        ws['C29'] = df_fila["Producto.1"]
        ws['I29'] = df_fila["Cantidad.1"]
        ws['M29'] = df_fila["Unidad de medida.1"]
        ws['S29'] = df_fila["Valor.1"]
        ws['C30'] = df_fila["Producto.2"]
        ws['I30'] = df_fila["Cantidad.2"]
        ws['M30'] = df_fila["Unidad de medida.2"]
        ws['S30'] = df_fila["Valor.2"]
        ws['C31'] = df_fila["Producto.3"]
        ws['I31'] = df_fila["Cantidad.3"]
        ws['M31'] = df_fila["Unidad de medida.3"]
        ws['S31'] = df_fila["Valor.3"]

        ws['D34'] = df_fila["Producto.4"]
        ws['R34'] = df_fila["Precio"]
        ws['D35'] = df_fila["Producto.5"]
        ws['R35'] = df_fila["Precio.1"]
        ws['D36'] = df_fila["Producto.6"]
        ws['R36'] = df_fila["Precio.2"]

        actividad3 = df_fila["Señale el tipo de emplazamiento"]
        if actividad3 == 'Local':
            ws['Q39'] = 'X'
        elif actividad3 == 'Puesto Fijo':
            ws['Q40'] = 'X'
        elif actividad3 == 'Vivienda económica':
            ws['Q41'] = 'X'
        elif actividad3 == 'Venta ambulante':
            ws['Q42'] = 'X'

        actividad4 = df_fila["¿Cuál fue el valor promedio de ventas en el último mes?"]
        if actividad4 == 'Inferiores a $600.000':
            ws['AN10'] = 'X'
        elif actividad4 == 'Entre $ 601.000 y $ 1.500.000':
            ws['AN11'] = 'X'
        elif actividad4 == 'Entre $ 1.501.000 y $ 3.000.000':
            ws['AN12'] = 'X'
        elif actividad4 == 'Superior a $ 3.000.000':
            ws['AN13'] = 'X'
            ws['AP13'] = df_fila['Si fue superior a 3 millones, ¿cuánto fue?']


        actividad5 = df_fila["Vende principalmente en:"]
        if actividad5 == 'Sitio':
            ws['AH16'] = 'X'
        elif actividad5 == 'Vereda':
            ws['AH17'] = 'X'
        elif actividad5 == 'Casco Urbano':
            ws['AH18'] = 'X'
        elif actividad5 == 'Otros Municipios y/o Veredas':
            ws['AN16'] = 'X'
            ws['AO18'] = df_fila["¿Cuáles?"]

        if df_fila["¿Lleva libros contables del establecimiento?"] == 'Si':
            ws['AO22'] = 'X'

        elif df_fila["¿Lleva libros contables del establecimiento?"] == 'No':
            ws['AQ22'] = 'X'

        if str(df_fila["Producto 1"]) != "nan":
            p1 = df_fila["Producto 1"]
        else:
            p1 = " "
        if str(df_fila["Producto 2"]) != "nan":
            p2 = df_fila["Producto 2"]
        else:
            p2 = " "
        if str(df_fila["Producto 3"]) != "nan":
            p3 = df_fila["Producto 3"]
        else:
            p3 = " "
        if str(df_fila["Producto 4"]) != "nan":
            p4 = df_fila["Producto 4"]
        else:
            p4 = " "
        if str(df_fila["Producto 5"]) != "nan":
            p5 = df_fila["Producto 5"]
        else:
            p5 = " "
            3
        ws['AE26'] = str(p1) + " " + str(p2) + " " + str(p3) + " " + str(p4) + " " + str(p5)

        if str(df_fila["Hidrocarburos"]) != "nan":
            ws['AH29'] = 'X'
            ws['AM29'] = df_fila['Hidrocarburos']
        if str(df_fila["Otro.1"]) != "nan":
            ws['AH30'] = 'X'
            ws['AM30'] = df_fila['Otro.1']

        actividad5 = df_fila['¿Con qué frecuencia compra productos en otras veredas y/o municipios?']
        if actividad5 == 'Diario':
            ws['AJ33'] = 'X'
        elif actividad5 == 'Semanal':
            ws['AJ34'] = 'X'
        elif actividad5 == 'Quincenal':
            ws['AJ35'] = 'X'
        elif actividad5 == 'Mensual':
            ws['AJ36'] = 'X'
        elif actividad5 == 'Otro':
            ws['AQ33'] = 'X'
            ws['AP34'] = df_fila["Otro, ¿Cuál?.1"]

        if df_fila["Sobre la actividad, piensa: Continuidad"] == "Continuar con la actividad":
            ws['AN39'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Continuidad"] == "Finalizar la actividad":
            ws['AP39'] = 'X'
        if df_fila["Sobre la actividad, piensa: Producción"] == "Ampliar la producción":
            ws['AP40'] = 'X'
            ws['AN41'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Permanecer con la misma producción":
            ws['AN40'] = 'X'
            ws['AP41'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Ninguna de las anteriores":
            ws['AP40'] = 'X'
            ws['AP41'] = 'X'


        actividad6 = df_fila['¿De dónde se abastece del recurso hídrico?']
        if actividad6 == 'Aljibe':
            ws['I45'] = 'X'
        elif actividad6 == 'Acueducto Veredal':
            ws['I46'] = 'X'
        elif actividad6 == 'Otro':
            ws['I47'] = 'X'
            ws['R47'] = df_fila['Otro, ¿Cuál?.2']
    
        ws['U45'] = df_fila["Forma de extracción"]
        ws['P46'] = df_fila["Cantidad estimada (escribir m3)"]

        if df_fila["¿Cuenta con servicio de alcantarillado?"] == "Si":
            ws['AP44'] = 'X'
            ws['AF45'] = df_fila["¿Cuál?"]
        elif df_fila["¿Cuenta con servicio de alcantarillado?"] == "No":
            ws['AR44'] = 'X'


        actividad7 = df_fila['¿Qué tipo de energía utiliza?']
        if actividad7 == 'Energía Eléctrica':
            ws['AG47'] = 'X'
        elif actividad7 == 'Energía Solar':
            ws['AN47'] = 'X'
        elif actividad7 == 'Otro':
            ws['AS47'] = df_fila['Otro, ¿Cuál?.3']

        actividad8 = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
        if actividad8 == 'Energía eléctrica':
            ws['AG48'] = 'X'
        elif actividad8 == 'Leña':
            ws['AM48'] = 'X'
        elif actividad8 == 'Gas':
            ws['AO48'] = 'X'
        elif actividad8 == 'Otro':
            ws['AS48'] = df_fila['Otro, ¿Cuál?.4']

        if df_fila["Contrata algún tipo de mano de obra"] == "Si":
            ws['Z50'] = 'X'

            #### Persona 1 ####

            if df_fila["Tipo de mano de obra"] == "Familiar":
                ws['B54'] = 'X'
            elif df_fila["Tipo de mano de obra"] == "Contratado":
                ws['D54'] = 'X'

            ws['E54'] = df_fila["Cargo.1"]

            if df_fila["Género"] == "Masculino":
                ws['M54'] = 'X'
            elif df_fila["Género"] == "Femenino":
                ws['K54'] = 'X'

            ws['N54'] = df_fila["Edad (años)"]
            ws['Q54'] = df_fila["Duración jornada (horas)"]

            actividad9 = df_fila['Escolaridad']
            if actividad9 == 'Primaria':
                ws['S54'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U54'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W54'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y54'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA54'] = 'X'

            if df_fila['Contrato'] == 'Tem.':
                ws['AC54'] = 'X'
            elif df_fila['Contrato'] == 'Fij':
                ws['AG54'] = 'X'
            
            if df_fila['Pago de seguridad'] == 'Si':
                ws['AH54'] = 'Si'
                ws['AJ54'] = ''
            elif df_fila['Pago de seguridad'] == 'No':
                ws['AH54'] = ''
                ws['AJ54'] = 'No'

            ws['AL54'] = df_fila["Procedencia"]
            ws['AM54'] = df_fila["Residencia"]
            ws['AN54'] = df_fila["Tiempo trabajado"]
            ws['AO54'] = df_fila["# Personas núcleo familiar"]
            ws['AP54'] = df_fila["Personas a cargo"]
            ws['AQ54'] = df_fila["Lugar de residencia familiar"]
        
            actividad10 = df_fila['Remuneración']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT54'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU54'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV54'] = 'X'
            elif actividad10 == 'Superiores a $2.701.000':
                ws['AW54'] = 'X'

            ##### Persona 2 #####

            if df_fila["Tipo de mano de obra.1"] == "Familiar":
                ws['B55'] = 'X'
            elif df_fila["Tipo de mano de obra.1"] == "Contratado":
                ws['D55'] = 'X'

            ws['E55'] = df_fila["Cargo.2"]

            if df_fila["Género.1"] == "Masculino":
                ws['M55'] = 'X'
            elif df_fila["Género.1"] == "Femenino":
                ws['K55'] = 'X'

            ws['N55'] = df_fila["Edad (años).1"]
            ws['Q55'] = df_fila["Duración jornada (horas).1"]

            actividad9 = df_fila['Escolaridad.1']
            if actividad9 == 'Primaria':
                ws['S55'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U55'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W55'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y55'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA55'] = 'X'

            if df_fila['Contrato.1'] == 'Tem.':
                ws['AC55'] = 'X'
            elif df_fila['Contrato.1'] == 'Fij':
                ws['AG55'] = 'X'
            
            if df_fila['Pago de seguridad.1'] == 'Si':
                ws['AH55'] = 'Si'
                ws['AJ55'] = ''
            elif df_fila['Pago de seguridad.1'] == 'No':
                ws['AH55'] = ''
                ws['AJ55'] = 'No'

            ws['AL55'] = df_fila["Procedencia.1"]
            ws['AM55'] = df_fila["Residencia.1"]
            ws['AN55'] = df_fila["Tiempo trabajado.1"]
            ws['AO55'] = df_fila["# Personas núcleo familiar.1"]
            ws['AP55'] = df_fila["Personas a cargo.1"]
            ws['AQ55'] = df_fila["Lugar de residencia familiar.1"]
        
            actividad10 = df_fila['Remuneración.1']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT55'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU55'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV55'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW55'] = 'X'

            ##### Persona 3 #####

            if df_fila["Tipo de mano de obra.2"] == "Familiar":
                ws['B56'] = 'X'
            elif df_fila["Tipo de mano de obra.2"] == "Contratado":
                ws['D56'] = 'X'

            ws['E56'] = df_fila["Cargo.3"]

            if df_fila["Género.2"] == "Masculino":
                ws['M56'] = 'X'
            elif df_fila["Género.2"] == "Femenino":
                ws['K56'] = 'X'

            ws['N56'] = df_fila["Edad (años).2"]
            ws['Q56'] = df_fila["Duración jornada (horas).2"]

            actividad9 = df_fila['Escolaridad.2']
            if actividad9 == 'Primaria':
                ws['S56'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U56'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W56'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y56'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA56'] = 'X'

            if df_fila['Contrato.2'] == 'Tem.':
                ws['AC56'] = 'X'
            elif df_fila['Contrato.2'] == 'Fij':
                ws['AG56'] = 'X'
            
            if df_fila['Pago de seguridad.2'] == 'Si':
                ws['AH56'] = 'Si'
                ws['AJ56'] = ''
            elif df_fila['Pago de seguridad.2'] == 'No':
                ws['AH56'] = ''
                ws['AJ56'] = 'No'

            ws['AL56'] = df_fila["Procedencia.2"]
            ws['AM56'] = df_fila["Residencia.2"]
            ws['AN56'] = df_fila["Tiempo trabajado.2"]
            ws['AO56'] = df_fila["# Personas núcleo familiar.2"]
            ws['AP56'] = df_fila["Personas a cargo.2"]
            ws['AQ56'] = df_fila["Lugar de residencia familiar.2"]
        
            actividad10 = df_fila['Remuneración.2']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT56'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU56'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV56'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW56'] = 'X'

            ##### Persona 4 #####

            if df_fila["Tipo de mano de obra.3"] == "Familiar":
                ws['B57'] = 'X'
            elif df_fila["Tipo de mano de obra.3"] == "Contratado":
                ws['D57'] = 'X'

            ws['E57'] = df_fila["Cargo.4"]

            if df_fila["Género.3"] == "Masculino":
                ws['M57'] = 'X'
            elif df_fila["Género.3"] == "Femenino":
                ws['K57'] = 'X'

            ws['N57'] = df_fila["Edad (años).3"]
            ws['Q57'] = df_fila["Duración jornada (horas).3"]

            actividad9 = df_fila['Escolaridad.3']
            if actividad9 == 'Primaria':
                ws['S57'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U57'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W57'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y57'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA57'] = 'X'

            if df_fila['Contrato.3'] == 'Tem.':
                ws['AC57'] = 'X'
            elif df_fila['Contrato.3'] == 'Fij':
                ws['AG57'] = 'X'
            
            if df_fila['Pago de seguridad.3'] == 'Si':
                ws['AH57'] = 'Si'
                ws['AJ57'] = ''
            elif df_fila['Pago de seguridad.3'] == 'No':
                ws['AH57'] = ''
                ws['AJ57'] = 'No'

            ws['AL57'] = df_fila["Procedencia.3"]
            ws['AM57'] = df_fila["Residencia.3"]
            ws['AN57'] = df_fila["Tiempo trabajado.3"]
            ws['AO57'] = df_fila["# Personas núcleo familiar.3"]
            ws['AP57'] = df_fila["Personas a cargo.3"]
            ws['AQ57'] = df_fila["Lugar de residencia familiar.3"]
        
            actividad10 = df_fila['Remuneración.3']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT57'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU57'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV57'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW57'] = 'X'

            ##### Persona 5 #####

            if df_fila["Tipo de mano de obra.4"] == "Familiar":
                ws['B58'] = 'X'
            elif df_fila["Tipo de mano de obra.4"] == "Contratado":
                ws['D58'] = 'X'

            ws['E58'] = df_fila["Cargo.5"]

            if df_fila["Género.4"] == "Masculino":
                ws['M58'] = 'X'
            elif df_fila["Género.4"] == "Femenino":
                ws['K58'] = 'X'

            ws['N58'] = df_fila["Edad (años).4"]
            ws['Q58'] = df_fila["Duración jornada (horas).4"]

            actividad9 = df_fila['Escolaridad.4']
            if actividad9 == 'Primaria':
                ws['S58'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U58'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W58'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y58'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA58'] = 'X'

            if df_fila['Contrato.4'] == 'Tem.':
                ws['AC58'] = 'X'
            elif df_fila['Contrato.4'] == 'Fij':
                ws['AG58'] = 'X'
            
            if df_fila['Pago de seguridad.4'] == 'Si':
                ws['AH58'] = 'Si'
                ws['AJ58'] = ''
            elif df_fila['Pago de seguridad.4'] == 'No':
                ws['AH58'] = ''
                ws['AJ58'] = 'No'

            ws['AL58'] = df_fila["Procedencia.4"]
            ws['AM58'] = df_fila["Residencia.4"]
            ws['AN58'] = df_fila["Tiempo trabajado.4"]
            ws['AO58'] = df_fila["# Personas núcleo familiar.4"]
            ws['AP58'] = df_fila["Personas a cargo.4"]
            ws['AQ58'] = df_fila["Lugar de residencia familiar.4"]
        
            actividad10 = df_fila['Remuneración.4']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT58'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU58'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV58'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW58'] = 'X'

            ##### Persona 6 #####

            if df_fila["Tipo de mano de obra.5"] == "Familiar":
                ws['B59'] = 'X'
            elif df_fila["Tipo de mano de obra.5"] == "Contratado":
                ws['D59'] = 'X'

            ws['E59'] = df_fila["Cargo.6"]

            if df_fila["Género.5"] == "Masculino":
                ws['M59'] = 'X'
            elif df_fila["Género.5"] == "Femenino":
                ws['K59'] = 'X'

            ws['N59'] = df_fila["Edad (años).5"]
            ws['Q59'] = df_fila["Duración jornada (horas).5"]

            actividad9 = df_fila['Escolaridad.5']
            if actividad9 == 'Primaria':
                ws['S59'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U59'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W59'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y59'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA59'] = 'X'

            if df_fila['Contrato.5'] == 'Tem.':
                ws['AC59'] = 'X'
            elif df_fila['Contrato.5'] == 'Fij':
                ws['AG59'] = 'X'
            
            if df_fila['Pago de seguridad.5'] == 'Si':
                ws['AH59'] = 'Si'
                ws['AJ59'] = ''
            elif df_fila['Pago de seguridad.5'] == 'No':
                ws['AH59'] = ''
                ws['AJ59'] = 'No'

            ws['AL59'] = df_fila["Procedencia.5"]
            ws['AM59'] = df_fila["Residencia.5"]
            ws['AN59'] = df_fila["Tiempo trabajado.5"]
            ws['AO59'] = df_fila["# Personas núcleo familiar.5"]
            ws['AP59'] = df_fila["Personas a cargo.5"]
            ws['AQ59'] = df_fila["Lugar de residencia familiar.5"]
        
            actividad10 = df_fila['Remuneración.5']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT59'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU59'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV59'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW59'] = 'X'

            ##### Persona 7 #####

            if df_fila["Tipo de mano de obra.6"] == "Familiar":
                ws['B60'] = 'X'
            elif df_fila["Tipo de mano de obra.6"] == "Contratado":
                ws['D60'] = 'X'

            ws['E60'] = df_fila["Cargo.7"]

            if df_fila["Género.6"] == "Masculino":
                ws['M60'] = 'X'
            elif df_fila["Género.6"] == "Femenino":
                ws['K60'] = 'X'

            ws['N60'] = df_fila["Edad (años).6"]
            ws['Q60'] = df_fila["Duración jornada (horas).6"]

            actividad9 = df_fila['Escolaridad.6']
            if actividad9 == 'Primaria':
                ws['S60'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U60'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W60'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y60'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA60'] = 'X'

            if df_fila['Contrato.6'] == 'Tem.':
                ws['AC60'] = 'X'
            elif df_fila['Contrato.6'] == 'Fij':
                ws['AG60'] = 'X'
            
            if df_fila['Pago de seguridad.6'] == 'Si':
                ws['AH60'] = 'Si'
                ws['AJ60'] = ''
            elif df_fila['Pago de seguridad.6'] == 'No':
                ws['AH60'] = ''
                ws['AJ60'] = 'No'

            ws['AL60'] = df_fila["Procedencia.6"]
            ws['AM60'] = df_fila["Residencia.6"]
            ws['AN60'] = df_fila["Tiempo trabajado.6"]
            ws['AO60'] = df_fila["# Personas núcleo familiar.6"]
            ws['AP60'] = df_fila["Personas a cargo.6"]
            ws['AQ60'] = df_fila["Lugar de residencia familiar.6"]
        
            actividad10 = df_fila['Remuneración.6']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT60'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU60'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV60'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW60'] = 'X'

            ##### Persona 8 #####

            if df_fila["Tipo de mano de obra.7"] == "Familiar":
                ws['B61'] = 'X'
            elif df_fila["Tipo de mano de obra.7"] == "Contratado":
                ws['D61'] = 'X'

            ws['E61'] = df_fila["Cargo.8"]

            if df_fila["Género.7"] == "Masculino":
                ws['M61'] = 'X'
            elif df_fila["Género.7"] == "Femenino":
                ws['K61'] = 'X'

            ws['N61'] = df_fila["Edad (años).7"]
            ws['Q61'] = df_fila["Duración jornada (horas).7"]

            actividad9 = df_fila['Escolaridad.7']
            if actividad9 == 'Primaria':
                ws['S61'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U61'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W61'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y61'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA61'] = 'X'

            if df_fila['Contrato.7'] == 'Tem.':
                ws['AC61'] = 'X'
            elif df_fila['Contrato.7'] == 'Fij':
                ws['AG61'] = 'X'
            
            if df_fila['Pago de seguridad.7'] == 'Si':
                ws['AH61'] = 'Si'
                ws['AJ61'] = ''
            elif df_fila['Pago de seguridad.7'] == 'No':
                ws['AH61'] = ''
                ws['AJ61'] = 'No'

            ws['AL61'] = df_fila["Procedencia.7"]
            ws['AM61'] = df_fila["Residencia.7"]
            ws['AN61'] = df_fila["Tiempo trabajado.7"]
            ws['AO61'] = df_fila["# Personas núcleo familiar.7"]
            ws['AP61'] = df_fila["Personas a cargo.7"]
            ws['AQ61'] = df_fila["Lugar de residencia familiar.7"]
        
            actividad10 = df_fila['Remuneración.7']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT61'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU61'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV61'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW61'] = 'X'

            ##### Persona 9 #####

            if df_fila["Tipo de mano de obra.8"] == "Familiar":
                ws['B62'] = 'X'
            elif df_fila["Tipo de mano de obra.8"] == "Contratado":
                ws['D62'] = 'X'

            ws['E62'] = df_fila["Cargo.9"]

            if df_fila["Género.8"] == "Masculino":
                ws['M62'] = 'X'
            elif df_fila["Género.8"] == "Femenino":
                ws['K62'] = 'X'

            ws['N62'] = df_fila["Edad (años).8"]
            ws['Q62'] = df_fila["Duración jornada (horas).8"]

            actividad9 = df_fila['Escolaridad.8']
            if actividad9 == 'Primaria':
                ws['S62'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U62'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W62'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y62'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA62'] = 'X'

            if df_fila['Contrato.8'] == 'Tem.':
                ws['AC62'] = 'X'
            elif df_fila['Contrato.8'] == 'Fij':
                ws['AG62'] = 'X'
            
            if df_fila['Pago de seguridad.8'] == 'Si':
                ws['AH62'] = 'Si'
                ws['AJ62'] = ''
            elif df_fila['Pago de seguridad.8'] == 'No':
                ws['AH62'] = ''
                ws['AJ62'] = 'No'

            ws['AL62'] = df_fila["Procedencia.8"]
            ws['AM62'] = df_fila["Residencia.8"]
            ws['AN62'] = df_fila["Tiempo trabajado.8"]
            ws['AO62'] = df_fila["# Personas núcleo familiar.8"]
            ws['AP62'] = df_fila["Personas a cargo.8"]
            ws['AQ62'] = df_fila["Lugar de residencia familiar.8"]
        
            actividad10 = df_fila['Remuneración.8']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT62'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU62'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV62'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW62'] = 'X'

            ##### Persona 10 #####

            if df_fila["Tipo de mano de obra.9"] == "Familiar":
                ws['B63'] = 'X'
            elif df_fila["Tipo de mano de obra.9"] == "Contratado":
                ws['D63'] = 'X'

            ws['E63'] = df_fila["Cargo.9"]

            if df_fila["Género.9"] == "Masculino":
                ws['M63'] = 'X'
            elif df_fila["Género.9"] == "Femenino":
                ws['K63'] = 'X'

            ws['N63'] = df_fila["Edad (años).9"]
            ws['Q63'] = df_fila["Duración jornada (horas).9"]

            actividad9 = df_fila['Escolaridad.9']
            if actividad9 == 'Primaria':
                ws['S63'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['U63'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['W63'] = 'X'
            elif actividad9 == 'Profesional':
                ws['Y63'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['AA63'] = 'X'

            if df_fila['Contrato.9'] == 'Tem.':
                ws['AC63'] = 'X'
            elif df_fila['Contrato.9'] == 'Fij':
                ws['AG63'] = 'X'
            
            if df_fila['Pago de seguridad.9'] == 'Si':
                ws['AH63'] = 'Si'
                ws['AJ63'] = ''
            elif df_fila['Pago de seguridad.9'] == 'No':
                ws['AH63'] = ''
                ws['AJ63'] = 'No'

            ws['AL63'] = df_fila["Procedencia.9"]
            ws['AM63'] = df_fila["Residencia.9"]
            ws['AN63'] = df_fila["Tiempo trabajado.9"]
            ws['AO63'] = df_fila["# Personas núcleo familiar.9"]
            ws['AP63'] = df_fila["Personas a cargo.9"]
            ws['AQ63'] = df_fila["Lugar de residencia familiar.9"]
        
            actividad10 = df_fila['Remuneración.9']
            if actividad10 == 'Inferiores a $900.000':
                ws['AT63'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AU63'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AV63'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AW63'] = 'X'


        elif df_fila["Contrata algún tipo de mano de obra"] == "No":
            ws['AF50'] = 'X'

def llenarInforme4(ws, df_fila):
        ws['AO1'] = df_fila["Encuesta No."]

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AQ2'] = fecha_str.split('/')[1]
                ws['AT2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AQ2'] = fecha_str.split('-')[1]
                ws['AT2'] = fecha_str.split('-')[0] ####### LLENAR FECHA EN ESPACIOS VACÍOS Y NO SOBRE EL SÍMBOLO.
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')

        ws['AN3'] = df_fila["Encuestador"]
        ws['F7'] = df_fila["Nombre"]
        ws['AB7'] = df_fila["Empresa"]
        ws['AR7'] = df_fila["Cargo"]

        if df_fila["¿Pertenece a alguna asociación?"] == 'Si':
            ws['AB8'] = 'X'
            ws['AO8'] = df_fila["Otro, ¿Cuál?"]
        elif df_fila["¿Pertenece a alguna asociación?"] == 'No':
            ws['AD8'] = 'X'

        actividad = df_fila['¿Qué tipo de servicios oferta?']
        if actividad == 'Restaurante, cafetería':
            ws['P13'] = 'X'
        elif actividad == 'Bar y centro nocturno':
            ws['P14'] = 'X'
        elif actividad == 'Servicios sexuales':
            ws['P15'] = 'X'
        elif actividad == 'Estación de servicios (combustible, montallantas, servicios)':
            ws['P17'] = 'X'
        elif actividad == 'Servicio de giros y/o financieros':
            ws['P19'] = 'X'
        elif actividad == 'Hospedaje (diligenciar título E)':
            ws['P21'] = 'X'

            ################
            #### ws E ####
            ################

            if df_fila['¿Qué tipo de hospedaje oferta?'] == 'Hotel':
                ws['F60'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Motel':
                ws['F61'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Apartahotel':
                ws['F62'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Pensión':
                ws['F63'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Cabaña':
                ws['M60'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Finca ecoturística':
                ws['M61'] = 'X'
            elif df_fila['¿Qué tipo de hospedaje oferta?'] == 'Otro':
                ws['M62'] = 'X'
                ws['J24'] = df_fila['Otro, ¿Cuál?.4']
            
            ws['R60'] = df_fila["¿Qué capacidad de alojamiento tiene?"]
            ws['AE60'] = df_fila["Semanalmente, ¿Cuántas personas en promedio hospeda?"]
            ws['AO60'] = df_fila["¿Cuáles son los principales sitios de procedencia de los huéspedes?"]


        elif actividad == 'Educacion':
            ws['P22'] = 'X'
        elif actividad == 'Otros':
            ws['P23'] = 'X'
            ws['J24'] = df_fila['Otros, ¿Cuáles?']

        if df_fila["Vende principalmente en:"] =='Sitio':
            ws['I28'] = 'X'
        if df_fila["Vende principalmente en:"] =='Vereda':
            ws['I29'] = 'X'
        if df_fila["Vende principalmente en:"] =='Casco Urbano':
            ws['I30'] = 'X'
        if df_fila["Vende principalmente en:"] =='Otros Municipios y/o Veredas':
            ws['T28'] = 'X'
            ws['P30'] = df_fila['Otros, ¿Cuáles?.1']


        ws['AB13'] = df_fila["Servicio 1"]
        ws['AN13'] = df_fila["Precio"]
        ws['AB14'] = df_fila["Servicio 2"]
        ws['AN14'] = df_fila["Precio.1"]
        ws['AB15'] = df_fila["Servicio 3"]
        ws['AN151'] = df_fila["Precio.2"]
        ws['AB16'] = df_fila["Servicio 4"]
        ws['AN16'] = df_fila["Precio.3"]

        ws['AC19'] = df_fila["Servicio 1.1"]
        ws['AC20'] = df_fila["Servicio 2.1"]
        ws['AO19'] = df_fila["Servicio 3"]
        ws['AO20'] = df_fila["Servicio 4"]

        actividad2 = df_fila['Frecuencia con que vende los servicios:']
        if actividad2 == 'Diario':
            ws['AI23'] = 'X'
        elif actividad2 == 'Semanal':
            ws['AI25'] = 'X'
        elif actividad2 == 'Quincenal':
            ws['AI24'] = 'X'
        elif actividad2 == 'Mensual':
            ws['AR24'] = 'X'


        if str(df_fila['Hidrocarburos']) != "nan":
            ws['AI28'] = 'X'
            ws['AO28'] = df_fila['Hidrocarburos']
        elif str(df_fila['Vereda']) != "nan":
            ws['AI29'] = 'X'
            ws['AO29'] = df_fila['Vereda']
        elif str(df_fila['Finca/Propiet.']) != "nan":
            ws['AI30'] = 'X'
            ws['AO30'] = df_fila['Finca/Propiet.']

        if df_fila["Sobre la actividad, piensa: Continuidad"] == "Continuar con la actividad":
            ws['L33'] = 'X'
            ws['AG34'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Continuidad"] == "Finalizar la actividad":
            ws['N33'] = 'X'
            ws['AE34'] = 'X'
        if df_fila["Sobre la actividad, piensa: Producción"] == "Ampliar la producción":
            ws['L34'] = 'X'
            ws['AG33'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Permanecer con la misma producción":
            ws['N34'] = 'X'
            ws['AE33'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Ninguna de las anteriores":
            ws['N34'] = 'X'
            ws['AG33'] = 'X'

        ws['AO34'] = df_fila['¿Por qué?']

        actividad6 = df_fila['¿De dónde se abastece del recurso hídrico?']
        if actividad6 == 'Aljibe':
            ws['W36'] = 'X'
        elif actividad6 == 'Acueducto Veredal':
            ws['AE36'] = 'X'
        elif actividad6 == 'Otro':
            ws['AL36'] = 'X'
            ws['AQ36'] = df_fila['Otro, ¿Cuál?.1']
    
        ws['W37'] = df_fila["Forma de extracción"]
        ws['AQ37'] = df_fila["Cantidad estimada (agregar m3)"]


        actividad7 = df_fila['¿Qué tipo de energía utiliza?']
        if actividad7 == 'Energía Eléctrica':
            ws['AA38'] = 'X'
        elif actividad7 == 'Energía Solar':
            ws['AJ38'] = 'X'
        elif actividad7 == 'Otro':
            ws['AP38'] = df_fila['Otro, ¿Cuál?.2']

        actividad8 = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
        if actividad8 == 'Energía elétrica':
            ws['AA39'] = 'X'
        elif actividad8 == 'Leña':
            ws['AF39'] = 'X'
        elif actividad8 == 'Gas':
            ws['AL39'] = 'X'
        elif actividad8 == 'Otro':
            ws['AQ39'] = df_fila['Otro, ¿Cuál?.3']

        if df_fila["¿Cuenta con servicio de alcantarillado?"] == "Si":
            ws['AB40'] = 'X'
            ws['AO40'] = df_fila['¿Cuál?']
        elif df_fila["¿Cuenta con servicio de alcantarillado?"] == "No":
            ws['AD40'] = 'X'

        ##### ABASTECIMIENTO DE INSUMOS #####

        ## SERVICIO 1
        ws['B44'] = df_fila["Servicio 1.2"]
        ws['J44'] = df_fila["Insumo/Materia prima"]
        ws['S44'] = df_fila["Precio compra"]
        ws['AB44'] = df_fila["Cantidad"]
        ws['AI44'] = df_fila["Frecuencia de compra"]
        ws['AQ44'] = df_fila["Procedencia"]

        ## SERVICIO 2
        ws['B45'] = df_fila["Servicio 2.2"]
        ws['J45'] = df_fila["Insumo/Materia prima.1"]
        ws['S45'] = df_fila["Precio compra.1"]
        ws['AB45'] = df_fila["Cantidad.1"]
        ws['AI45'] = df_fila["Frecuencia de compra.1"]
        ws['AQ45'] = df_fila["Procedencia.1"]

        ## SERVICIO 3
        ws['B46'] = df_fila["Servicio 3.1"]
        ws['J46'] = df_fila["Insumo/Materia prima.2"]
        ws['S46'] = df_fila["Precio compra.2"]
        ws['AB46'] = df_fila["Cantidad.2"]
        ws['AI46'] = df_fila["Frecuencia de compra.2"]
        ws['AQ46'] = df_fila["Procedencia.2"]

        ## SERVICIO 4
        ws['B47'] = df_fila["Servicio 4.1"]
        ws['J47'] = df_fila["Insumo/Materia prima.3"]
        ws['S47'] = df_fila["Precio compra.3"]
        ws['AB47'] = df_fila["Cantidad.3"]
        ws['AI47'] = df_fila["Frecuencia de compra.3"]
        ws['AQ47'] = df_fila["Procedencia.3"]

        ## SERVICIO 5
        ws['B48'] = df_fila["Servicio 5"]
        ws['J48'] = df_fila["Insumo/Materia prima.4"]
        ws['S48'] = df_fila["Precio compra.4"]
        ws['AB48'] = df_fila["Cantidad.4"]
        ws['AI48'] = df_fila["Frecuencia de compra.4"]
        ws['AQ48'] = df_fila["Procedencia.4"]

        ws['W49'] = df_fila["¿Cuál fue el monto total gastado en insumos del último mes?"]

        ##### EQUIPOS Y MAQUINARIA #####

        ## EQUIPO 1
        ws['B53'] = df_fila["Equipo/maquinaria"]
        ws['N53'] = df_fila["Precio compra"]
        ws['X53'] = df_fila["Cantidad que posee la unidad económica"]
        ws['AF53'] = df_fila["Vida útil"]
        ws['AO53'] = df_fila["Procedencia.5"]

        ## EQUIPO 2
        ws['B54'] = df_fila["Equipo/maquinaria.1"]
        ws['N54'] = df_fila["Precio compra.1"]
        ws['X54'] = df_fila["Cantidad que posee la unidad económica.1"]
        ws['AF54'] = df_fila["Vida útil.1"]
        ws['AO54'] = df_fila["Procedencia.6"]

        ## EQUIPO 3
        ws['B55'] = df_fila["Equipo/maquinaria.2"]
        ws['N55'] = df_fila["Precio compra.2"]
        ws['X55'] = df_fila["Cantidad que posee la unidad económica.2"]
        ws['AF55'] = df_fila["Vida útil.2"]
        ws['AO55'] = df_fila["Procedencia.7"]

        ## EQUIPO 4
        ws['B56'] = df_fila["Equipo/maquinaria.3"]
        ws['N56'] = df_fila["Precio compra.3"]
        ws['X56'] = df_fila["Cantidad que posee la unidad económica.3"]
        ws['AF56'] = df_fila["Vida útil.3"]
        ws['AO56'] = df_fila["Procedencia.8"]

        if df_fila["Contrata algún tipo de mano de obra"] == "Si":
            ws['AC64'] = 'X'

            #### Persona 1 ####

            if df_fila["Tipo de mano de obra"] == "Familiar":
                ws['B69'] = 'X'
            elif df_fila["Tipo de mano de obra"] == "Contratado":
                ws['D69'] = 'X'

            ws['E69'] = df_fila["Cargo.1"]

            if df_fila["Género"] == "Masculino":
                ws['J69'] = 'X'
            elif df_fila["Género"] == "Femenino":
                ws['H69'] = 'X'

            ws['K69'] = df_fila["Edad (años)"]
            ws['L69'] = df_fila["Duración jornada (horas)"]

            actividad9 = df_fila['Escolaridad']
            if actividad9 == 'Primaria':
                ws['N69'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q69'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S69'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U69'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W69'] = 'X'

            if df_fila['Contrato'] == 'Tem.':
                ws['Y69'] = 'X'
            elif df_fila['Contrato'] == 'Fij':
                ws['AC69'] = 'X'
            
            if df_fila['Pago de seguridad'] == 'Si':
                ws['AE69'] = 'X'
            elif df_fila['Pago de seguridad'] == 'No':
                ws['AG69'] = 'X'

            ws['AH69'] = df_fila["Procedencia.9"]
            ws['AI69'] = df_fila["Residencia"]
            ws['AL69'] = df_fila["Tiempo trabajado"]
            ws['AM69'] = df_fila["# Personas núcleo familiar"]
            ws['AO69'] = df_fila["Personas a cargo"]
            ws['AP69'] = df_fila["Lugar de residencia familiar"]
        
            actividad10 = df_fila['Remuneración']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR69'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS69'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT69'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU69'] = 'X'

            ##### Persona 2 #####

            if df_fila["Tipo de mano de obra.1"] == "Familiar":
                ws['B70'] = 'X'
            elif df_fila["Tipo de mano de obra.1"] == "Contratado":
                ws['D70'] = 'X'

            ws['E70'] = df_fila["Cargo.2"]

            if df_fila["Género.1"] == "Masculino":
                ws['J70'] = 'X'
            elif df_fila["Género.1"] == "Femenino":
                ws['H70'] = 'X'

            ws['K70'] = df_fila["Edad (años).1"]
            ws['L70'] = df_fila["Duración jornada (horas).1"]

            actividad9 = df_fila['Escolaridad.1']
            if actividad9 == 'Primaria':
                ws['N70'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q70'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S70'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U70'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W70'] = 'X'

            if df_fila['Contrato.1'] == 'Tem.':
                ws['Y70'] = 'X'
            elif df_fila['Contrato.1'] == 'Fij':
                ws['AC70'] = 'X'
            
            if df_fila['Pago de seguridad.1'] == 'Si':
                ws['AE70'] = 'X'
            elif df_fila['Pago de seguridad.1'] == 'No':
                ws['Ag70'] = 'X'

            ws['AH70'] = df_fila["Procedencia.10"]
            ws['AI70'] = df_fila["Residencia.1"]
            ws['AL70'] = df_fila["Tiempo trabajado.1"]
            ws['AM70'] = df_fila["# Personas núcleo familiar.1"]
            ws['AO70'] = df_fila["Personas a cargo.1"]
            ws['AP70'] = df_fila["Lugar de residencia familiar.1"]
        
            actividad10 = df_fila['Remuneración.1']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR70'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS70'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT70'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU70'] = 'X'

            ##### Persona 3 #####

            if df_fila["Tipo de mano de obra.2"] == "Familiar":
                ws['B71'] = 'X'
            elif df_fila["Tipo de mano de obra.2"] == "Contratado":
                ws['D71'] = 'X'

            ws['E71'] = df_fila["Cargo.3"]

            if df_fila["Género.2"] == "Masculino":
                ws['J71'] = 'X'
            elif df_fila["Género.2"] == "Femenino":
                ws['H71'] = 'X'

            ws['K71'] = df_fila["Edad (años).2"]
            ws['L71'] = df_fila["Duración jornada (horas).2"]

            actividad9 = df_fila['Escolaridad.2']
            if actividad9 == 'Primaria':
                ws['N71'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q71'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S71'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U71'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W71'] = 'X'

            if df_fila['Contrato.2'] == 'Tem.':
                ws['Y71'] = 'X'
            elif df_fila['Contrato.2'] == 'Fij':
                ws['AC71'] = 'X'
            
            if df_fila['Pago de seguridad.2'] == 'Si':
                ws['AE71'] = 'X'
            elif df_fila['Pago de seguridad.2'] == 'No':
                ws['AG71'] = 'X'

            ws['AH71'] = df_fila["Procedencia.11"]
            ws['AI71'] = df_fila["Residencia.2"]
            ws['AL71'] = df_fila["Tiempo trabajado.2"]
            ws['AM71'] = df_fila["# Personas núcleo familiar.2"]
            ws['AO71'] = df_fila["Personas a cargo.2"]
            ws['AP71'] = df_fila["Lugar de residencia familiar.2"]
        
            actividad10 = df_fila['Remuneración.2']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR71'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS71'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT71'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU71'] = 'X'

            ##### Persona 4 #####

            if df_fila["Tipo de mano de obra.3"] == "Familiar":
                ws['B72'] = 'X'
            elif df_fila["Tipo de mano de obra.3"] == "Contratado":
                ws['D72'] = 'X'

            ws['E72'] = df_fila["Cargo.4"]

            if df_fila["Género.3"] == "Masculino":
                ws['J72'] = 'X'
            elif df_fila["Género.3"] == "Femenino":
                ws['H72'] = 'X'

            ws['K72'] = df_fila["Edad (años).3"]
            ws['L72'] = df_fila["Duración jornada (horas).3"]

            actividad9 = df_fila['Escolaridad.3']
            if actividad9 == 'Primaria':
                ws['N72'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q72'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S72'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U72'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W72'] = 'X'

            if df_fila['Contrato.3'] == 'Tem.':
                ws['Y72'] = 'X'
            elif df_fila['Contrato.3'] == 'Fij':
                ws['AC72'] = 'X'
            
            if df_fila['Pago de seguridad.3'] == 'Si':
                ws['AE72'] = 'X'
            elif df_fila['Pago de seguridad.3'] == 'No':
                ws['AG72'] = 'X'

            ws['AH72'] = df_fila["Procedencia.12"]
            ws['AI72'] = df_fila["Residencia.3"]
            ws['AL72'] = df_fila["Tiempo trabajado.3"]
            ws['AM72'] = df_fila["# Personas núcleo familiar.3"]
            ws['AO72'] = df_fila["Personas a cargo.3"]
            ws['AP72'] = df_fila["Lugar de residencia familiar.3"]
        
            actividad10 = df_fila['Remuneración.3']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR72'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS72'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT72'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU72'] = 'X'

            ##### Persona 5 #####

            if df_fila["Tipo de mano de obra.4"] == "Familiar":
                ws['B73'] = 'X'
            elif df_fila["Tipo de mano de obra.4"] == "Contratado":
                ws['D73'] = 'X'

            ws['E73'] = df_fila["Cargo.5"]

            if df_fila["Género.4"] == "Masculino":
                ws['J73'] = 'X'
            elif df_fila["Género.4"] == "Femenino":
                ws['H73'] = 'X'

            ws['K73'] = df_fila["Edad (años).4"]
            ws['L73'] = df_fila["Duración jornada (horas).4"]

            actividad9 = df_fila['Escolaridad.4']
            if actividad9 == 'Primaria':
                ws['N73'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q73'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S73'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U73'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W73'] = 'X'

            if df_fila['Contrato.4'] == 'Tem.':
                ws['Y73'] = 'X'
            elif df_fila['Contrato.4'] == 'Fij':
                ws['AC73'] = 'X'
            
            if df_fila['Pago de seguridad.4'] == 'Si':
                ws['AE73'] = 'X'
            elif df_fila['Pago de seguridad.4'] == 'No':
                ws['AG73'] = 'X'

            ws['AH73'] = df_fila["Procedencia.13"]
            ws['AI73'] = df_fila["Residencia.4"]
            ws['AL73'] = df_fila["Tiempo trabajado.4"]
            ws['AM73'] = df_fila["# Personas núcleo familiar.4"]
            ws['AO73'] = df_fila["Personas a cargo.4"]
            ws['AP73'] = df_fila["Lugar de residencia familiar.4"]
        
            actividad10 = df_fila['Remuneración.4']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR73'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS73'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT73'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU73'] = 'X'

            ##### Persona 6 #####

            if df_fila["Tipo de mano de obra.5"] == "Familiar":
                ws['B74'] = 'X'
            elif df_fila["Tipo de mano de obra.5"] == "Contratado":
                ws['D74'] = 'X'

            ws['E74'] = df_fila["Cargo.6"]

            if df_fila["Género.5"] == "Masculino":
                ws['j74'] = 'X'
            elif df_fila["Género.5"] == "Femenino":
                ws['H74'] = 'X'

            ws['K74'] = df_fila["Edad (años).5"]
            ws['L74'] = df_fila["Duración jornada (horas).5"]

            actividad9 = df_fila['Escolaridad.5']
            if actividad9 == 'Primaria':
                ws['N74'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q74'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S74'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U74'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W74'] = 'X'

            if df_fila['Contrato.5'] == 'Tem.':
                ws['Y74'] = 'X'
            elif df_fila['Contrato.5'] == 'Fij':
                ws['AC74'] = 'X'
            
            if df_fila['Pago de seguridad.5'] == 'Si':
                ws['AE74'] = 'X'
            elif df_fila['Pago de seguridad.5'] == 'No':
                ws['AG74'] = 'X'

            ws['AH74'] = df_fila["Procedencia.14"]
            ws['AI74'] = df_fila["Residencia.5"]
            ws['AL74'] = df_fila["Tiempo trabajado.5"]
            ws['AM74'] = df_fila["# Personas núcleo familiar.5"]
            ws['AO74'] = df_fila["Personas a cargo.5"]
            ws['AP74'] = df_fila["Lugar de residencia familiar.5"]
        
            actividad10 = df_fila['Remuneración.5']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR74'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS74'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT74'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU74'] = 'X'

            ##### Persona 7 #####

            if df_fila["Tipo de mano de obra.6"] == "Familiar":
                ws['B75'] = 'X'
            elif df_fila["Tipo de mano de obra.6"] == "Contratado":
                ws['D75'] = 'X'

            ws['E75'] = df_fila["Cargo.7"]

            if df_fila["Género.6"] == "Masculino":
                ws['J75'] = 'X'
            elif df_fila["Género.6"] == "Femenino":
                ws['H75'] = 'X'

            ws['K75'] = df_fila["Edad (años).6"]
            ws['L75'] = df_fila["Duración jornada (horas).6"]

            actividad9 = df_fila['Escolaridad.6']
            if actividad9 == 'Primaria':
                ws['N75'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q75'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S75'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U75'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W75'] = 'X'

            if df_fila['Contrato.6'] == 'Tem.':
                ws['Y75'] = 'X'
            elif df_fila['Contrato.6'] == 'Fij':
                ws['AC75'] = 'X'
            
            if df_fila['Pago de seguridad.6'] == 'Si':
                ws['AE75'] = 'X'
            elif df_fila['Pago de seguridad.6'] == 'No':
                ws['AG75'] = 'X'

            ws['AH75'] = df_fila["Procedencia.15"]
            ws['AI75'] = df_fila["Residencia.6"]
            ws['AL75'] = df_fila["Tiempo trabajado.6"]
            ws['AM75'] = df_fila["# Personas núcleo familiar.6"]
            ws['AO75'] = df_fila["Personas a cargo.6"]
            ws['AP75'] = df_fila["Lugar de residencia familiar.6"]
        
            actividad10 = df_fila['Remuneración.6']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR75'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS75'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT75'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU75'] = 'X'

            ##### Persona 8 #####

            if df_fila["Tipo de mano de obra.7"] == "Familiar":
                ws['B76'] = 'X'
            elif df_fila["Tipo de mano de obra.7"] == "Contratado":
                ws['D76'] = 'X'

            ws['E76'] = df_fila["Cargo.8"]

            if df_fila["Género.7"] == "Masculino":
                ws['J76'] = 'X'
            elif df_fila["Género.7"] == "Femenino":
                ws['H76'] = 'X'

            ws['K76'] = df_fila["Edad (años).7"]
            ws['L76'] = df_fila["Duración jornada (horas).7"]

            actividad9 = df_fila['Escolaridad.7']
            if actividad9 == 'Primaria':
                ws['N76'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q76'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S76'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U76'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W76'] = 'X'

            if df_fila['Contrato.7'] == 'Tem.':
                ws['Y76'] = 'X'
            elif df_fila['Contrato.7'] == 'Fij':
                ws['AC76'] = 'X'
            
            if df_fila['Pago de seguridad.7'] == 'Si':
                ws['AE76'] = 'X'
            elif df_fila['Pago de seguridad.7'] == 'No':
                ws['AG76'] = 'X'

            ws['AH76'] = df_fila["Procedencia.16"]
            ws['AI76'] = df_fila["Residencia.7"]
            ws['AL76'] = df_fila["Tiempo trabajado.7"]
            ws['AM76'] = df_fila["# Personas núcleo familiar.7"]
            ws['AO76'] = df_fila["Personas a cargo.7"]
            ws['AP76'] = df_fila["Lugar de residencia familiar.7"]
        
            actividad10 = df_fila['Remuneración.7']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR76'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS76'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT76'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU76'] = 'X'

            ##### Persona 9 #####

            if df_fila["Tipo de mano de obra.8"] == "Familiar":
                ws['B77'] = 'X'
            elif df_fila["Tipo de mano de obra.8"] == "Contratado":
                ws['D77'] = 'X'

            ws['E77'] = df_fila["Cargo.9"]

            if df_fila["Género.8"] == "Masculino":
                ws['J77'] = 'X'
            elif df_fila["Género.8"] == "Femenino":
                ws['H77'] = 'X'

            ws['K77'] = df_fila["Edad (años).8"]
            ws['L77'] = df_fila["Duración jornada (horas).8"]

            actividad9 = df_fila['Escolaridad.8']
            if actividad9 == 'Primaria':
                ws['N77'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q77'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S77'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U77'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W77'] = 'X'

            if df_fila['Contrato.8'] == 'Tem.':
                ws['Y77'] = 'X'
            elif df_fila['Contrato.8'] == 'Fij':
                ws['AC77'] = 'X'
            
            if df_fila['Pago de seguridad.8'] == 'Si':
                ws['AE77'] = 'X'
            elif df_fila['Pago de seguridad.8'] == 'No':
                ws['AG77'] = 'X'

            ws['AH77'] = df_fila["Procedencia.17"]
            ws['AI77'] = df_fila["Residencia.8"]
            ws['AL77'] = df_fila["Tiempo trabajado.8"]
            ws['AM77'] = df_fila["# Personas núcleo familiar.8"]
            ws['AO77'] = df_fila["Personas a cargo.8"]
            ws['AP77'] = df_fila["Lugar de residencia familiar.8"]
        
            actividad10 = df_fila['Remuneración.8']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR77'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS77'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT77'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU77'] = 'X'

            ##### Persona 10 #####

            if df_fila["Tipo de mano de obra.9"] == "Familiar":
                ws['B78'] = 'X'
            elif df_fila["Tipo de mano de obra.9"] == "Contratado":
                ws['D78'] = 'X'

            ws['E78'] = df_fila["Cargo.9"]

            if df_fila["Género.9"] == "Masculino":
                ws['J78'] = 'X'
            elif df_fila["Género.9"] == "Femenino":
                ws['H78'] = 'X'

            ws['K78'] = df_fila["Edad (años).9"]
            ws['L78'] = df_fila["Duración jornada (horas).9"]

            actividad9 = df_fila['Escolaridad.9']
            if actividad9 == 'Primaria':
                ws['N78'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q78'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S78'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U78'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W78'] = 'X'

            if df_fila['Contrato.9'] == 'Tem.':
                ws['Y78'] = 'X'
            elif df_fila['Contrato.9'] == 'Fij':
                ws['AC78'] = 'X'
            
            if df_fila['Pago de seguridad.9'] == 'Si':
                ws['AE78'] = 'X'
            elif df_fila['Pago de seguridad.9'] == 'No':
                ws['AG78'] = 'X'

            ws['AH78'] = df_fila["Procedencia.18"]
            ws['AI78'] = df_fila["Residencia.9"]
            ws['AL78'] = df_fila["Tiempo trabajado.9"]
            ws['AM78'] = df_fila["# Personas núcleo familiar.9"]
            ws['AO78'] = df_fila["Personas a cargo.9"]
            ws['AP78'] = df_fila["Lugar de residencia familiar.9"]
        
            actividad10 = df_fila['Remuneración.9']
            if actividad10 == 'Inferiores a $900.000':
                ws['AR78'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AS78'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AT78'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AU78'] = 'X'


        elif df_fila["Contrata algún tipo de mano de obra"] == "No":
            ws['AE64'] = 'X'

def llenarInforme5(ws, df_fila):
        ws['AO1'] = df_fila['Encuesta No.']

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AQ2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AQ2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[0]
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')  
        
        ws['AO3'] = df_fila['Encuestador']


        # A. IDENTIFICACIÓN ENTREVISTADO

        ws['F7'] = df_fila['Nombre']
        ws['Y7'] = df_fila['Empresa']
        ws['AO7'] = df_fila['Cargo']

        asociacion = df_fila['¿Pertenece a alguna asociación?']
        if pd.notna(asociacion):
            if asociacion == 'Si':
                ws['AB8'] = 'X'
                ws['AO8'] = df_fila['Otro, ¿Cuál?']
            elif asociacion == 'No':
                ws['AD8'] = 'X'
        else:
            print(f'Campo vacío')


        # B. DESCRIPCIÓN DE LA ACTIVIDAD

        ws['A12'] = df_fila['Con cuántos empleados cuenta su empresa actualmente']

        partip_capital = df_fila['Participación de capital\nSi (Responder 3)']
        if pd.notna(partip_capital):
            if partip_capital == 'Si':
                ws['P15'] = 'X'
                ws['K16'] = df_fila['País del cual proviene el capital']

            elif partip_capital == 'No':
                ws['R15'] = 'X'
            
        else:
            print('Campo vacío')

        # 4. Tipo de servicio

        tipo_servicio = df_fila['Tipo de servicio']
        if pd.notna(tipo_servicio):
            if tipo_servicio == 'Servicios relacionados con perforación de pozos':
                ws['R19'] = 'X'
            elif tipo_servicio == 'Venta y alquiler de equipos y herramientas':
                ws['R20'] = 'X'
            elif tipo_servicio == 'Transporte de maquinaria o combustibles':
                ws['R21'] = 'X'
            elif tipo_servicio == 'Otros servicios':
                ws['R22'] = 'X'
                ws['G23'] = df_fila['Otros, ¿Cuáles?']
        else:
            print('Campo vacio')

        # 5. ¿Cuál es el principal servicio demandado por el sector de hidrocarburos?

        ws['A25'] = df_fila['¿Cuál es el principal servicio demandado por el sector de hidrocarburos?']

        # 6. Frecuencia con la que son demandados los servicios

        actividad_frecuencia = df_fila['Frecuencia con la que son demandados los servicios por parte del sector de hidrocarburos']
        if pd.notna(actividad_frecuencia):
            if actividad_frecuencia == 'Mensual':
                ws['F29'] = 'X'
            elif actividad_frecuencia == 'Bimestral':
                ws['F30'] = 'X'
            elif actividad_frecuencia == 'Trimestral':
                ws['F31'] = 'X'
            elif actividad_frecuencia == 'Semestral':
                ws['K28'] = 'X'
            elif actividad_frecuencia == 'Semanal':
                ws['M29'] = 'X'
            elif actividad_frecuencia == 'Anual':
                ws['M30'] = 'X'            
            elif actividad_frecuencia == 'Contrato permanente':
                ws['V30'] = 'X'
            elif actividad_frecuencia == 'Otro':
                ws['M31'] = 'X'
                ws['P32'] = df_fila['Otro, ¿Cuál?.1']
        else: 
            print('Campo vacío')

        # 7. En promedio cuántos servicios son contratados al mes por el sector de hidrocarburos
        ws['C36'] = df_fila['Servicio 1']
        ws['N36'] = df_fila['Cantidad']

        ws['C37'] = df_fila['Servicio 2']
        ws['N37'] = df_fila['Cantidad.1']

        ws['C38'] = df_fila['Servicio 3']
        ws['N38'] = df_fila['Cantidad.2']

        ws['C39'] = df_fila['Servicio 4']
        ws['N39'] = df_fila['Cantidad.3']


        # 8. Ingreso promedio mensual

        ws['N40'] = df_fila['Ingreso promedio mensual ($)']


        # 9. ¿El pago por sus servicios es oportuno?
        pago_oportuno = df_fila['¿El pago por sus servicios es oportuno?']
        if pd.notna(pago_oportuno):
            if pago_oportuno == 'Si':
                ws['P42'] = 'X'
            elif pago_oportuno == 'No':
                ws['R42'] = 'X'
        else:
            print('Campo vacío')

        # 10. ¿Cuál es el precio de los servicios ofertados para el sector de hidrocarburos?
        
        ws['AA13'] = df_fila['Servicio 1.1']
        ws['AJ13'] = df_fila['Cantidad (según frecuencia)']
        ws['AR13'] = df_fila['Precio']

        ws['AA14'] = df_fila['Servicio 2.1']
        ws['AJ14'] = df_fila['Cantidad (según frecuencia).1']
        ws['AR14'] = df_fila['Precio.1']


        ws['AA15'] = df_fila['Servicio 3.1']
        ws['AJ15'] = df_fila['Cantidad (según frecuencia).2']
        ws['AR15'] = df_fila['Precio.2']


        ws['AA16'] = df_fila['Servicio 4.1']
        ws['AJ16'] = df_fila['Cantidad (según frecuencia).3']
        ws['AR16'] = df_fila['Precio.3']



        # 11. ¿Cuál es el costo aproximado de los servicios ofertados según la frecuencia de venta?
        ws['AB20'] = df_fila['Servicio 1.2']
        ws['AQ20'] = df_fila['Costo']

        ws['AB21'] = df_fila['Servicio 2.2']
        ws['AQ21'] = df_fila['Costo.1']

        ws['AB22'] = df_fila['Servicio 3.2']
        ws['AQ22'] = df_fila['Costo.2']
        
        ws['AB23'] = df_fila['Servicio 4.2']
        ws['AQ23'] = df_fila['Costo.3']

        # 12. ¿Cuál es el costo o valor aproximado de los insumos o equipos utilizados para prestar los servicios en el último mes?
        ws['AD26'] = df_fila['¿Cuál es el costo o valor aproximado de los insumos o equipos utilizados para prestar los servicios en el último mes?']

        # 13. ¿De dónde provienen los insumos y/o maquinaria y/o equipos que emplea?
        ws['Z29'] = df_fila['¿De dónde provienen los insumos y/o maquinaria y/o equipos que emplea?']

        # 14. Vende principalmente en:
        lugar_venta = df_fila['Vende principalmente en:']
        if pd.notna(lugar_venta):
            if lugar_venta == 'Sitio':
                ws['AH32'] = 'X'
            elif lugar_venta == 'Vereda':
                ws['AH33'] = 'X'
            elif lugar_venta == 'Casco Urbano':
                ws['AH35'] = 'X'
            elif lugar_venta == 'Otros municipios y/o veredas':
                ws['AS33'] = 'X'
                ws['AO35'] = df_fila['¿Cuáles?.1']
        else:
            print('Campo vacío')

        # 15. ¿Posee algún permiso ambiental?
        permiso_ambiental = df_fila['¿Posee algún permiso ambiental?']
        if pd.notna(permiso_ambiental):
            if permiso_ambiental == 'No':
                ws['AM37'] = 'X'
            elif permiso_ambiental == 'Si':
                ws['AK37'] = 'X'
                ws['AR37'] = df_fila['¿Cuál?']
        else:
            print('Campo vacío')

        # 16. Mantenimiento de la actividad
        continuar = df_fila['¿Piensa continuar con la actividad?']
        if pd.notna(permiso_ambiental):
            if continuar == 'Si':
                ws['AQ39'] = 'X'
            elif continuar == 'No':
                ws['AS39'] = 'X'
        else:
            print('Campo vacio')

        produccion = df_fila['¿Piensa ampliar la producción?']
        if pd.notna(permiso_ambiental):
            if produccion == 'Si':
                ws['AQ40'] = 'X'
            elif produccion == 'No':
                ws['AS40'] = 'X'
        else:
            print('Campo vacio')


        permanecer = df_fila['¿Piensa permanecer con la misma producción?']
        if pd.notna(permiso_ambiental):
            if permanecer == 'Si':
                ws['AQ41'] = 'X'
            elif permanecer == 'No':
                ws['AS41'] = 'X'
        else:
            print('Campo vacio')

        finalizar = df_fila['¿Piensa finalizar el servicio?']
        if pd.notna(permiso_ambiental):
            if finalizar == 'Si':
                ws['AQ42'] = 'X'
            elif finalizar == 'No':
                ws['AS42'] = 'X'
        else:
            print('Campo vacio')
        
        
        # C. INFORMACIÓN LABORAL

        # Mano de obra contratada
        
        ws['I48'] = df_fila['#']

        genero = df_fila['Género']
        if pd.notna(genero):
            if genero == 'Masculino':
                ws['M48'] = 'X'
            elif genero == 'Femenino':
                ws['K48'] = 'X'
        else:
            print('Campo vacio')

        contrato = df_fila['Contrato']
        if pd.notna(genero):
            if contrato == 'Termino Fijo':
                ws['O48'] = df_fila['¿Cuánto?']
            elif contrato == 'Indefinido':
                ws['S48'] = 'X'
        else:
            print('Campo vacio')

        ws['W48'] = df_fila['Jornal y turno laboral']

        escolaridad = df_fila['Escolaridad']
        if pd.notna(df_fila['Escolaridad']):
            if df_fila['Escolaridad'] == 'Primaria':
                ws['AB48'] = 'X'
            elif df_fila['Escolaridad'] == 'Bachillerato':
                ws['AC48'] = 'X'
            elif df_fila['Escolaridad'] == 'Técnico o tecnológico ':
                ws['AD48'] = 'X'
            elif df_fila['Escolaridad'] == 'Profesional':
                ws['AE48'] = 'X'
            elif df_fila['Escolaridad'] == 'Posgrado':
                ws['AG48'] = 'X'
        else:
            print('Campo vacio')

        procedencia = df_fila['Procedencia']
        if pd.notna(procedencia):
            if procedencia == 'Vereda':
                ws['AJ48'] = 'X'
            elif procedencia == 'Municipio':
                ws['AM48'] = 'X'
            elif procedencia == 'Otro':
                ws['AO48'] = 'X'
        else:
            print('Campo vacio')

        residencia = df_fila['Residencia']
        if pd.notna(escolaridad):
            if residencia == 'Vereda':
                ws['AQ48'] = 'X'
            elif residencia == 'Municipio':
                ws['AS48'] = 'X'
            elif residencia == 'Otro':
                ws['AU48'] = 'X'
        else:
            print('Campo vacio')
    

        # Mano de obra no calificada

        ws['I49'] = df_fila['#.1']

        genero = df_fila['Género.1']
        if pd.notna(genero):
            if genero == 'Masculino':
                ws['M49'] = 'X'
            elif genero == 'Femenino':
                ws['K49'] = 'X'
        else:
            print('Campo vacio')

        contrato = df_fila['Contrato.1']
        if pd.notna(genero):
            if contrato == 'Termino Fijo':
                ws['O49'] = df_fila['¿Cuánto?.1']
            elif contrato == 'Indefinido':
                ws['S49'] = 'X'
        else:
            print('Campo vacio')

        ws['W49'] = df_fila['Jornal y turno laboral.1']

        escolaridad = df_fila['Escolaridad.1']
        if pd.notna(escolaridad):
            if escolaridad == 'Primaria':
                ws['AB49'] = 'X'
            elif escolaridad == 'Bachillerato':
                ws['AC49'] = 'X'
            elif escolaridad == 'Técnico o tecnológico ':
                ws['AD49'] = 'X'
            elif escolaridad == 'Profesional':
                ws['AE49'] = 'X'
            elif escolaridad == 'Posgrado':
                ws['AG49'] = 'X'
        else:
            print('Campo vacio')

        procedencia = df_fila['Procedencia.1']
        if pd.notna(escolaridad):
            if procedencia == 'Vereda':
                ws['AJ49'] = 'X'
            elif procedencia == 'Municipio':
                ws['AM49'] = 'X'
            elif procedencia == 'Otro':
                ws['AO49'] = 'X'
        else:
            print('Campo vacio')

        residencia = df_fila['Residencia.1']
        if pd.notna(escolaridad):
            if residencia == 'Vereda':
                ws['AQ49'] = 'X'
            elif residencia == 'Municipio':
                ws['AS49'] = 'X'
            elif residencia == 'Otro':
                ws['AU49'] = 'X'
        else:
            print('Campo vacio')



        # Empleados administrativos y contables


        ws['I50'] = df_fila['#.2']

        genero = df_fila['Género.2']
        if pd.notna(genero):
            if genero == 'Masculino':
                ws['M50'] = 'X'
            elif genero == 'Femenino':
                ws['K50'] = 'X'
        else:
            print('Campo vacio')

        contrato = df_fila['Contrato.2']
        if pd.notna(genero):
            if contrato == 'Termino Fijo':
                ws['O50'] = df_fila['¿Cuánto?.2']
            elif contrato == 'Indefinido':
                ws['S50'] = 'X'
        else:
            print('Campo vacio')

        ws['W50'] = df_fila['Jornal y turno laboral.2']

        escolaridad = df_fila['Escolaridad.2']
        if pd.notna(escolaridad):
            if escolaridad == 'Primaria':
                ws['AB50'] = 'X'
            elif escolaridad == 'Bachillerato':
                ws['AC50'] = 'X'
            elif escolaridad == 'Técnico o tecnológico ':
                ws['AD50'] = 'X'
            elif escolaridad == 'Profesional':
                ws['AE50'] = 'X'
            elif escolaridad == 'Posgrado':
                ws['AG50'] = 'X'
        else:
            print('Campo vacio')

        procedencia = df_fila['Procedencia.2']
        if pd.notna(escolaridad):
            if procedencia == 'Vereda':
                ws['AJ50'] = 'X'
            elif procedencia == 'Municipio':
                ws['AM50'] = 'X'
            elif procedencia == 'Otro':
                ws['AO50'] = 'X'
        else:
            print('Campo vacio')

        residencia = df_fila['Residencia.2']
        if pd.notna(escolaridad):
            if residencia == 'Vereda':
                ws['AQ50'] = 'X'
            elif residencia == 'Municipio':
                ws['AS50'] = 'X'
            elif residencia == 'Otro':
                ws['AU50'] = 'X'
        else:
            print('Campo vacio')

        # Gerentes y directivos

        ws['I52'] = df_fila['#.3']

        genero = df_fila['Género.3']
        if pd.notna(genero):
            if genero == 'Masculino':
                ws['M52'] = 'X'
            elif genero == 'Femenino':
                ws['K52'] = 'X'
        else:
            print('Campo vacio')

        contrato = df_fila['Contrato.3']
        if pd.notna(genero):
            if contrato == 'Termino Fijo':
                ws['O52'] = df_fila['¿Cuánto?.3']
            elif contrato == 'Indefinido':
                ws['S52'] = 'X'
        else:
            print('Campo vacio')

        ws['W52'] = df_fila['Jornal y turno laboral.3']

        escolaridad = df_fila['Escolaridad.3']
        if pd.notna(escolaridad):
            if escolaridad == 'Primaria':
                ws['AB52'] = 'X'
            elif escolaridad == 'Bachillerato':
                ws['AC52'] = 'X'
            elif escolaridad == 'Técnico o tecnológico ':
                ws['AD52'] = 'X'
            elif escolaridad == 'Profesional':
                ws['AE52'] = 'X'
            elif escolaridad == 'Posgrado':
                ws['AG52'] = 'X'
        else:
            print('Campo vacio')

        procedencia = df_fila['Procedencia.3']
        if pd.notna(escolaridad):
            if procedencia == 'Vereda':
                ws['AJ52'] = 'X'
            elif procedencia == 'Municipio':
                ws['AM52'] = 'X'
            elif procedencia == 'Otro':
                ws['AO52'] = 'X'
        else:
            print('Campo vacio')

        residencia = df_fila['Residencia.3']
        if pd.notna(escolaridad):
            if residencia == 'Vereda':
                ws['AQ52'] = 'X'
            elif residencia == 'Municipio':
                ws['AS52'] = 'X'
            elif residencia == 'Otro':
                ws['AU52'] = 'X'
        else:
            print('Campo vacio')
        
        # 28. ¿Contrata servicios profesionales?
        if df_fila['Contrata servicios profesionales * Sí (Responder 30 y 31)'] == 'Si':
            ws['L56'] = 'X'

            # ¿Qué tipo de servicios? 
            servicios = df_fila['¿Qué tipo de servicios?']
            if pd.notna(servicios):
                if servicios == 'Contaduría':
                    ws['AK56'] = 'X'
                elif servicios == 'Consultoría':
                    ws['AT56'] = 'X'
                elif servicios == 'Asesoría legal':
                    ws['AK57'] = 'X'
                elif servicios == 'Otros':
                    ws['AT57'] = 'X'
                    ws['AE58'] = df_fila['¿Cuál?.1']
            else:
                print('Campo vacio')

            # 30. Con qué frecuencia contrata servicios profesionales

            ws['A59'] = df_fila['Servicio 1.3']

            frecuencia_servicios1 = df_fila['Frecuencia']
            if pd.notna(frecuencia_servicios1):
                if frecuencia_servicios1 == 'Mensual':
                    ws['F59'] = 'X'
                elif frecuencia_servicios1 == 'Semestral':
                    ws['J59'] = 'X'
                elif frecuencia_servicios1 == 'Trimestral':
                    ws['P59'] = 'X'
                elif frecuencia_servicios1 == 'Anual':
                    ws['U59'] = 'X'

            ws['A60'] = df_fila['Servicio 2.3']

            frecuencia_servicios2 = df_fila['Frecuencia.1']
            if pd.notna(frecuencia_servicios2):
                if frecuencia_servicios2 == 'Mensual':
                    ws['F60'] = 'X'
                elif frecuencia_servicios2 == 'Semestral':
                    ws['J60'] = 'X'
                elif frecuencia_servicios2 == 'Trimestral':
                    ws['P60'] = 'X'
                elif frecuencia_servicios2 == 'Anual':
                    ws['U60'] = 'X'
            else:
                print('Campo vacio')

            ws['A61'] = df_fila['Servicio 3.3']

            frecuencia_servicios3 = df_fila['Frecuencia.2']
            if pd.notna(frecuencia_servicios3):
                if frecuencia_servicios3 == 'Mensual':
                    ws['F61'] = 'X'
                elif frecuencia_servicios3 == 'Semestral':
                    ws['J61'] = 'X'
                elif frecuencia_servicios3 == 'Trimestral':
                    ws['P61'] = 'X'
                elif frecuencia_servicios3 == 'Anual':
                    ws['U61'] = 'X'
            else:
                print('Campo vacio')

            ws['A62'] = df_fila['Servicio 4.3']

            frecuencia_servicios4 = df_fila['Frecuencia.3']
            if pd.notna(frecuencia_servicios4):
                if frecuencia_servicios4 == 'Mensual':
                    ws['F62'] = 'X'
                elif frecuencia_servicios4 == 'Semestral':
                    ws['J62'] = 'X'
                elif frecuencia_servicios4 == 'Trimestral':
                    ws['P62'] = 'X'
                elif frecuencia_servicios4 == 'Anual':
                    ws['U62'] = 'X'
            else:
                print('Campo vacio')

            ws['A63'] = df_fila['Servicio 5.3']
            frecuencia_servicios5 = df_fila['Frecuencia.4']
            if pd.notna(frecuencia_servicios5):
                if frecuencia_servicios5 == 'Mensual':
                    ws['F63'] = 'X'
                elif frecuencia_servicios5 == 'Semestral':
                    ws['J63'] = 'X'
                elif frecuencia_servicios5 == 'Trimestral':
                    ws['P63'] = 'X'
                elif frecuencia_servicios5 == 'Anual':
                    ws['U63'] = 'X'       
            else:
                print('Campo vacio')

            # 31. ¿Cuál es el monto pagado por estos servicios durante el último semestre?
            ws['AE62'] = df_fila['¿Cuál es el monto pagado por estos servicios durante el último semestre?']


        elif df_fila['Contrata servicios profesionales * Sí (Responder 30 y 31)'] == 'No':
            ws['N56'] = 'X'
        


        # D. REMUNERACIONES
        ws['Z65'] = df_fila['Salarios pagados a la mano de obra calificada']
        ws['Z66'] = df_fila['Salarios pagados a la mano de obra no calificada']
        ws['Z67'] = df_fila['Salarios pagados a empleados y administrativos']
        ws['Z68'] = df_fila['Salarios pagados a gerentes y directivos']
        ws['Z69'] = df_fila['Total remuneraciones']


def llenarInforme6(ws, df_fila):    
        # A. IDENTIFICACIÓN ENTREVISTADO
        ws['AQ1'] = df_fila['Encuesta No.']

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AO2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AR2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AO2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AR2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[0]
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')  
        
        ws['AP3'] = df_fila['Encuestador']

        ws['F7'] = df_fila['Nombre']
        ws['Z7'] = df_fila['Empresa']
        ws['AQ7'] = df_fila['Cargo']


        pertenece_asociacion = df_fila['¿Pertenece a alguna asociación?']
        if pd.notna(pertenece_asociacion):
            if pertenece_asociacion == 'Si':
                ws['AA8'] = 'X'
                ws['AO8'] = df_fila['Otro, ¿Cuál?']
            elif pertenece_asociacion == 'No':
                ws['AC8'] = 'X'
        else:
            print("Campo vacío")

        
        # Pregunta 1: Bien final producido
        ws['A12'] = df_fila['Bien final producido']

        # Pregunta 2: ¿Con cuántos empleados cuenta la empresa?
        ws['A16'] = df_fila['¿Con cuántos empleados cuenta la empresa?']

        # Pregunta 3: La empresa cuenta con algún tipo de permiso ambiental
        permiso_ambiental = df_fila['La empresa cuenta con algún tipo de permiso ambiental']
        if permiso_ambiental == 'Si':
            ws['Q13'] = 'x'
            ws['W13'] = df_fila['¿Cuál?']
        elif permiso_ambiental == 'No':
            ws['S13'] = 'x'
        

        # Pregunta 4: Tipo de empresa
        tipo_empresa = df_fila['Tipo de Empresa']
        if tipo_empresa == 'Pública':
            ws['R17'] = 'X'
        elif tipo_empresa == 'Privada':
            ws['X17'] = 'X'
        elif tipo_empresa == 'Mixta':
            ws['AD17'] = 'X'


        # Pregunta 5: Vende principalmente en
        if df_fila['Vende principalmente en'] == 'Sitio':
            ws['AU13'] = 'x'
        elif df_fila['Vende principalmente en'] == 'Vereda':
            ws['AU14'] = 'x'
        elif df_fila['Vende principalmente en'] == 'Casco Urbano':
            ws['AU15'] = 'x'
        elif df_fila['Vende principalmente en'] == 'Otros Municipios y/o Veredas':
            ws['AU16'] = 'x'
            ws['AN17'] = df_fila['Otros, ¿Cuáles?']

        # Pregunta 6: Procedencia de los compradores
        if pd.notna(df_fila['Hidrocarburos']):

            ws['V18'] = 'X'
            ws['AC18'] = df_fila['Hidrocarburos']
        
        if pd.notna(df_fila['Otro']):

            ws['V19'] = 'X'
            ws['AC19'] = df_fila['Otro']

        
        # Sobre la actividad, piensa:

        continuidad = df_fila['Sobre la actividad, piensa: Continuidad']
        if continuidad == 'Continuar con la actividad':
            ws['L21'] = 'X'
            
        elif continuidad == 'Finalizar la actividad':
            ws['N21'] = 'X'

        produccion = df_fila['Sobre la actividad, piensa: Producción']
        if produccion == 'Ampliar la producción':
            ws['AB21'] = 'X'
            ws['AU21'] = 'X'               
        elif produccion == 'Permanecer con la misma producción':
            ws['AD21'] = 'X'
            ws['AS21'] = 'X'   
        elif produccion == "Ninguna de las anteriores":
            ws['AD21'] = 'X'
            ws['AU21'] = 'X'

        columnas = {
            "Tipo de producto fabricado": "B",
            "Unidad de medida": "K",
            "Cantidad producida": "P",
            "Frecuencia de producción": "V",
            "Costos de producción por unidad": "AE",
            "Cantidad vendida por semana": "AM",
            "Precio de venta": "AS"
        }

        
        for i in range(3):
            fila_id = 26 + i  # Empezar desde la fila 25 y avanzar
            ws[f"{columnas['Tipo de producto fabricado']}{fila_id}"] = valorCol('Tipo de producto fabricado', i, df_fila)
            ws[f"{columnas['Unidad de medida']}{fila_id}"] = valorCol('Unidad de medida', i, df_fila)
            ws[f"{columnas['Cantidad producida']}{fila_id}"] = valorCol('Cantidad producida', i, df_fila)
            ws[f"{columnas['Frecuencia de producción']}{fila_id}"] = valorCol('Frecuencia de producción', i, df_fila)
            ws[f"{columnas['Costos de producción por unidad']}{fila_id}"] = valorCol('Costos de producción por unidad', i, df_fila)
            ws[f"{columnas['Cantidad vendida por semana']}{fila_id}"] = valorCol('Cantidad vendida por semana', i, df_fila)
            ws[f"{columnas['Precio de venta']}{fila_id}"] = valorCol('Precio de venta', i, df_fila)

        # Equipo/maquinaria
        ws['B33'] = df_fila['Equipo/maquinaria 1']
        ws['B34'] = df_fila['Equipo/maquinaria 2']
        ws['B35'] = df_fila['Equipo/maquinaria 3']

        # Precio al que lo compró
        ws['P33'] = df_fila['Precio al que lo compró']
        ws['P34'] = df_fila['Precio al que lo compró.1']
        ws['P35'] = df_fila['Precio al que lo compró.2']

        # Cantidad que posee la unidad económica
        ws['Z33'] = df_fila['Cantidad que posee la unidad económica']
        ws['Z34'] = df_fila['Cantidad que posee la unidad económica.1']
        ws['Z35'] = df_fila['Cantidad que posee la unidad económica.2']

        # Vida útil
        ws['AJ33'] = df_fila['Vida útil']
        ws['AJ34'] = df_fila['Vida útil.1']
        ws['AJ35'] = df_fila['Vida útil.2']

        # Procedencia
        ws['AR33'] = df_fila['Procedencia']
        ws['AR34'] = df_fila['Procedencia.1']
        ws['AR35'] = df_fila['Procedencia.2']
            
        # Servicios
        ws['B40'] = df_fila['Servicios']
        ws['B41'] = df_fila['Servicios.1']
        ws['B42'] = df_fila['Servicios.2']

        # Insumo/Materia prima
        ws['J40'] = df_fila['Insumo/Materia prima 1']
        ws['J41'] = df_fila['Insumo/Materia prima 2']
        ws['J41'] = df_fila['Insumo/Materia prima 3']

        # Precio compra
        ws['T40'] = df_fila['Precio compra']
        ws['T41'] = df_fila['Precio compra.1']
        ws['T42'] = df_fila['Precio compra.2']

        # Cantidad
        ws['AB40'] = df_fila['Cantidad']
        ws['AB41'] = df_fila['Cantidad.1']
        ws['AB42'] = df_fila['Cantidad.2']

        # Frecuencia de compra
        ws['AJ40'] = df_fila['Frecuencia de compra']
        ws['AJ41'] = df_fila['Frecuencia de compra.1']
        ws['AJ42'] = df_fila['Frecuencia de compra.2']

        # Procedencia
        ws['AR40'] = df_fila['Procedencia.3']
        ws['AR41'] = df_fila['Procedencia.4']
        ws['AR42'] = df_fila['Procedencia.5']

        agua_fuente = df_fila['¿De dónde se abastece del recurso hídrico?']
        if agua_fuente == 'Aljibe':
            ws['W43'] = 'X'
        elif agua_fuente == 'Acueducto Veredal':
            ws['AG43'] = 'X'
        elif agua_fuente == 'Otro':
            ws['AN43'] = 'X'
            ws['AT43'] = df_fila['¿Cuál?.1']    

        ws['W44'] = df_fila['Forma de extracción']

        ws['AO44'] = df_fila['Cantidad estimada (m3)']

        energia = df_fila['¿Qué tipo de energía utiliza?']
        if energia == 'Energía Eléctrica':
            ws['AC45'] = 'X'
        elif energia == 'Energía Solar':
            ws['AL45'] = 'X'
        elif energia == 'Otro':
            ws['AT45'] = df_fila['¿Cuál?.2']

        energia_coccion = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
        if energia_coccion == 'Energía Eléctrica':
            ws['AC46'] = 'X'
        elif energia_coccion == 'Leña':
            ws['AH46'] = 'X'
        elif energia_coccion == 'Gas':
            ws['AN46'] = 'X'
        elif energia_coccion == 'Otro':       
            ws['AT46'] = df_fila['¿Cuál?.3']

        alcantarillado = df_fila['¿Cuenta con servicio de alcantarillado?']
        if pd.notna(alcantarillado):
            if alcantarillado == 'Si':
                ws['AB47'] = 'X'
                ws['AO47'] = df_fila['¿Cuál?.4']  
            elif alcantarillado == 'No':
                ws['AD47'] = 'X'

        ws['AC49'] = df_fila['¿Cuál fue el monto total gastado en insumos del último mes?']


        servicio = df_fila['¿Demanda algún tipo de servicio de la región?']
        if pd.notna(servicio):
            if servicio == 'Seguridad':
                ws['L52'] = 'X'
            elif servicio == 'Mano de obra calificada':
                ws['L53'] = 'X'
            elif servicio == 'Mano de obra no calificada':
                ws['L54'] = 'X'
            elif servicio == 'Transporte':
                ws['L55'] = 'X'
            elif servicio == 'Alojamiento':
                ws['V52'] = 'X'
            elif servicio == 'Alimentación':
                ws['V53'] = 'X'
            elif servicio == 'Otro':
                ws['V54'] = 'X'
                ws['P55'] = df_fila['Otro, ¿Cuál?.1']
        else:
            print('Campo vacío')

        ws['AC51'] = df_fila['¿Con que frecuencia demanda servicios de la región?']
        
        for i in range(10):
            prefijo_persona = 60 + i
            ws[f'E{prefijo_persona}'] = valorCol('Cargo', i+1, df_fila)
            ws[f'K{prefijo_persona}'] = valorCol('Edad (años)', i, df_fila)
            ws[f'L{prefijo_persona}'] = valorCol('Duración jornada (horas)', i, df_fila)

            manoObra = valorCol('Tipo de mano de obra', i, df_fila)
            if pd.notna(manoObra):
                if manoObra == 'Familiar':
                    ws[f'B{prefijo_persona}'] = 'X'
                elif manoObra == 'Contratado':
                    ws[f'D{prefijo_persona}'] = 'X'

            # Genero
            genero = valorCol('Género', i,df_fila)
            if pd.notna(genero):
                if genero == 'Masculino':
                    ws[f'J{prefijo_persona}'] = 'X'
                elif genero ==  'Femenino':
                    ws[f'H{prefijo_persona}'] = 'X'

            # Escolaridad 
            escolaridad = valorCol('Escolaridad', i, df_fila)
            if pd.notna(escolaridad):
                if escolaridad:
                    if escolaridad == 'Primaria':
                        ws[f'N{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Bachillerato':
                        ws[f'Q{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Técnico':
                        ws[f'S{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Pregrado':
                        ws[f'U{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Posgrado':
                        ws[f'W{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Contrato 
            contrato = valorCol('Contrato', i,df_fila)
            if contrato:
                if contrato == 'Tem.':
                    ws[f'AC{prefijo_persona}'] = 'X'
                elif contrato == 'Fij':
                    ws[f'AE{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Pago de seguridad social 
            pago_seguridad = valorCol('Pago de seguridad', i, df_fila)
            if pago_seguridad:
                if pago_seguridad == 'Si':
                    ws[f'AG{prefijo_persona}'] = 'X'
                elif pago_seguridad == 'No':
                    ws[f'AI{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Remuneración 
            remuneracion = valorCol('Remuneración', i, df_fila)
            if remuneracion:
                if remuneracion == 'Inferiores a $900.000':
                    ws[f'AU{prefijo_persona}'] = 'X'
                elif remuneracion == '$900.000 - $1.800.000':
                    ws[f'AV{prefijo_persona}'] = 'X'
                elif remuneracion == '$1.801.000 - $2.700.000':
                    ws[f'AW{prefijo_persona}'] = 'X'
                elif remuneracion == 'Superiores a $2.701.000':
                    ws[f'AX{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Información adicional
        
            ws[f'AJ{prefijo_persona}'] = valorCol('Procedencia', 6 + i, df_fila)
            ws[f'AK{prefijo_persona}'] = valorCol('Residencia', i, df_fila)
            ws[f'AN{prefijo_persona}'] = valorCol('Tiempo trabajado', i, df_fila)
            ws[f'AO{prefijo_persona}'] = valorCol('# Personas núcleo familiar', i, df_fila)
            ws[f'AQ{prefijo_persona}'] = valorCol('Personas a cargo', i, df_fila)
            ws[f'AS{prefijo_persona}'] = valorCol('Lugar de residencia familiar', i,df_fila)

def llenarInforme7(ws, df_fila):
        ws['AQ1'] = df_fila["Encuesta No."]

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AO2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AR2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AO2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AR2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[0] ####### LLENAR FECHA EN ESPACIOS VACÍOS Y NO SOBRE EL SÍMBOLO.
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')

        ws['AP3'] = df_fila["Encuestador"]
        ws['F7'] = df_fila["Nombre"]
        ws['Z7'] = df_fila["Empresa"]
        ws['AQ7'] = df_fila["Cargo"]

        if df_fila["¿Pertenece a alguna asociación?"] == 'Si':
            ws['AA8'] = 'X'
            ws['AO8'] = df_fila["Otro, ¿Cuál?"]
            
        elif df_fila["¿Pertenece a alguna asociación?"] == 'No':
            ws['AC8'] = 'X'

        

        actividad = df_fila['¿Qué tipo de servicio de transporte ofrece?']
        if actividad == 'Transporte público':
            ws['T12'] = 'X'
        elif actividad == 'Servicios especiales o transporte de pasajeros':
            ws['T13'] = 'X'
        elif actividad == 'Transporte de carga o maquinaria':
            ws['T14'] = 'X'


        actividad2 = df_fila['Presta los servicios de transporte como']
        if actividad2 == 'Particular':
            ws['F16'] = 'X'
        elif actividad2 == 'Afiliado':
            ws['F17'] = 'X'
        elif actividad2 == 'Cooperativa':
            ws['M16'] = actividad2
        elif actividad2 == 'Empresa':
            ws['M17'] = 'X'
        elif actividad2 == 'Otro':
            ws['N17'] = df_fila["Otro, ¿Cuál?"]


        if df_fila["¿Presta los servicios al sector de hidrocarburos?"] == 'Si':
            ws['AM12'] = 'X'

            if df_fila["¿El pago por parte del prestador de hidrocarburos es oportuno?"] == 'Si':
                ws['AM15'] = 'X'

            elif df_fila["¿El pago por parte del prestador de hidrocarburos es oportuno?"] == 'No':
                ws['AO15'] = 'X'

            ws['AI16'] = df_fila['Observaciones']

        elif df_fila["¿Presta los servicios al sector de hidrocarburos?"] == 'No':
            ws['AO12'] = 'X'



        if df_fila["Sobre la actividad, piensa: Continuidad"] == "Continuar con la actividad":
            ws['L19'] = 'X'
            ws['AU20'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Continuidad"] == "Finalizar la actividad":
            ws['N19'] = 'X'
            ws['AS20'] = 'X'
        if df_fila["Sobre la actividad, piensa: Producción"] == "Ampliar la producción":
            ws['AB19'] = 'X'
            ws['AU19'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Permanecer con la misma producción":
            ws['AD19'] = 'X'
            ws['AS19'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Ninguna de las anteriores":
            ws['AD19'] = 'X'
            ws['AU19'] = 'X'

        if df_fila["Se encuentra afiliado a alguna empresa y/o cooperativa de transporte:"] == 'Si':
            ws['H24'] = 'X'

            ws['D25'] = df_fila['¿Cuál?.1']
            ws['R25'] = df_fila['Número de Contacto']


            ws['A31'] = df_fila['¿Cuál es el porcentaje pagado a la cooperativa por cada servicio? (%)']

            ws['C37'] = df_fila['Cuántos afiliados tiene la cooperativa y/o empresa de transporte']

            ws['AC37'] = df_fila['¿Cuál es el porcentaje cobrado a los afiliados por cada servicio prestado? (%)']


        elif df_fila["Se encuentra afiliado a alguna empresa y/o cooperativa de transporte:"] == 'No':
            ws['J24'] = 'X'


        if df_fila["¿Es propietario del vehículo?"] == 'Si':
            ws['H28'] = 'X'

        elif df_fila["¿Es propietario del vehículo?"] == 'No':
            ws['J28'] = 'X'

        ws['A33'] = df_fila['Durante la última semana cuántos Km recorrió']
        ws['AP22'] = df_fila['Hace cuánto presta servicios de transporte']
        
        if df_fila['¿Cuál es el costo aproximado de mantenimiento y gastos del vehículo en una semana?'] == "Entre $100.000 - $200.000":
            ws['AS26'] = 'X'
        elif df_fila['¿Cuál es el costo aproximado de mantenimiento y gastos del vehículo en una semana?'] == "Entre $201.000 - $400.000":
            ws['AS27'] = 'X'
        elif df_fila['¿Cuál es el costo aproximado de mantenimiento y gastos del vehículo en una semana?'] == "Entre $401.000 - $600.000":
            ws['AS28'] = 'X'
        elif df_fila['¿Cuál es el costo aproximado de mantenimiento y gastos del vehículo en una semana?'] == "Mayor a $600.000":
            ws['AS29'] = 'X'

        if df_fila["¿El estado de las vías le genera sobre costos?"] == 'Si':
            ws['AN32'] = 'X'
            ws['AL33'] = df_fila['Costos Incurridos']
        elif df_fila["¿El estado de las vías le genera sobre costos?"] == 'No':
            ws['AP32'] = 'X'

        ws['D40'] = df_fila['Costo 1']
        ws['Q40'] = df_fila['Valor']
        ws['D41'] = df_fila['Costo 2']
        ws['Q41'] = df_fila['Valor.1']
        ws['D42'] = df_fila['Costo 3']
        ws['Q42'] = df_fila['Valor.2']

        if df_fila["Emplea directamente algún tipo de mano de obra (si la respuesta es SI, diligenciar el título G)"] == 'Si':
            ws['AN40'] = 'X'
            
            #### Persona 1 ####
            if df_fila["Tipo de mano de obra"] == "Familiar":
                ws['B64'] = 'X'
            elif df_fila["Tipo de mano de obra"] == "Contratado":
                ws['D64'] = 'X'

            ws['E64'] = df_fila["Cargo.1"]

            if df_fila["Género"] == "Masculino":
                ws['J64'] = 'X'
            elif df_fila["Género"] == "Femenino":
                ws['H64'] = 'X'

            ws['K64'] = df_fila["Edad (años)"]
            ws['L64'] = df_fila["Duración jornada (horas)"]

            actividad9 = df_fila['Escolaridad']
            if actividad9 == 'Primaria':
                ws['N64'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q64'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S64'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U64'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W64'] = 'X'

            if df_fila['Contrato'] == 'Tem.':
                ws['AC64'] = 'X'
            elif df_fila['Contrato'] == 'Fij':
                ws['AE64'] = 'X'
            
            if df_fila['Pago de seguridad'] == 'Si':
                ws['AG64'] = 'X'
            elif df_fila['Pago de seguridad'] == 'No':
                ws['AI64'] = 'X'

            ws['AJ64'] = df_fila["Procedencia.4"]
            ws['AK64'] = df_fila["Residencia"]
            ws['AN64'] = df_fila["Tiempo trabajado"]
            ws['AO64'] = df_fila["# Personas núcleo familiar"]
            ws['AQ64'] = df_fila["Personas a cargo"]
            ws['AS64'] = df_fila["Lugar de residencia familiar"]
        
            actividad10 = df_fila['Remuneración']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU64'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV64'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW64'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX64'] = 'X'

            ##### Persona 2 #####

            if df_fila["Tipo de mano de obra.1"] == "Familiar":
                ws['B65'] = 'X'
            elif df_fila["Tipo de mano de obra.1"] == "Contratado":
                ws['D65'] = 'X'

            ws['E65'] = df_fila["Cargo.2"]

            if df_fila["Género.1"] == "Masculino":
                ws['J65'] = 'X'
            elif df_fila["Género.1"] == "Femenino":
                ws['H65'] = 'X'

            ws['K65'] = df_fila["Edad (años).1"]
            ws['L65'] = df_fila["Duración jornada (horas).1"]

            actividad9 = df_fila['Escolaridad.1']
            if actividad9 == 'Primaria':
                ws['N65'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q65'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S65'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U65'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W65'] = 'X'

            if df_fila['Contrato.1'] == 'Tem.':
                ws['AC65'] = 'X'
            elif df_fila['Contrato.1'] == 'Fij':
                ws['AE65'] = 'X'
            
            if df_fila['Pago de seguridad.1'] == 'Si':
                ws['AG65'] = 'X'
            elif df_fila['Pago de seguridad.1'] == 'No':
                ws['AI65'] = 'X'

            ws['AJ65'] = df_fila["Procedencia.5"]
            ws['AK65'] = df_fila["Residencia.1"]
            ws['AN65'] = df_fila["Tiempo trabajado.1"]
            ws['AO65'] = df_fila["# Personas núcleo familiar.1"]
            ws['AQ65'] = df_fila["Personas a cargo.1"]
            ws['AS65'] = df_fila["Lugar de residencia familiar.1"]
        
            actividad10 = df_fila['Remuneración.1']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU65'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV65'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW65'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX65'] = 'X'

            ##### Persona 3 #####

            if df_fila["Tipo de mano de obra.2"] == "Familiar":
                ws['B66'] = 'X'
            elif df_fila["Tipo de mano de obra.2"] == "Contratado":
                ws['D66'] = 'X'

            ws['E66'] = df_fila["Cargo.3"]

            if df_fila["Género.2"] == "Masculino":
                ws['J66'] = 'X'
            elif df_fila["Género.2"] == "Femenino":
                ws['H66'] = 'X'

            ws['K66'] = df_fila["Edad (años).2"]
            ws['L66'] = df_fila["Duración jornada (horas).2"]

            actividad9 = df_fila['Escolaridad.2']
            if actividad9 == 'Primaria':
                ws['N66'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q66'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S66'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U66'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W66'] = 'X'

            if df_fila['Contrato.2'] == 'Tem.':
                ws['AC66'] = 'X'
            elif df_fila['Contrato.2'] == 'Fij':
                ws['AE66'] = 'X'
            
            if df_fila['Pago de seguridad.2'] == 'Si':
                ws['AG66'] = 'X'
            elif df_fila['Pago de seguridad.2'] == 'No':
                ws['AI66'] = 'X'

            ws['AJ66'] = df_fila["Procedencia.6"]
            ws['AK66'] = df_fila["Residencia.2"]
            ws['AN66'] = df_fila["Tiempo trabajado.2"]
            ws['AO66'] = df_fila["# Personas núcleo familiar.2"]
            ws['AQ66'] = df_fila["Personas a cargo.2"]
            ws['AS66'] = df_fila["Lugar de residencia familiar.2"]
        
            actividad10 = df_fila['Remuneración.2']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU66'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV66'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW66'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX66'] = 'X'

            ##### Persona 4 #####

            if df_fila["Tipo de mano de obra.3"] == "Familiar":
                ws['B67'] = 'X'
            elif df_fila["Tipo de mano de obra.3"] == "Contratado":
                ws['D67'] = 'X'

            ws['E67'] = df_fila["Cargo.4"]

            if df_fila["Género.3"] == "Masculino":
                ws['J67'] = 'X'
            elif df_fila["Género.3"] == "Femenino":
                ws['H67'] = 'X'

            ws['K67'] = df_fila["Edad (años).3"]
            ws['L67'] = df_fila["Duración jornada (horas).3"]

            actividad9 = df_fila['Escolaridad.3']
            if actividad9 == 'Primaria':
                ws['N67'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q67'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S67'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U67'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W67'] = 'X'

            if df_fila['Contrato.3'] == 'Tem.':
                ws['AC67'] = 'X'
            elif df_fila['Contrato.3'] == 'Fij':
                ws['AE67'] = 'X'
            
            if df_fila['Pago de seguridad.3'] == 'Si':
                ws['AG67'] = 'X'
            elif df_fila['Pago de seguridad.3'] == 'No':
                ws['AI67'] = 'X'

            ws['AJ67'] = df_fila["Procedencia.7"]
            ws['AK67'] = df_fila["Residencia.3"]
            ws['AN67'] = df_fila["Tiempo trabajado.3"]
            ws['AO67'] = df_fila["# Personas núcleo familiar.3"]
            ws['AQ67'] = df_fila["Personas a cargo.3"]
            ws['AS67'] = df_fila["Lugar de residencia familiar.3"]
        
            actividad10 = df_fila['Remuneración.3']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU67'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV67'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW67'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX67'] = 'X'

            ##### Persona 5 #####

            if df_fila["Tipo de mano de obra.4"] == "Familiar":
                ws['B68'] = 'X'
            elif df_fila["Tipo de mano de obra.4"] == "Contratado":
                ws['D68'] = 'X'

            ws['E68'] = df_fila["Cargo.5"]

            if df_fila["Género.4"] == "Masculino":
                ws['J68'] = 'X'
            elif df_fila["Género.4"] == "Femenino":
                ws['H68'] = 'X'

            ws['K68'] = df_fila["Edad (años).4"]
            ws['L68'] = df_fila["Duración jornada (horas).4"]

            actividad9 = df_fila['Escolaridad.4']
            if actividad9 == 'Primaria':
                ws['N68'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q68'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S68'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U68'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W68'] = 'X'

            if df_fila['Contrato.4'] == 'Tem.':
                ws['AC68'] = 'X'
            elif df_fila['Contrato.4'] == 'Fij':
                ws['AE68'] = 'X'
            
            if df_fila['Pago de seguridad.4'] == 'Si':
                ws['AG68'] = 'X'
            elif df_fila['Pago de seguridad.4'] == 'No':
                ws['AI68'] = 'X'

            ws['AJ68'] = df_fila["Procedencia.8"]
            ws['AK68'] = df_fila["Residencia.4"]
            ws['AN68'] = df_fila["Tiempo trabajado.4"]
            ws['AO68'] = df_fila["# Personas núcleo familiar.4"]
            ws['AQ68'] = df_fila["Personas a cargo.4"]
            ws['AS68'] = df_fila["Lugar de residencia familiar.4"]
        
            actividad10 = df_fila['Remuneración.4']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU68'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV68'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW68'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX68'] = 'X'

            ##### Persona 6 #####

            if df_fila["Tipo de mano de obra.5"] == "Familiar":
                ws['B69'] = 'X'
            elif df_fila["Tipo de mano de obra.5"] == "Contratado":
                ws['D69'] = 'X'

            ws['E69'] = df_fila["Cargo.6"]

            if df_fila["Género.5"] == "Masculino":
                ws['j69'] = 'X'
            elif df_fila["Género.5"] == "Femenino":
                ws['H69'] = 'X'

            ws['K69'] = df_fila["Edad (años).5"]
            ws['L69'] = df_fila["Duración jornada (horas).5"]

            actividad9 = df_fila['Escolaridad.5']
            if actividad9 == 'Primaria':
                ws['N69'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q69'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S69'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U69'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W69'] = 'X'

            if df_fila['Contrato.5'] == 'Tem.':
                ws['AC69'] = 'X'
            elif df_fila['Contrato.5'] == 'Fij':
                ws['AE69'] = 'X'
            
            if df_fila['Pago de seguridad.5'] == 'Si':
                ws['AG69'] = 'X'
            elif df_fila['Pago de seguridad.5'] == 'No':
                ws['AI69'] = 'X'

            ws['AJ69'] = df_fila["Procedencia.9"]
            ws['AK69'] = df_fila["Residencia.5"]
            ws['AN69'] = df_fila["Tiempo trabajado.5"]
            ws['AO69'] = df_fila["# Personas núcleo familiar.5"]
            ws['AQ69'] = df_fila["Personas a cargo.5"]
            ws['AS69'] = df_fila["Lugar de residencia familiar.5"]
        
            actividad10 = df_fila['Remuneración.5']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU69'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV69'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW69'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX69'] = 'X'

            ##### Persona 7 #####

            if df_fila["Tipo de mano de obra.6"] == "Familiar":
                ws['B70'] = 'X'
            elif df_fila["Tipo de mano de obra.6"] == "Contratado":
                ws['D70'] = 'X'

            ws['E70'] = df_fila["Cargo.7"]

            if df_fila["Género.6"] == "Masculino":
                ws['J70'] = 'X'
            elif df_fila["Género.6"] == "Femenino":
                ws['H70'] = 'X'

            ws['K70'] = df_fila["Edad (años).6"]
            ws['L70'] = df_fila["Duración jornada (horas).6"]

            actividad9 = df_fila['Escolaridad.6']
            if actividad9 == 'Primaria':
                ws['N70'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q70'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S70'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U70'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W70'] = 'X'

            if df_fila['Contrato.6'] == 'Tem.':
                ws['AC70'] = 'X'
            elif df_fila['Contrato.6'] == 'Fij':
                ws['AE70'] = 'X'
            
            if df_fila['Pago de seguridad.6'] == 'Si':
                ws['AG70'] = 'X'
            elif df_fila['Pago de seguridad.6'] == 'No':
                ws['AI70'] = 'X'

            ws['AJ70'] = df_fila["Procedencia.10"]
            ws['AK70'] = df_fila["Residencia.6"]
            ws['AN70'] = df_fila["Tiempo trabajado.6"]
            ws['AO70'] = df_fila["# Personas núcleo familiar.6"]
            ws['AQ70'] = df_fila["Personas a cargo.6"]
            ws['AS70'] = df_fila["Lugar de residencia familiar.6"]
        
            actividad10 = df_fila['Remuneración.6']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU70'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV70'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW70'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX70'] = 'X'

            ##### Persona 8 #####

            if df_fila["Tipo de mano de obra.7"] == "Familiar":
                ws['B71'] = 'X'
            elif df_fila["Tipo de mano de obra.7"] == "Contratado":
                ws['D71'] = 'X'

            ws['E71'] = df_fila["Cargo.8"]

            if df_fila["Género.7"] == "Masculino":
                ws['J71'] = 'X'
            elif df_fila["Género.7"] == "Femenino":
                ws['H71'] = 'X'

            ws['K71'] = df_fila["Edad (años).7"]
            ws['L71'] = df_fila["Duración jornada (horas).7"]

            actividad9 = df_fila['Escolaridad.7']
            if actividad9 == 'Primaria':
                ws['N71'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q71'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S71'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U71'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W71'] = 'X'

            if df_fila['Contrato.7'] == 'Tem.':
                ws['AC71'] = 'X'
            elif df_fila['Contrato.7'] == 'Fij':
                ws['AE71'] = 'X'
            
            if df_fila['Pago de seguridad.7'] == 'Si':
                ws['AG71'] = 'X'
            elif df_fila['Pago de seguridad.7'] == 'No':
                ws['AI71'] = 'X'

            ws['AJ71'] = df_fila["Procedencia.11"]
            ws['AK71'] = df_fila["Residencia.7"]
            ws['AN71'] = df_fila["Tiempo trabajado.7"]
            ws['AO71'] = df_fila["# Personas núcleo familiar.7"]
            ws['AQ71'] = df_fila["Personas a cargo.7"]
            ws['AS71'] = df_fila["Lugar de residencia familiar.7"]
        
            actividad10 = df_fila['Remuneración.7']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU71'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV71'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW71'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX71'] = 'X'

            ##### Persona 9 #####

            if df_fila["Tipo de mano de obra.8"] == "Familiar":
                ws['B72'] = 'X'
            elif df_fila["Tipo de mano de obra.8"] == "Contratado":
                ws['D72'] = 'X'

            ws['E72'] = df_fila["Cargo.9"]

            if df_fila["Género.8"] == "Masculino":
                ws['J72'] = 'X'
            elif df_fila["Género.8"] == "Femenino":
                ws['H72'] = 'X'

            ws['K72'] = df_fila["Edad (años).8"]
            ws['L72'] = df_fila["Duración jornada (horas).8"]

            actividad9 = df_fila['Escolaridad.8']
            if actividad9 == 'Primaria':
                ws['N72'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q72'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S72'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U72'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W72'] = 'X'

            if df_fila['Contrato.8'] == 'Tem.':
                ws['AC72'] = 'X'
            elif df_fila['Contrato.8'] == 'Fij':
                ws['AE72'] = 'X'
            
            if df_fila['Pago de seguridad.8'] == 'Si':
                ws['AG72'] = 'X'
            elif df_fila['Pago de seguridad.8'] == 'No':
                ws['AI72'] = 'X'

            ws['AJ72'] = df_fila["Procedencia.12"]
            ws['AK72'] = df_fila["Residencia.8"]
            ws['AN72'] = df_fila["Tiempo trabajado.8"]
            ws['AO72'] = df_fila["# Personas núcleo familiar.8"]
            ws['AQ72'] = df_fila["Personas a cargo.8"]
            ws['AS72'] = df_fila["Lugar de residencia familiar.8"]
        
            actividad10 = df_fila['Remuneración.8']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU72'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV72'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW72'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX72'] = 'X'

            ##### Persona 10 #####

            if df_fila["Tipo de mano de obra.9"] == "Familiar":
                ws['B73'] = 'X'
            elif df_fila["Tipo de mano de obra.9"] == "Contratado":
                ws['D73'] = 'X'

            ws['E73'] = df_fila["Cargo.9"]

            if df_fila["Género.9"] == "Masculino":
                ws['J73'] = 'X'
            elif df_fila["Género.9"] == "Femenino":
                ws['H73'] = 'X'

            ws['K73'] = df_fila["Edad (años).9"]
            ws['L73'] = df_fila["Duración jornada (horas).9"]

            actividad9 = df_fila['Escolaridad.9']
            if actividad9 == 'Primaria':
                ws['N73'] = 'X'
            elif actividad9 == 'Bachillerato':
                ws['Q73'] = 'X'
            elif actividad9 == 'Técnico o tecnológico':
                ws['S73'] = 'X'
            elif actividad9 == 'Profesional':
                ws['U73'] = 'X'
            elif actividad9 == 'Posgrado':
                ws['W73'] = 'X'

            if df_fila['Contrato.9'] == 'Tem.':
                ws['AC73'] = 'X'
            elif df_fila['Contrato.9'] == 'Fij':
                ws['AE73'] = 'X'
            
            if df_fila['Pago de seguridad.9'] == 'Si':
                ws['AG73'] = 'X'
            elif df_fila['Pago de seguridad.9'] == 'No':
                ws['AI73'] = 'X'

            ws['AJ73'] = df_fila["Procedencia.13"]
            ws['AK73'] = df_fila["Residencia.9"]
            ws['AN73'] = df_fila["Tiempo trabajado.9"]
            ws['AO73'] = df_fila["# Personas núcleo familiar.9"]
            ws['AQ73'] = df_fila["Personas a cargo.9"]
            ws['AS73'] = df_fila["Lugar de residencia familiar.9"]
        
            actividad10 = df_fila['Remuneración.9']
            if actividad10 == 'Inferiores a $900.000':
                ws['AU73'] = 'X'
            elif actividad10 == '$901.000 - $1.800.000':
                ws['AV73'] = 'X'
            elif actividad10 == '$1.801.000 - $2.700.000':
                ws['AW73'] = 'X'
            elif actividad10 == 'superiores a $2.701.000':
                ws['AX73'] = 'X'


        elif df_fila["Emplea directamente algún tipo de mano de obra (si la respuesta es SI, diligenciar el título G)"] == 'No':
            ws['AP40'] = 'X'

        ##### INFORMACIÓN TRANSPORTE DE PASAJEROS #####

        ws['N47'] = df_fila['Tarifa única ($)']
        ws['V47'] = df_fila['Servicio contratado por día ($)']
        ws['AE47'] = df_fila['Servicio contratado por semana ($)']
        ws['AL47'] = df_fila['Servicio contratado por mes ($)']
        ws['AT47'] = df_fila['Servicio contratado por km recorrido ($)']

        ws['M48'] = df_fila['Tarifa única ($).1']
        ws['U48'] = df_fila['Servicio contratado por día ($).1']
        ws['AD48'] = df_fila['Servicio contratado por semana ($).1']
        ws['AK48'] = df_fila['Servicio contratado por mes ($).1']
        ws['AS48'] = df_fila['Servicio contratado por km recorrido ($).1']

        ws['M50'] = df_fila['Tarifa única ($).2']
        ws['U50'] = df_fila['Servicio contratado por día ($).2']
        ws['AD50'] = df_fila['Servicio contratado por semana ($).2']
        ws['AK50'] = df_fila['Servicio contratado por mes ($).2']
        ws['AS50'] = df_fila['Servicio contratado por km recorrido ($).2']

        ws['AC52'] = df_fila['¿Cuál es el destino más frecuente?']


        ##### INFORMACIÓN TRANSPORTE DE PASAJEROS 2 #####

        ws['A56'] = df_fila['Elemento transportado']
        ws['I56'] = df_fila['Cantidad mensual']
        ws['M56'] = df_fila['Procedencia']
        ws['T56'] = df_fila['Destino']
        ws['AA56'] = df_fila['Frecuencia de movilización']
        ws['AG56'] = df_fila['Costo promedio del flete o tarifa']
        ws['AM56'] = df_fila['Promedio mensual e ingreso']
        ws['AT56'] = df_fila['Medio de transporte']

        ws['A57'] = df_fila['Elemento transportado.1']
        ws['I57'] = df_fila['Cantidad mensual.1']
        ws['M57'] = df_fila['Procedencia.1']
        ws['T57'] = df_fila['Destino.1']
        ws['AA57'] = df_fila['Frecuencia de movilización.1']
        ws['AG57'] = df_fila['Costo promedio del flete o tarifa.1']
        ws['AM57'] = df_fila['Promedio mensual e ingreso.1']
        ws['AT57'] = df_fila['Medio de transporte.1']

        ws['A58'] = df_fila['Elemento transportado.2']
        ws['I58'] = df_fila['Cantidad mensual.2']
        ws['M58'] = df_fila['Procedencia.2']
        ws['T58'] = df_fila['Destino.2']
        ws['AA58'] = df_fila['Frecuencia de movilización.2']
        ws['AG58'] = df_fila['Costo promedio del flete o tarifa.2']
        ws['AM58'] = df_fila['Promedio mensual e ingreso.2']
        ws['AT58'] = df_fila['Medio de transporte.2']
    
        ws['A59'] = df_fila['Elemento transportado.3']
        ws['I59'] = df_fila['Cantidad mensual.3']
        ws['M59'] = df_fila['Procedencia.3']
        ws['T59'] = df_fila['Destino.3']
        ws['AA59'] = df_fila['Frecuencia de movilización.3']
        ws['AG59'] = df_fila['Costo promedio del flete o tarifa.3']
        ws['AM59'] = df_fila['Promedio mensual e ingreso.3']
        ws['AT59'] = df_fila['Medio de transporte.3']

def llenarInforme8(ws, df_fila):
        ws['AO1'] = df_fila["Encuesta No."]

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AQ2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AN2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AQ2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[0] ####### LLENAR FECHA EN ESPACIOS VACÍOS Y NO SOBRE EL SÍMBOLO.
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')

        ws['AO3'] = df_fila["Encuestador"]
        ws['F7'] = df_fila["Nombre"]
        ws['Y7'] = df_fila["Empresa"]
        ws['AO7'] = df_fila["Cargo"]

        if df_fila["¿Pertenece a alguna asociación?"] == 'Si':
            ws['AB8'] = 'X'
            ws['AO8'] = df_fila["Otro, ¿Cuál?"]
        elif df_fila["¿Pertenece a alguna asociación?"] == 'No':
            ws['AD8'] = 'X'

        if df_fila["¿Tiene registro industrial y/o permiso ambiental?"] == 'Si':
            ws['I12'] = 'X'
        elif df_fila["¿Pertenece a alguna asociación?"] == 'No':
            ws['K12'] = 'X'
            ws['G13'] = df_fila['¿Cuál?.1']


        actividad = df_fila['Tipo de Cultivo']
        if actividad == 'Caucho':
            ws['J15'] = 'X'
        elif actividad == 'Palma africana':
            ws['J16'] = 'X'
        elif actividad == 'Acacia':
            ws['R15'] = 'X'
        elif actividad == 'Otros':
            ws['P16'] = df_fila['Otro, ¿Cuál?']


        ws['B19'] = df_fila["¿Con cuántos empleados cuenta la planta?"]

        actividad2 = df_fila["Vende principalmente en:"]
        if actividad2 == 'Sitio':
            ws['R22'] = 'X'
        elif actividad2 == 'Vereda':
            ws['R23'] = 'X'
        elif actividad2 == 'Casco Urbano':
            ws['R24'] = 'X'
        elif actividad2 == 'Otros Municipios y/o Veredas':
            ws['R25'] = 'X'
            ws['G26'] = df_fila["Otros, ¿Cuáles?"]
        
        actividad3 = df_fila["La planta obtiene el producto de:"]
        if actividad3 == 'Plantaciones propias':
            ws['AQ12'] = 'X'
        elif actividad3 == 'Venta del producto por parte de particulares':
            ws['AQ13'] = 'X'
        elif actividad3 == 'Otro':
            ws['AQ14'] = 'X'
            ws['AD15'] = df_fila["Otro, ¿Cuál?.1"]

        ws['AB19'] = df_fila["¿Cuál es el precio de venta?"]
        ws['AM19'] = df_fila["Unidad"]

        if df_fila["Sobre la actividad, piensa: Continuidad"] == "Continuar con la actividad":
            ws['AQ23'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Continuidad"] == "Finalizar la actividad":
            ws['AR23'] = 'X'
        if df_fila["Sobre la actividad, piensa: Producción"] == "Ampliar la producción":
            ws['AQ24'] = 'X'
            ws['AR25'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Permanecer con la misma producción":
            ws['AQ25'] = 'X'
            ws['AR24'] = 'X'
        elif df_fila["Sobre la actividad, piensa: Producción"] == "Ninguna de las anteriores":
            ws['AR24'] = 'X'
            ws['AR25'] = 'X'

        if str(df_fila['Hidrocarburos']) != "nan":
            ws['R28'] = 'X'
            ws['AD28'] = df_fila['Hidrocarburos']
        elif str(df_fila['Plantas de procesamiento']) != "nan":
            ws['R29'] = 'X'
            ws['AD29'] = df_fila['Plantas de procesamiento']
        elif str(df_fila['Distribuidores regionales']) != "nan":
            ws['R30'] = 'X'
        elif str(df_fila['Otro, ¿Cuál?.2']) != "nan":
            ws['AD30'] = df_fila['Otro, ¿Cuál?.2']

        ws['AK33'] = df_fila['¿Cuál es el área total cultivada? (Ha)']

        if df_fila['Unidad.1'] == "m2":
            ws['V34'] = 'X'
        elif df_fila['Unidad.1'] == "Ha":
            ws['Z34'] = 'X'
        elif df_fila['Unidad.1'] == "Cosecha":
            ws['AC34'] = 'X'

        ws['AL34'] = df_fila['Costo aproximado de establecimiento']

        if df_fila['Unidad.2'] == "m2":
            ws['V35'] = 'X'
        elif df_fila['Unidad.2'] == "Ha":
            ws['Z35'] = 'X'
        elif df_fila['Unidad.2'] == "Cosecha":
            ws['AC35'] = 'X'

        ws['AL35'] = df_fila['Costo aproximado de mantenimiento']

        if df_fila['Unidad.3'] == "m2":
            ws['V36'] = 'X'
        elif df_fila['Unidad.3'] == "Ha":
            ws['Z36'] = 'X'
        elif df_fila['Unidad.3'] == "Cosecha":
            ws['AC36'] = 'X'

        ws['AL36'] = df_fila['Costo aproximado de cosecha']

        ws['AK37'] = df_fila["Duración de cada ciclo de producción (Indicar la unidad)"]

        if df_fila['Unidad.4'] == "Tn":
            ws['V38'] = 'X'
        elif df_fila['Unidad.4'] == "Lt":
            ws['Z38'] = 'X'
        elif df_fila['Unidad.4'] == "Gn":
            ws['AD38'] = 'X'

        ws['AK38'] = df_fila['Volumen de producción por Ha']

        if df_fila['Unidad.5'] == "Tb":
            ws['V39'] = 'X'
        elif df_fila['Unidad.5'] == "Kg":
            ws['Z39'] = 'X'
        elif df_fila['Unidad.5'] == "Carga":
            ws['AD39'] = 'X'

        ws['AL39'] = df_fila['Precio de venta del producto']

        df_fila['Costo aproximado de establecimiento']

        ##### ABASTECIMIENTO DE INSUMOS #####

        ## INSUMO 1 - AGUA
        if df_fila['Agua'] == "Si":
            ws['N44'] = df_fila["Precio compra"]
            ws['Y44'] = df_fila["Cantidad (m3)"]
            ws['AG44'] = df_fila["Frecuencia de abastecimiento"]
            ws['AO44'] = df_fila["Procedencia de los insumos"]

        ## INSUMO 2 - COMBUSTIBLE
        if df_fila['Combustible'] == "Si":
            ws['N45'] = df_fila["Precio compra.1"]
            ws['Y45'] = df_fila["Cantidad (Gn)"]
            ws['AG45'] = df_fila["Frecuencia de abastecimiento.1"]
            ws['AO45'] = df_fila["Procedencia de los insumos.1"]

        ## INSUMO 3
        ws['N46'] = df_fila["Precio compra.2"]
        ws['Y46'] = df_fila["Cantidad"]
        ws['AG46'] = df_fila["Frecuencia de abastecimiento.2"]
        ws['AO46'] = df_fila["Procedencia de los insumos.2"]

        ## INSUMO 4
        ws['N47'] = df_fila["Precio compra.2"]
        ws['Y47'] = df_fila["Cantidad.1"]
        ws['AG47'] = df_fila["Frecuencia de abastecimiento.2"]
        ws['AO47'] = df_fila["Procedencia de los insumos.2"]


        actividad6 = df_fila['¿De dónde se abastece del recurso hídrico?']
        if actividad6 == 'Aljibe':
            ws['W48'] = 'X'
        elif actividad6 == 'Acueducto Veredal':
            ws['AE48'] = 'X'
        elif actividad6 == 'Otro':
            ws['AL48'] = 'X'
            ws['AQ48'] = df_fila['¿Cuál?.2']
    
        ws['W49'] = df_fila["Forma de extracción"]

        actividad7 = df_fila['¿Qué tipo de energía utiliza?']
        if actividad7 == 'Energía Eléctrica':
            ws['AA50'] = 'X'
        elif actividad7 == 'Energía Solar':
            ws['AJ50'] = 'X'
        elif actividad7 == 'Otro':
            ws['AP50'] = df_fila['¿Cuál?.3']

        if df_fila['¿Cuenta con servicio de alcantarillado?'] == "Si":
            ws['AB51'] = 'X'
        elif df_fila['¿Cuenta con servicio de alcantarillado?'] == "No":
            ws['AD51'] = 'X'

        ws['AO51'] = df_fila['¿Cuál?.4']
        ws['Y52'] = df_fila["¿Cuál es el manejo de aguas residuales y solidos?"]
        ws['Y53'] = df_fila["¿Cuál es el gasto aproximado de suministros en el proceso durante un mes?"]

        ##### ¿Qué tipo de equipos o maquinaría utiliza? #####

        ## EQUIPO 1
        ws['B56'] = df_fila["Equipo/maquinaria 1"]
        ws['N56'] = df_fila["Precio al que lo compró"]
        ws['X56'] = df_fila["Cantidad que posee la unidad económica"]
        ws['AF56'] = df_fila["Vida útil"]
        ws['AO56'] = df_fila["Procedencia"]

        ## EQUIPO 2
        ws['B57'] = df_fila["Equipo/maquinaria 2"]
        ws['N57'] = df_fila["Precio al que lo compró.1"]
        ws['X57'] = df_fila["Cantidad que posee la unidad económica.1"]
        ws['AF57'] = df_fila["Vida útil.1"]
        ws['AO57'] = df_fila["Procedencia.1"]

        ## EQUIPO 3
        ws['B58'] = df_fila["Equipo/maquinaria 3"]
        ws['N58'] = df_fila["Precio al que lo compró.2"]
        ws['X58'] = df_fila["Cantidad que posee la unidad económica.2"]
        ws['AF58'] = df_fila["Vida útil.2"]
        ws['AO58'] = df_fila["Procedencia.2"]

        ## EQUIPO 4
        ws['B59'] = df_fila["Equipo/maquinaria 4"]
        ws['N59'] = df_fila["Precio al que lo compró.3"]
        ws['X59'] = df_fila["Cantidad que posee la unidad económica.3"]
        ws['AF59'] = df_fila["Vida útil.3"]
        ws['AO59'] = df_fila["Procedencia.3"]

        ## EQUIPO 5
        ws['B60'] = df_fila["Equipo/maquinaria 5"]
        ws['N60'] = df_fila["Precio al que lo compró.4"]
        ws['X60'] = df_fila["Cantidad que posee la unidad económica.4"]
        ws['AF60'] = df_fila["Vida útil.4"]
        ws['AO60'] = df_fila["Procedencia.4"]


        ##### INFORMACIÓN LABORAL #####

        #### MANO DE OBRA CALIFICADA ####
        ws['I67'] = df_fila["#"]

        if df_fila["Género"] == "Femenino":
            ws['K66'] = 'X'
        elif df_fila["Género"] == "Masculino":
            ws['M66'] = 'X'

        if df_fila["Contrato"] == "Termino Fijo":
            ws['O66'] = df_fila["¿Cuánto?"]
        elif df_fila["Contrato"] == "Indefinido":
            ws['S66'] = 'X'

        ws['W66'] = df_fila["Jornal y turno laboral"]

        if df_fila["Escolaridad"] == "Primaria":
            ws['AB66'] = 'X'
        elif df_fila["Escolaridad"] == "Bachillerato":
            ws['AC66'] = 'X'
        elif df_fila["Escolaridad"] == "Técnico o tecnológico":
            ws['AD66'] = 'X'
        elif df_fila["Escolaridad"] == "Profesional":
            ws['AE66'] = 'X'
        elif df_fila["Escolaridad"] == "Posgrado":
            ws['AG66'] = 'X'
        
        if df_fila["Procedencia.5"] == "Vereda":
            ws['AJ66'] = 'X'
        elif df_fila["Procedencia.5"] == "Municipio":
            ws['AM66'] = 'X'
        elif df_fila["Procedencia.5"] == "Otro":
            ws['AO66'] = 'X'
        
        if df_fila["Residencia"] == "Vereda":
            ws['AQ66'] = 'X'
        elif df_fila["Residencia"] == "Municipio":
            ws['AS66'] = 'X'
        elif df_fila["Residencia"] == "Otro":
            ws['AU66'] = 'X'
        

        #### MANO DE OBRA NO CALIFICADA ####
        ws['I67'] = df_fila["#.1"]

        if df_fila["Género.1"] == "Femenino":
            ws['K67'] = 'X'
        elif df_fila["Género.1"] == "Masculino":
            ws['M67'] = 'X'

        if df_fila["Contrato.1"] == "Termino Fijo":
            ws['O67'] = df_fila["¿Cuánto?.1"]
        elif df_fila["Contrato.1"] == "Indefinido":
            ws['S67'] = 'X'

        ws['W67'] = df_fila["Jornal y turno laboral.1"]

        if df_fila["Escolaridad.1"] == "Primaria":
            ws['AB67'] = 'X'
        elif df_fila["Escolaridad.1"] == "Bachillerato":
            ws['AC67'] = 'X'
        elif df_fila["Escolaridad.1"] == "Técnico o tecnológico":
            ws['AD67'] = 'X'
        elif df_fila["Escolaridad.1"] == "Profesional":
            ws['AE67'] = 'X'
        elif df_fila["Escolaridad.1"] == "Posgrado":
            ws['AG67'] = 'X'
        
        if df_fila["Procedencia.6"] == "Vereda":
            ws['AJ67'] = 'X'
        elif df_fila["Procedencia.6"] == "Municipio":
            ws['AM67'] = 'X'
        elif df_fila["Procedencia.6"] == "Otro":
            ws['AO67'] = 'X'
        
        if df_fila["Residencia.1"] == "Vereda":
            ws['AQ67'] = 'X'
        elif df_fila["Residencia.1"] == "Municipio":
            ws['AS67'] = 'X'
        elif df_fila["Residencia.1"] == "Otro":
            ws['AU67'] = 'X'
        

        #### EMPLEADOS ADMINISTRATIVOS Y CONTABLES ####
        ws['I68'] = df_fila["#.2"]

        if df_fila["Género.2"] == "Femenino":
            ws['K68'] = 'X'
        elif df_fila["Género.2"] == "Masculino":
            ws['M68'] = 'X'

        if df_fila["Contrato.2"] == "Termino Fijo":
            ws['O68'] = df_fila["¿Cuánto?.2"]
        elif df_fila["Contrato.2"] == "Indefinido":
            ws['S68'] = 'X'

        ws['W68'] = df_fila["Jornal y turno laboral.2"]

        if df_fila["Escolaridad.2"] == "Primaria":
            ws['AB68'] = 'X'
        elif df_fila["Escolaridad.2"] == "Bachillerato":
            ws['AC68'] = 'X'
        elif df_fila["Escolaridad.2"] == "Técnico o tecnológico":
            ws['AD68'] = 'X'
        elif df_fila["Escolaridad.2"] == "Profesional":
            ws['AE68'] = 'X'
        elif df_fila["Escolaridad.2"] == "Posgrado":
            ws['AG68'] = 'X'
        
        if df_fila["Procedencia.7"] == "Vereda":
            ws['AJ68'] = 'X'
        elif df_fila["Procedencia.7"] == "Municipio":
            ws['AM68'] = 'X'
        elif df_fila["Procedencia.7"] == "Otro":
            ws['AO68'] = 'X'
        
        if df_fila["Residencia.2"] == "Vereda":
            ws['AQ68'] = 'X'
        elif df_fila["Residencia.2"] == "Municipio":
            ws['AS68'] = 'X'
        elif df_fila["Residencia.2"] == "Otro":
            ws['AU68'] = 'X'
        
        #### GERENTES Y DIRECTIVOS ####
        ws['I70'] = df_fila["#.3"]

        if df_fila["Género.3"] == "Femenino":
            ws['K70'] = 'X'
        elif df_fila["Género.3"] == "Masculino":
            ws['M70'] = 'X'
        
        if df_fila["Contrato.3"] == "Termino Fijo":
            ws['O70'] = df_fila["¿Cuánto?.3"]
        elif df_fila["Contrato.3"] == "Indefinido":
            ws['S70'] = 'X'
        
        ws['W70'] = df_fila["Jornal y turno laboral.3"]

        if df_fila["Escolaridad.3"] == "Primaria":
            ws['AB70'] = 'X'
        elif df_fila["Escolaridad.3"] == "Bachillerato":
            ws['AC70'] = 'X'
        elif df_fila["Escolaridad.3"] == "Técnico o tecnológico":
            ws['AD70'] = 'X'
        elif df_fila["Escolaridad.3"] == "Profesional":
            ws['AE70'] = 'X'
        elif df_fila["Escolaridad.3"] == "Posgrado":
            ws['AG70'] = 'X'
        
        if df_fila["Procedencia.8"] == "Vereda":
            ws['AJ70'] = 'X'
        elif df_fila["Procedencia.8"] == "Municipio":
            ws['AM70'] = 'X'
        elif df_fila["Procedencia.8"] == "Otro":
            ws['AO70'] = 'X'
        
        if df_fila["Residencia.3"] == "Vereda":
            ws['AQ70'] = 'X'
        elif df_fila["Residencia.3"] == "Municipio":
            ws['AS70'] = 'X'
        elif df_fila["Residencia.3"] == "Otro":
            ws['AU70'] = 'X'
        
        ##### OBRA O LABOR #####
        ## Obra o labor 1
        ws['A74'] = df_fila['Tipo de obra o labor 1']
        ws['K74'] = df_fila['Frecuencia de contratación/año']
        ws['R74'] = df_fila['Duración en Jornales del contrato']
        ws['AA74'] = df_fila['Valor del jornal']
        ws['AG74'] = df_fila['Cantidad de jornaleros empleados por contrato']
        ws['AO74'] = df_fila['Residencia de los jornaleros']

        ## Obra o labor 2
        ws['A75'] = df_fila['Tipo de obra o labor 2']
        ws['K75'] = df_fila['Frecuencia de contratación/año.1']
        ws['R75'] = df_fila['Duración en Jornales del contrato.1']
        ws['AA75'] = df_fila['Valor del jornal.1']
        ws['AG75'] = df_fila['Cantidad de jornaleros empleados por contrato.1']
        ws['AO75'] = df_fila['Residencia de los jornaleros.1']

        ## Obra o labor 3
        ws['A76'] = df_fila['Tipo de obra o labor 3']
        ws['K76'] = df_fila['Frecuencia de contratación/año.2']
        ws['R76'] = df_fila['Duración en Jornales del contrato.2']
        ws['AA76'] = df_fila['Valor del jornal.2']
        ws['AG76'] = df_fila['Cantidad de jornaleros empleados por contrato.2']
        ws['AO76'] = df_fila['Residencia de los jornaleros.2']


        if df_fila['Contrata servicios profesionales ("Sí" implica responder 70, 71, y 72)'] == "Si":
            ws['L78'] = 'X'


            if df_fila['¿Qué tipo de servicios?'] == "Contaduría":
                ws['AK78'] = 'X'
            elif df_fila['¿Qué tipo de servicios?'] == "Consultoría":
                ws['AT78'] = 'X'
            elif df_fila['¿Qué tipo de servicios?'] == "Asesoría legal":
                ws['AK79'] = 'X'
            elif df_fila['¿Qué tipo de servicios?'] == "Otros":
                ws['AT79'] = 'X'
                ws['AE80'] = df_fila['Otros, ¿Cuáles?.1']
            

            ##### SERVICIOS #####
            ws['A81'] = df_fila['Servicio 1']
            if df_fila['Frecuencia'] == "Mensual":
                ws['F81'] = 'X'
            elif df_fila['Frecuencia'] == "Semestral":
                ws['J81'] = 'X'
            elif df_fila['Frecuencia'] == "Trimestral":
                ws['P81'] = 'X'
            elif df_fila['Frecuencia'] == "Anual":
                ws['U81'] = 'X'

            ws['A82'] = df_fila['Servicio 2']
            if df_fila['Frecuencia.1'] == "Mensual":
                ws['F82'] = 'X'
            elif df_fila['Frecuencia.1'] == "Semestral":
                ws['J82'] = 'X'
            elif df_fila['Frecuencia.1'] == "Trimestral":
                ws['P82'] = 'X'
            elif df_fila['Frecuencia.1'] == "Anual":
                ws['U82'] = 'X'
            
            ws['A83'] = df_fila['Servicio 3']
            if df_fila['Frecuencia.2'] == "Mensual":
                ws['F83'] = 'X'
            elif df_fila['Frecuencia.2'] == "Semestral":
                ws['J83'] = 'X'
            elif df_fila['Frecuencia.2'] == "Trimestral":
                ws['P83'] = 'X'
            elif df_fila['Frecuencia.2'] == "Anual":
                ws['U83'] = 'X'

            
            ws['A84'] = df_fila['Servicio 4']
            if df_fila['Frecuencia.3'] == "Mensual":
                ws['F84'] = 'X'
            elif df_fila['Frecuencia.3'] == "Semestral":
                ws['J84'] = 'X'
            elif df_fila['Frecuencia.3'] == "Trimestral":
                ws['P84'] = 'X'
            elif df_fila['Frecuencia.3'] == "Anual":
                ws['U84'] = 'X'

            
            ws['A85'] = df_fila['Servicio 5']
            if df_fila['Frecuencia.4'] == "Mensual":
                ws['F85'] = 'X'
            elif df_fila['Frecuencia.4'] == "Semestral":
                ws['J85'] = 'X'
            elif df_fila['Frecuencia.4'] == "Trimestral":
                ws['P85'] = 'X'
            elif df_fila['Frecuencia.4'] == "Anual":
                ws['U85'] = 'X'


            ws['AE84'] = df_fila['¿Cuál es el monto pagado por estos servicios durante el último semestre?']

        elif df_fila['Contrata servicios profesionales ("Sí" implica responder 70, 71, y 72)'] == "No":
            ws['N78'] = 'X'

        ws['Z87'] = df_fila['Salarios pagados a la mano de obra calificada']
        ws['Z88'] = df_fila['Salarios pagados a la mano de obra no calificada']
        ws['Z89'] = df_fila['Salarios pagados a empleados y administrativos']
        ws['Z90'] = df_fila['Salarios pagados a gerentes y directivos']
        ws['Z91'] = df_fila['Total remuneraciones']

def llenarInforme9(ws, df_fila):
        ws['AN1'] = df_fila['Encuesta No.']

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AM2'] = re.findall('\d+',fecha_str.split("/")[2])[0]
                ws['AP2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[0]
            elif '-' in fecha_str:
                ws['AM2'] = re.findall('\d+',fecha_str.split("-")[2])[0]
                ws['AP2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[0]
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')  
        
        ws['AN3'] = df_fila['Encuestador']
        
        # A. Identificación del entrevistado
        ws['F7'] = df_fila['Nombre']
        ws['X7'] = df_fila['Empresa']
        ws['AN7'] = df_fila['Cargo']

        # ¿Pertenece a alguna asociación?
        asociacion = df_fila['¿Pertenece a alguna asociación?']
        if pd.notna(asociacion):
            if asociacion == 'Si':
                ws['AA8'] = 'X'
                ws['AN8'] = df_fila['¿Cuál?']
            elif asociacion == 'No':
                ws['AC8'] = 'X'
        else:
            print('Campo vacío')

        # B. INFORMACIÓN DE LA ACTIVIDAD
        # Minerales extraídos o explotados y/o transformados
        for i in range(6): 
            row_idx = 15 + i
            ws[f"K{row_idx}"] = valorCol('Unidad de medida', i, df_fila)
            ws[f"M{row_idx}"] = valorCol('Cantidad día', 2*i, df_fila)
            ws[f"O{row_idx}"] = valorCol('Cantidad Mes', 2*i, df_fila)
            ws[f"R{row_idx}"] = valorCol('Cantidad Año', 2*i, df_fila)
            ws[f"X{row_idx}"] = valorCol('Costo de producción/por unidad de medida', i, df_fila)
            ws[f"AE{row_idx}"] = valorCol('Cantidad día', 2*i+1, df_fila)
            ws[f"AG{row_idx}"] = valorCol('Cantidad Mes', 2*i+1, df_fila)
            ws[f"AK{row_idx}"] = valorCol('Cantidad Año', 2*i+1, df_fila)
            ws[f"AN{row_idx}"] = valorCol('Valor total de ventas según frecuencia', i, df_fila)


        tiene_calculo = df_fila['¿Tiene un cálculo aproximado del tiempo que puede seguir explotando el mineral?']
        if pd.notna(tiene_calculo):
            if tiene_calculo == 'Si':
                ws['O22'] = 'X'
                ws['AE22'] = df_fila['¿Cuánto?']
            elif tiene_calculo == 'No':
                ws['Q22'] = 'X'
        else:
            print('Campo vacío')

        
        objeto_explotacion = df_fila['¿Cuál es el objeto de la explotación?']
        if pd.notna(objeto_explotacion):
            if objeto_explotacion == 'Extracción de minerales': 
                ws['T23'] = 'X'
            elif objeto_explotacion == 'Transformación de minerales':
                ws['AE23'] = 'X'
            elif objeto_explotacion == 'Extracción y transformación de minerales':
                ws['AT23'] = 'X'
        else:
            print('Campo vacío')

        recurso_hidrico = df_fila['¿De dónde se abastece del recurso hídrico?']
        if pd.notna(recurso_hidrico):
            if recurso_hidrico == 'Aljibe': 
                ws['V24'] = 'X'
            elif recurso_hidrico == 'Acueducto veredal':
                ws['AD24'] = 'X'
            elif recurso_hidrico == 'Otro':
                ws['AK24'] = 'X'
                ws['AP24'] = df_fila['¿Cuál?.1']
        else:
            print('Campo vacío: ¿Cuál es el objeto de la explotación?')   

        ws['V25'] = df_fila['Forma de extracción']
        ws['AP25'] = df_fila['Cantidad estimada (m3)']

        # Tipo de energía que utiliza
        energia = df_fila['¿Qué tipo de energía utiliza?']
        if pd.notna(energia):
            if energia == 'Energía eléctrica':
                ws['AA26'] = 'X'
            elif energia == 'Energía solar':
                ws['AI26'] = 'X'
            elif energia == 'Otro':
                if pd.notna(df_fila['¿Cuál?.2']):
                    ws['AQ26'] = df_fila['¿Cuál?.2']
        else:
            print('Campo vacio')

        energia_coccion = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
        if pd.notna(energia_coccion):
            if energia_coccion == 'Energía eléctrica':
                ws['AA27'] = 'X'
            elif energia_coccion == 'Leña':
                ws['AE27'] = 'X'
            elif energia_coccion == 'Gas':
                ws['AK27'] = 'X'
            elif energia_coccion == 'Otro':
                if pd.notna(df_fila['¿Cuál?.3']):
                    ws['AQ27'] = df_fila['¿Cuál?.3']
        else:
            print('Campo vacío')

        alcantarillado = df_fila['¿Cuenta con servicio de alcantarillado?']
        if pd.notna(alcantarillado):
            if alcantarillado == 'Si':
                ws['AA28'] = 'X'
                ws['AN28'] = df_fila['¿Cuál?.4']
            elif alcantarillado == 'No':
                ws['AC28'] = 'X'

        ws['A30'] = df_fila['¿Cuál es el manejo de aguas residuales y solidos?']

        ws['Z31'] = df_fila['Sitio de venta']
        

        if pd.notna(df_fila['Hidrocarburos']):
            ws['Q33'] = 'X'
            ws['AC33'] = df_fila['¿Cuál? (Empresa y proyecto)']
        
        if pd.notna(df_fila['Otros']):
            ws['Q34'] = 'X'
            ws['AC34'] = df_fila['¿Cuáles?']


        ws['O36'] = df_fila['Valor total de la maquinaria']

        if pd.notna(df_fila['¿Cuenta con permisos ambientales?']):
            if df_fila['¿Cuenta con permisos ambientales?'] == 'Si':
                ws['Y37'] = 'X'
            elif df_fila['¿Cuenta con permisos ambientales?'] == 'No':
                ws['AA37'] = 'X'
        else:
            print('Campo vacío')   
        
        # Frecuencia de demanda de servicios

        ws['AB37'] = df_fila['Lugar de procedencia de la maquinaria']
        
        ws['L40'] = df_fila['Seguridad']

        ws['L41'] = df_fila['Mano de obra calificada']

        ws['L42'] = df_fila['Mano de obra no calificada']

        ws['L43'] = df_fila['Transporte']

        ws['AJ40'] = df_fila['Alojamiento']

        ws['AJ41'] = df_fila['Alimentación']

        ws['AF42'] = df_fila['Otro, ¿Cuál?']


        continuar = df_fila['¿Piensa continuar con la actividad?']
        if pd.notna(continuar):
            if continuar == 'Si':
                ws['M45'] = 'X'
            elif continuar == 'No':
                ws['O45'] = 'X'
        else:
            print('Campo vacio')

        produccion = df_fila['¿Piensa ampliar la producción?']
        if pd.notna(produccion):
            if produccion == 'Si':
                ws['AA45'] = 'X'
            elif produccion == 'No':
                ws['AC45'] = 'X'
        else:
            print('Campo vacio')


        permanecer = df_fila['¿Piensa permanecer con la misma producción?']
        if pd.notna(permanecer):
            if permanecer == 'Si':
                ws['AQ45'] = 'X'
            elif permanecer == 'No':
                ws['AS45'] = 'X'
        else:
            print('Campo vacio')

        # C. INFORMACIÓN LABORAL

        for i in range(10):
            prefijo_persona = 51 + i
            ws[f'E{prefijo_persona}'] = valorCol('Cargo', i+1, df_fila)
            ws[f'K{prefijo_persona}'] = valorCol('Edad (años)', i, df_fila)
            ws[f'L{prefijo_persona}'] = valorCol('Duración jornada (horas)', i, df_fila)

            manoObra = valorCol('Tipo de mano de obra', i, df_fila)
            if pd.notna(manoObra):
                if manoObra == 'Familiar':
                    ws[f'B{prefijo_persona}'] = 'X'
                elif manoObra == 'Contratado':
                    ws[f'D{prefijo_persona}'] = 'X'

            # Genero
            genero = valorCol('Género', i,df_fila)
            if pd.notna(genero):
                if genero == 'Masculino':
                    ws[f'J{prefijo_persona}20'] == 'X'
                elif genero ==  'Femenino':
                    ws[f'H{prefijo_persona}'] == 'X'

            # Escolaridad 
            escolaridad = valorCol('Escolaridad', i, df_fila)
            if pd.notna(escolaridad):
                if escolaridad:
                    if escolaridad == 'Primaria':
                        ws[f'N{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Bachillerato':
                        ws[f'P{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Técnico':
                        ws[f'R{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Pregrado':
                        ws[f'T{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Posgrado':
                        ws[f'V{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Contrato 
            contrato = valorCol('Contrato', i,df_fila)
            if contrato:
                if contrato == 'Tem.':
                    ws[f'AA{prefijo_persona}'] = valorCol('Contrato', i, df_fila)
                elif contrato == 'Fij':
                    ws[f'AC{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Pago de seguridad social 
            pago_seguridad = valorCol('Pago de seguridad social', i, df_fila)
            if pago_seguridad:
                if pago_seguridad == 'Si':
                    ws[f'AE{prefijo_persona}'] = 'X'
                elif pago_seguridad == 'No':
                    ws[f'AG{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Remuneración 
            remuneracion = valorCol('Remuneración', i, df_fila)
            if remuneracion:
                if remuneracion == 'Inferiores a $900.000':
                    ws[f'AS{prefijo_persona}'] = 'X'
                elif remuneracion == '$901.000 - $1.800.000':
                    ws[f'AT{prefijo_persona}'] = 'X'
                elif remuneracion == '$1.801.000 - $2.700.000':
                    ws[f'AU{prefijo_persona}'] = 'X'
                elif remuneracion == 'Superiores a $2.701.000':
                    ws[f'AV{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Información adicional
        
            ws[f'AH{prefijo_persona}'] = valorCol('Procedencia', i, df_fila)
            ws[f'AI{prefijo_persona}'] = valorCol('Residencia', i, df_fila)
            ws[f'AL{prefijo_persona}'] = valorCol('Tiempo trabajado', i, df_fila)
            ws[f'AM{prefijo_persona}'] = valorCol('# Personas núcleo familiar', i, df_fila)
            ws[f'AO{prefijo_persona}'] = valorCol('Personas a cargo', i, df_fila)
            ws[f'AP{prefijo_persona}'] = valorCol('Lugar de residencia familiar', i,df_fila)
