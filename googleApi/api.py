from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
from func.llenarPlantillas import llenarInforme1, llenarFichaPredial, llenarUsosUsuarios, llenarFormatoAgropecuario, llenarFormatoComercial, llenarFormatoServicios, llenarActEconomica
from util.Pdf import Pdf

class GoogleSheetsAExcel:
    def __init__(
        self,
        service_account_file: str,
        spreadsheet_id: str,
        drive_folder_id: str,
        range_informe1: str = None,
        range_ficha1: str = None,
        range_ficha2: str = None,

        # Rangos para el Formato UU
        range_usos1: str = None,
        range_usos2: str = None,

        # Rangos para el Formato Agropecuario
        range_formato_agro: str = None,         
        range_info_comercial: str = None,
        range_explot_avicola: str = None,
        range_info_laboral: str = None,
        range_explot_agricola: str = None,
        range_explot_porcina: str = None,
        range_detalle_jornal: str = None,

        # Rangos Formato Comercial

        range_formato_comercial: str = None,
        range_descripcion_abastecimiento: str = None,
        range_descripcion_actividad_precio: str = None,
        range_info_laboral2: str = None,

        # Rangos para el Formato Servicios
        range_formato_servicios: str = None, # Rango principal Sheet1
        range_desc_actividad_precio_servicios: str = None, # data-desc_actividad-precio_servicios
        range_insumos_abastecimiento_servicios: str = None, # data-begin_insumos-abastecimiento_insumos
        range_desc_actividad_servicios: str = None, # data-desc_actividad-num_servicios
        range_equipos_maquinaria_servicios: str = None, # data-equipos_maquinaria
        range_info_laboral_servicios: str = None, # data-informacion_laboral

        # Rangos para encuesta de identificacion de actividade económica
        range_identificacion_actividad: str = None, # Sheet1

        # Plantillas
        plantilla_informe1: str = None,
        plantilla_ficha: str = None,
        plantilla_usos_usuarios: str = None,
        plantilla_formato_agro: str = None,
        plantilla_formato_comercial: str = None,
        plantilla_formato_servicios: str = None,
        plantilla_identificacion_actividad: str = None
    ) -> None:
        """
        Constructor de la clase GoogleSheetsAExcel.
        """
        self.service_account_file = service_account_file
        self.spreadsheet_id = spreadsheet_id
        self.drive_folder_id = drive_folder_id

        # Rangos anteriores
        self.range_informe1 = range_informe1
        self.range_ficha1 = range_ficha1
        self.range_ficha2 = range_ficha2
        self.range_usos1 = range_usos1
        self.range_usos2 = range_usos2

        # Rangos para el Formato Agropecuario
        self.range_formato_agro = range_formato_agro
        self.range_info_comercial = range_info_comercial
        self.range_explot_avicola = range_explot_avicola
        self.range_info_laboral = range_info_laboral
        self.range_explot_agricola = range_explot_agricola
        self.range_explot_porcina = range_explot_porcina
        self.range_detalle_jornal = range_detalle_jornal

        # Rangos para el Formato Comercial
        self.range_formato_comercial = range_formato_comercial
        self.range_descripcion_abastecimiento = range_descripcion_abastecimiento
        self.range_descripcion_actividad_precio = range_descripcion_actividad_precio
        self.range_info_laboral2 = range_info_laboral2

        # Rangos para el Formato Servicios
        self.range_formato_servicios = range_formato_servicios
        self.range_desc_actividad_precio_servicios = range_desc_actividad_precio_servicios
        self.range_insumos_abastecimiento_servicios = range_insumos_abastecimiento_servicios
        self.range_desc_actividad_servicios = range_desc_actividad_servicios
        self.range_equipos_maquinaria_servicios = range_equipos_maquinaria_servicios
        self.range_info_laboral_servicios = range_info_laboral_servicios

        # Rango para la encuesta de identificación de actividad económica

        self.range_identificacion_actividad = range_identificacion_actividad

        # Plantillas FP, Id. AE, Usos/Usuarios
        self.plantilla_informe1 = plantilla_informe1
        self.plantilla_ficha = plantilla_ficha
        self.plantilla_usos_usuarios = plantilla_usos_usuarios

        # Plantilla Formato Agropecuario
        self.plantilla_formato_agro = plantilla_formato_agro

        # Plantillas para el Formato Comercial
        self.plantilla_formato_comercial = plantilla_formato_comercial

        # Plantillas para el Formato Servicios
        self.plantilla_formato_servicios = plantilla_formato_servicios

        # Plantilla para la encuesta de identificación de actividad económica
        self.plantilla_identificacion_actividad = plantilla_identificacion_actividad


              
        self.scopes = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive'
        ]
        self.credentials = None
        self.sheet_service = None
        self.drive_service = None


    def inicializarServicios(self):
        """
        Inicializa el cliente de Sheets y Drive con las credenciales.
        """
        self.credentials = service_account.Credentials.from_service_account_file(
            self.service_account_file, scopes=self.scopes
        )
        self.sheet_service = build('sheets', 'v4', credentials=self.credentials)
        self.drive_service = build('drive', 'v3', credentials=self.credentials)

    
    def fetchDatos(self, rango: str) -> pd.DataFrame:
        """
        Retorna un DataFrame con los datos del rango especificado en la hoja de cálculo.
        """
        try:
            result = self.sheet_service.spreadsheets().values().get(
                spreadsheetId=self.spreadsheet_id,
                range=rango
            ).execute()

            values = result.get('values', [])
            if not values:
                raise ValueError(f"No se encontraron datos en el rango '{rango}'")

            columnas = values[0]
            df = pd.DataFrame(values[1:], columns=columnas)
            if 'data-fecha' in df.columns:
                df['data-fecha'] = pd.to_datetime(df['data-fecha'], errors='coerce')
            
            return df

        except Exception as e:
            return pd.DataFrame() 

    def obtenerOCrearCarpetaPorCodigo(self, codigo: str) -> str:
        """
        Verifica si existe una carpeta en Drive con nombre = 'codigo' (dentro de self.drive_folder_id).
        Si no existe, la crea. Retorna el folder_id de esa carpeta.
        """
        query = (
            f"'{self.drive_folder_id}' in parents and "
            f"name = '{codigo}' and "
            "mimeType = 'application/vnd.google-apps.folder'"
        )
        respuesta = self.drive_service.files().list(q=query).execute()
        archivos = respuesta.get('files', [])

        if archivos:
            carpeta_id = archivos[0]['id']
            print(f"[+] Carpeta '{codigo}' encontrada (id: {carpeta_id}).")
        else:
            metadata = {
                'name': codigo,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [self.drive_folder_id]
            }
            carpeta = self.drive_service.files().create(body=metadata, fields='id').execute()
            carpeta_id = carpeta.get('id')
            print(f"[+] Carpeta '{codigo}' creada (id: {carpeta_id}).")

        return carpeta_id

    def subirArchivo(self, file_path: str, nombre_archivo: str, folder_id: str):
        """
        Sube un archivo a la carpeta 'folder_id' en Drive,
        usando el MIME correcto según sea PDF o Excel.
        """
        # Detectar la extensión
        extension = os.path.splitext(file_path)[1].lower()
        if extension == '.pdf':
            mime_type = 'application/pdf'
        else:
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        archivo_metadata = {
            'name': nombre_archivo,
            'parents': [folder_id]
        }
        media = MediaFileUpload(file_path, mimetype=mime_type)
        self.drive_service.files().create(body=archivo_metadata, media_body=media, fields='id').execute()
        print(f"[OK] Subido '{nombre_archivo}' a carpeta (id: {folder_id}).")

    def archivoExiste(self, nombre_archivo: str, folder_id: str) -> bool:
        """
        Verifica si un archivo PDF de nombre 'nombre_archivo' ya existe en la carpeta 'folder_id'.
        """
        query = (
            f"'{folder_id}' in parents and "
            f"name = '{nombre_archivo}' and "
            "mimeType = 'application/pdf'"
        )
        respuesta = self.drive_service.files().list(q=query, fields="files(id, name)").execute()
        archivos = respuesta.get('files', [])
        return len(archivos) > 0

    # -------------------------------------------------------------------------
    #  ! Métodos para llenar las encuestas y subirlas al Drive.
    # -------------------------------------------------------------------------
    def llenarYSubirInforme1(self):
        """
        Lee el rango 'range_informe1'. Por cada fila, toma 'data-info_general-num_encuesta' 
        como 'codigo', crea la subcarpeta en Drive (si no existe) y sube un PDF 
        con nombre <codigo>_informe1.pdf.
        """
        if not self.range_informe1 or not self.plantilla_informe1:
            print("No están configurados 'range_informe1' o 'plantilla_informe1'.")
            return

        df_informe = self.fetchDatos(self.range_informe1)
        pdfConv = Pdf()

        for _, fila in df_informe.iterrows():
            # 1) Tomar el codigo
            codigo = str(fila['data-info_general-num_encuesta'])

            # 2) Crear/obtener carpeta
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 3) Nombre del PDF
            nombre_pdf = f"{codigo}_informe1.pdf"

            # 4) Verificar si ya existe
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 5) Llenar plantilla
            wb = load_workbook(self.plantilla_informe1)
            ws = wb.active
            llenarInforme1(ws, fila)

            # 6) Guardar Excel local
            nombre_excel = f"{codigo}_informe1.xlsx"
            wb.save(nombre_excel)

            # 7) Convertir a PDF (1 sola página, Portrait, A4)
            ruta_pdf = f"{codigo}_informe1.pdf"
            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,           
                orientation=1,             
                paper_size=9,              
                fit_to_pages_wide=1,      
                fit_to_pages_tall=1,       
                zoom=None,                 
                center_horizontally=True,
                center_vertically=True
            )

            # 8) Subir y limpiar
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)
            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")


    def llenarYSubirIdentificacionActEconomica(self):
        """
        Lee el rango 'range_informe1'. Por cada fila, toma 'data-info_general-num_encuesta' 
        como 'codigo', crea la subcarpeta en Drive (si no existe) y sube un PDF 
        con nombre <codigo>_informe1.pdf.
        """
        if not self.range_identificacion_actividad or not self.plantilla_identificacion_actividad:
            print("No están configurados 'range_identificacion_actividad' o 'plantilla_informe1'.")
            return

        df_informe = self.fetchDatos(self.range_identificacion_actividad)
        pdfConv = Pdf()

        for _, fila in df_informe.iterrows():
            # 1) Tomar el codigo
            codigo = str(fila['data-id_encuesta'])

            # 2) Crear/obtener carpeta
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 3) Nombre del PDF
            nombre_pdf = f"{codigo}_informe1.pdf"

            # 4) Verificar si ya existe
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 5) Llenar plantilla
            wb = load_workbook(self.plantilla_identificacion_actividad)
            ws = wb.active
            llenarActEconomica(ws, fila)

            # 6) Guardar Excel local
            nombre_excel = f"{codigo}_informe1.xlsx"
            wb.save(nombre_excel)

            # 7) Convertir a PDF (1 sola página, Portrait, A4)
            ruta_pdf = f"{codigo}_informe1.pdf"
            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,           
                orientation=1,             
                paper_size=9,              
                fit_to_pages_wide=1,      
                fit_to_pages_tall=1,       
                zoom=None,                 
                center_horizontally=True,
                center_vertically=True
            )

            # 8) Subir y limpiar
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)
            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")


    def llenarYSubirFichaPredial(self):
        """
        Lee 'df_ficha1' y 'df_ficha2' (range_ficha1, range_ficha2).
        Para cada fila de df_ficha1:
        - Saca el codigo (columna data-info_general-num_encuesta) y la KEY.
        - Filtra df_ficha2 con PARENT_KEY == KEY para tener un subconjunto.
        - Llama a 'llenarFichaPredial(ws, row_ficha1, subset_ficha2)', 
            donde subset_ficha2 llena la tabla en la misma hoja.
        - Genera un PDF (hasta 3 páginas de alto) y lo sube a la carpeta Drive de 'codigo'.
        """
        if not self.range_ficha1 or not self.range_ficha2 or not self.plantilla_ficha:
            print("No están configurados 'range_ficha1', 'range_ficha2' o 'plantilla_ficha'.")
            return

        df_ficha1 = self.fetchDatos(self.range_ficha1)
        df_ficha2 = self.fetchDatos(self.range_ficha2)
        pdfConv = Pdf()

        for idx, row_ficha1 in df_ficha1.iterrows():
            codigo = str(row_ficha1['data-info_general-num_encuesta'])
            key = row_ficha1['KEY']

            # Filtramos df_ficha2
            subset_ficha2 = df_ficha2[df_ficha2['PARENT_KEY'] == key]
            if subset_ficha2.empty:
                print(f"No hay sub-filas en df_ficha2 para KEY='{key}'. Omitiendo...")
                continue

            # Crear/obtener la carpeta
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # Construir nombre PDF
            nombre_pdf = f"{codigo}_fichaPredial.pdf"
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # Llenar la plantilla
            wb = load_workbook(self.plantilla_ficha)
            ws = wb.active

            llenarFichaPredial(ws, row_ficha1, subset_ficha2, self.drive_service)

            # Guardar Excel local
            nombre_excel = f"{codigo}_fichaPredial.xlsx"
            wb.save(nombre_excel)

            # Convertir a PDF
            ruta_pdf = f"{codigo}_fichaPredial.pdf"

            # Ejemplo: ajusta a 1 página de ancho, hasta 3 páginas de alto (FitToPagesTall=3),
            # en orientación vertical (1 = Portrait), tamaño A4 (9).
            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,          # Usa la hoja activa (ws.active)
                orientation=1,            # 1 = Portrait, 2 = Landscape
                paper_size=9,             # 9 = A4, 2 = Letter, etc.
                fit_to_pages_wide=1,      # Ajustar el ancho a 1 página
                fit_to_pages_tall=3,      # Permite hasta 3 páginas de alto
                zoom=None,                # None => usar FitToPages
                center_horizontally=True,
                center_vertically=True
            )

            # Subir a Drive y limpiar
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)
            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")


    def llenarYSubirUsosUsuarios(self):
        """
        Lee dos rangos self.range_usos1 y self.range_usos2
        range_usos1 produce un df_usos1 con la columna KEY.
        range_usos2 produce un df_usos2 con la columna PARENT_KEY.
        
        Por cada fila de df_usos1:
        - obtiene código data-info_general-num_encuesta,
        - filtra df_usos2 en base a PARENT_KEY == KEY,
        - llama a llenarUsosUsuarios(ws, fila_principal, subset_usos2, ...),
        - genera un PDF y lo sube a Drive.
        """
        # 1) Validar que tengas definidos los rangos y la plantilla
        if not self.range_usos1 or not self.range_usos2 or not self.plantilla_usos_usuarios:
            print("No están configurados los rangos o la plantilla de usos/usuarios.")
            return

        # 2) Leer los dos DataFrames
        df_usos1 = self.fetchDatos(self.range_usos1)  # contiene la columna 'KEY'
        df_usos2 = self.fetchDatos(self.range_usos2)  # contiene la columna 'PARENT_KEY'

        pdfConv = Pdf()

        # 3) Iterar sobre df_usos1
        for idx, row_usos1 in df_usos1.iterrows():
            # 3a) Tomar código y KEY
            codigo = str(row_usos1['data-info_general-num_encuesta'])
            key = row_usos1['KEY'] 

            # 3b) Filtrar df_usos2 usando PARENT_KEY == key
            subset_usos2 = df_usos2[df_usos2['PARENT_KEY'] == key]
            if subset_usos2.empty:
                print(f"No hay subfilas en df_usos2 para KEY='{key}'. Omitiendo...")
                continue

            # 4) Crear/obtener carpeta en Drive
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 5) Construir nombre PDF y verificar si ya existe
            nombre_pdf = f"{codigo}_usosUsuarios.pdf"
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 6) Cargar la plantilla y llenar
            wb = load_workbook(self.plantilla_usos_usuarios)
            ws = wb.active

            # Llamamos a tu función de llenado 
            llenarUsosUsuarios(ws, row_usos1, subset_usos2, self.drive_service)

            # 7) Guardar Excel temporal
            nombre_excel = f"{codigo}_usosUsuarios.xlsx"
            wb.save(nombre_excel)

            # 8) Convertir a PDF 
            #    - Landscape (orientation=2)
            #    - A4 (paper_size=9)
            #    - Ancho en 1 página, alto en máx. 2 páginas
            ruta_pdf = f"{codigo}_usosUsuarios.pdf"
            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,
                orientation=2,       # 2 = Landscape
                paper_size=9,        # 9 = A4
                fit_to_pages_wide=1, # Ajustar ancho a 1 página
                fit_to_pages_tall=2, # Hasta 2 páginas de alto
                zoom=None,           # None => usar FitToPages
                center_horizontally=True,
                center_vertically=True
            )

            # 9) Subir el PDF y limpiar
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)

            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")



    def llenarYSubirFormatoAgropecuario(self):
        """
        Lee el rango 'range_formato_agro' como la tabla principal,
        donde cada fila tiene 'KEY' y 'data-info_general-num_encuesta' (código).
        Luego filtra los DataFrames secundarios (info_comercial, avícola, laboral,
        agrícola, porcina, jornal) usando 'PARENT_KEY == KEY'.
        Llama a 'llenarFormatoAgropecuario' y sube el PDF resultante a Drive,
        generando hasta 4 páginas en modo Landscape (o lo que configures).
        """
        # 1) Validar que estén configurados el rango principal y la plantilla
        if not self.range_formato_agro or not self.plantilla_formato_agro:
            print("No están configurados 'range_formato_agro' o 'plantilla_formato_agro'.")
            return

        # 2) Cargar DF principal
        df_principal = self.fetchDatos(self.range_formato_agro)

        # 3) Cargar DFs secundarios
        df_info_com = self.fetchDatos(self.range_info_comercial)
        df_avicola = self.fetchDatos(self.range_explot_avicola)
        df_laboral = self.fetchDatos(self.range_info_laboral)
        df_agricola = self.fetchDatos(self.range_explot_agricola)
        df_porcina = self.fetchDatos(self.range_explot_porcina)
        df_jornal = self.fetchDatos(self.range_detalle_jornal)

        pdfConv = Pdf()

        # 4) Iterar sobre cada fila del DF principal
        for idx, df_fila in df_principal.iterrows():
            # Tomamos el código y KEY
            codigo = str(df_fila['data-datos_encuesta-num_encuesta'])
            key = df_fila['KEY']

            # 4a) Filtrar cada DF secundario por 'PARENT_KEY' == key
            subset_info_com = df_info_com[df_info_com['PARENT_KEY'] == key]
            subset_avicola  = df_avicola[df_avicola['PARENT_KEY'] == key]
            subset_laboral  = df_laboral[df_laboral['PARENT_KEY'] == key]
            subset_agricola = df_agricola[df_agricola['PARENT_KEY'] == key]
            subset_porcina  = df_porcina[df_porcina['PARENT_KEY'] == key]
            subset_jornal   = df_jornal[df_jornal['PARENT_KEY'] == key]

            # 5) Crear/obtener la carpeta en Drive
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 6) Definir nombre del PDF y chequear si existe
            nombre_pdf = f"{codigo}_formatoAgropecuario.pdf"
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 7) Cargar plantilla de Excel
            wb = load_workbook(self.plantilla_formato_agro)
            # Asume que la hoja que te interesa es la activa
            # o si tienes un nombre de hoja en particular:
            # ws = wb["FORMATO2"] 
            ws = wb.active

            # 8) Llamar función de llenado
            llenarFormatoAgropecuario(
                ws, 
                df_fila,
                subset_info_com,
                subset_avicola,
                subset_laboral,
                subset_agricola,
                subset_porcina,
                subset_jornal
            )

            # 9) Guardar Excel temporal
            nombre_excel = f"{codigo}_formatoAgropecuario.xlsx"
            wb.save(nombre_excel)

            # 9b) Convertir a PDF
            ruta_pdf = f"{codigo}_formatoAgropecuario.pdf"
            

            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,           # Usa la hoja activa
                orientation=2,             # 2 = Landscape
                paper_size=9,              # 9 = A4
                fit_to_pages_wide=1,       # 1 => ajusta ancho a 1 pág
                fit_to_pages_tall=4,       # hasta 4 páginas de alto
                zoom=None,                 # None => usa FitToPages
                center_horizontally=True,
                center_vertically=True
            )

            # 10) Subir PDF y limpiar archivos locales
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)

            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")

    def llenarYSubirFormatoComercial(self):
        """
        Lee el rango 'rango_formato_comercial' como la tabla principal,
        Luego filtra los DataFrames secundarios (descripcion_abastecimiento, descripcion_actividad_precio, info_laboral2) usando 'PARENT_KEY == KEY'.
        Llama a 'llenarFormatoComercial' y sube el PDF resultante a Drive.
        """
        # 1) Validar que estén configurados el rango principal y la plantilla
        if not self.range_formato_comercial or not self.plantilla_formato_comercial:
            print("No están configurados 'rango_formato_comercial' o 'plantilla_formato_comercial'.")
            return

        # 2) Cargar DF principal
        df_principal = self.fetchDatos(self.range_formato_comercial)

        # 3) Cargar DFs secundarios
        df_descripcion_abastecimiento = self.fetchDatos(self.range_descripcion_abastecimiento)
        df_descripcion_actividad_precio = self.fetchDatos(self.range_descripcion_actividad_precio)
        df_info_laboral = self.fetchDatos(self.range_info_laboral2)

        pdfConv = Pdf()

        for idx, df_fila in df_principal.iterrows():
            # Tomamos el código y KEY
            codigo = str(df_fila['data-datos_encuesta-num_encuesta'])
            key = df_fila['KEY']

             # 4a) Filtrar cada DF secundario por 'PARENT_KEY' == key
            subset_descripcion_abastecimiento = df_descripcion_abastecimiento[df_descripcion_abastecimiento['PARENT_KEY'] == key]
            subset_descripcion_actividad_precio = df_descripcion_actividad_precio[df_descripcion_actividad_precio['PARENT_KEY'] == key]
            subset_info_laboral = df_info_laboral[df_info_laboral['PARENT_KEY'] == key]

            # 5) Crear/obtener la carpeta en Drive
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 6) Definir nombre del PDF y chequear si existe
            nombre_pdf = f"{codigo}_formatoComercial.pdf"
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 7) Cargar plantilla de Excel
            wb = load_workbook(self.plantilla_formato_comercial)
            # Asume que la hoja que te interesa es la activa
            ws = wb.active

            # 8) Llamar función de llenado
            llenarFormatoComercial(
                ws,
                df_fila,
                subset_descripcion_abastecimiento,
                subset_descripcion_actividad_precio,
                subset_info_laboral
            )

            # 9) Guardar Excel temporal
            nombre_excel = f"{codigo}_formatoComercial.xlsx"
            wb.save(nombre_excel)

            # 9b) Convertir a PDF
            ruta_pdf = f"{codigo}_formatoComercial.pdf"

            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,           
                orientation=2,          
                paper_size=9,              
                fit_to_pages_wide=1,      
                fit_to_pages_tall=1,       
                zoom=None,                 
                center_horizontally=True,
                center_vertically=True
            )

            # 10) Subir PDF y limpiar archivos locales
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)

            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")


    def llenarYSubirFormatoServicios(self):
        """
        Lee el rango 'range_formato_servicios' como la tabla principal,
        Luego filtra los DataFrames secundarios (descripcion_actividad_precio, insumos_abastecimiento, desc_actividad, equipos_maquinaria, info_laboral) usando 'PARENT_KEY == KEY'.
        Llama a 'llenarFormatoServicios' y sube el PDF resultante a Drive.
        """
        # 1) Validar que estén configurados el rango principal y la plantilla
        if not self.range_formato_servicios or not self.plantilla_formato_servicios:
            print("No están configurados 'range_formato_servicios' o 'plantilla_formato_servicios'.")
            return

        # 2) Cargar DF principal
        df_principal = self.fetchDatos(self.range_formato_servicios)

                # range_formato_servicios: str = None,
        # range_desc_actividad_precio_servicios: str = None,
        # range_insumos_abastecimiento_servicios: str = None,
        # range_desc_actividad_servicios: str = None,
        # range_equipos_maquinaria_servicios: str = None,
        # range_info_laboral_servicios: str = None,

        # 3) Cargar DFs secundarios
        df_descripcion_actividad_precio = self.fetchDatos(self.range_desc_actividad_precio_servicios)
        df_insumos_abastecimiento = self.fetchDatos(self.range_insumos_abastecimiento_servicios)
        df_desc_actividad = self.fetchDatos(self.range_desc_actividad_servicios)
        df_equipos_maquinaria = self.fetchDatos(self.range_equipos_maquinaria_servicios)
        df_info_laboral = self.fetchDatos(self.range_info_laboral_servicios)

        pdfConv = Pdf()

        # 4) Iterar sobre cada fila del DF principal
        for idx, df_fila in df_principal.iterrows():
            # Tomamos el código y KEY
            codigo = str(df_fila['data-datos_encuesta-num_encuesta'])
            key = df_fila['KEY']
            

            # 4a) Filtrar cada DF secundario por 'PARENT_KEY' == key
            subset_descripcion_actividad_precio = df_descripcion_actividad_precio[df_descripcion_actividad_precio['PARENT_KEY'] == key]
            subset_insumos_abastecimiento = df_insumos_abastecimiento[df_insumos_abastecimiento['PARENT_KEY'] == key]
            subset_desc_actividad = df_desc_actividad[df_desc_actividad['PARENT_KEY'] == key]
            subset_equipos_maquinaria = df_equipos_maquinaria[df_equipos_maquinaria['PARENT_KEY'] == key]
            if not df_info_laboral.empty:
                subset_info_laboral = df_info_laboral[df_info_laboral['PARENT_KEY'] == key]
            else:
                subset_info_laboral = pd.DataFrame()

            # 5) Crear/obtener la carpeta en Drive
            folder_id = self.obtenerOCrearCarpetaPorCodigo(codigo)

            # 6) Definir nombre del PDF y chequear si existe
            nombre_pdf = f"{codigo}_formatoServicios.pdf"
            if self.archivoExiste(nombre_pdf, folder_id):
                print(f"El archivo {nombre_pdf} ya existe en '{codigo}'. Omitiendo...")
                continue

            # 7) Cargar plantilla de Excel
            wb = load_workbook(self.plantilla_formato_servicios)
            # Asume que la hoja que te interesa es la activa
            ws = wb.active

            # 8) Llamar función de llenado
            llenarFormatoServicios(
                ws,
                df_fila,
                subset_descripcion_actividad_precio,
                subset_insumos_abastecimiento,
                subset_desc_actividad,
                subset_equipos_maquinaria,
                subset_info_laboral
            )

            # 9) Guardar Excel temporal
            nombre_excel = f"{codigo}_formatoServicios.xlsx"
            wb.save(nombre_excel)

            # 9b) Convertir a PDF
            ruta_pdf = f"{codigo}_formatoServicios.pdf"

            pdfConv.excelPdf(
                excel_path=nombre_excel,
                pdf_path=ruta_pdf,
                sheet_name=None,
                orientation=2,
                paper_size=9,
                fit_to_pages_wide=1,
                fit_to_pages_tall=1,
                zoom=None,
                center_horizontally=True,
                center_vertically=True
            )

            # 10) Subir PDF y limpiar archivos locales
            self.subirArchivo(ruta_pdf, nombre_pdf, folder_id)
            os.remove(nombre_excel)
            os.remove(ruta_pdf)

            print(f"[OK] Se generó y subió {nombre_pdf} en la carpeta '{codigo}'.")

