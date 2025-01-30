import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime


def llenarInforme1(ws, df_fila):
    ws['Y1'] = df_fila['data-info_general-num_encuesta']
    if pd.notna(df_fila['data-info_general-fecha']):
        fecha_valor = df_fila['data-info_general-fecha']
        ws['X2'] = str(fecha_valor.day)
        ws['Z2'] = str(fecha_valor.month)
        ws['AD2'] = str(fecha_valor.year)
    else:
        print("Campo de fecha vacío o inválido.")

    ws['W3'] = df_fila['data-start_act_economica-encuestador']
    ws['A8'] = df_fila['data-info_general-departamento']
    ws['N8'] = df_fila['data-info_general-municipio']
    ws['U8'] = df_fila['data-info_general-vereda']
    ws['A10'] = df_fila['data-start_act_economica-coordenadas']

    if df_fila['data-start_act_economica-permite_entrevista'] == 'yes':
        ws['W10'] = 'X'

        ws['G14'] = df_fila['data-start_act_economica-start_act_economica-nombre_establecimiento']
        ws['D15'] = df_fila['data-start_act_economica-direccion']
        ws['U15'] = df_fila['data-start_act_economica-telefono_contacto']
        ws['G16'] = df_fila['data-start_act_economica-actividad_economica']
        ws['W16'] = df_fila['data-start_act_economica-inicio_actividad']
        ws['D17'] = df_fila['data-start_act_economica-propietario']
        ws['Q17'] = df_fila['data-start_act_economica-procedencia_propietario']
        ws['Y17'] = df_fila['data-start_act_economica-lugar_residencia']
        ws['D18'] = df_fila['data-start_act_economica-administrador']
        ws['Q18'] = df_fila['data-start_act_economica-procedencia_administrador']
        ws['Z18'] = df_fila['data-start_act_economica-lugar_residencia_admin']

        actividad = df_fila['data-start_act_economica-actividad_como']
        mapeo_actividad = {
            'natural': 'G23',
            'sociedad_hecho': 'N23',
            'unipersonal': 'G24',
            'sociedad_comercial': 'N24',
            'cooperativa': 'G25',
            'predio': 'G26',
            'other': 'N25'
        }
        if actividad in mapeo_actividad:
            ws[mapeo_actividad[actividad]] = 'X'
            if actividad == 'other':
                ws['K26'] = df_fila['data-start_act_economica-tenencia_propiedad_other']

        actividad2 = df_fila['data-start_act_economica-tipo_actividad']
        mapeo_tipo_actividad = {
            'agricola': 'G30',
            'pecuaria': 'G31',
            'agroindustrial': 'G32',
            'servicios': 'G33',
            'comercial': 'M31',
            'manufactura': 'M32',
            'transporte': 'M33'
        }
        if actividad2 in mapeo_tipo_actividad:
            ws[mapeo_tipo_actividad[actividad2]] = 'X'

        ws['B37'] = df_fila['data-start_act_economica-producto_principal']

        tenencia = df_fila['data-start_act_economica-tenencia_propiedad']
        mapeo_tenencia = {
            'propia': 'E40',
            'administrada': 'E41',
            'arrendada': 'E42',
            'other': 'E44'
        }
        if tenencia in mapeo_tenencia:
            ws[mapeo_tenencia[tenencia]] = 'X'
            if tenencia == 'arrendada':
                ws['J41'] = df_fila['data-start_act_economica-canon_arrendamiento']
            if tenencia == 'other':
                ws['C45'] = df_fila['data-start_act_economica-tenencia_propiedad_other']

        ws['A48'] = df_fila['data-start_act_economica-actividad_ingresos']

        ingresos = df_fila['data-start_act_economica-ingresos']
        mapeo_ingresos = {
            'inferior_600k': 'Y23',
            'entre_601k_1500k' : 'Y24',
            'entre_1501k_3000k' : 'Y25',
            'superior_3000k' : 'Y26',
            'other' : 'AD24'
        }
        if ingresos in mapeo_ingresos:
            ws[mapeo_ingresos[ingresos]] = 'X'
            if ingresos == 'other':
                ws['AA25'] = df_fila['data-start_act_economica-ingresos_other']
        
        ws['P30'] = df_fila['data-start_act_economica-horario_actividad']

        if df_fila['data-start_act_economica-tiene_registro'] == 'yes':
            ws['T33'] = 'X'
        else:
            ws['W33'] = 'X'

        lugares_comercializa = df_fila['data-start_act_economica-lugares_comercializa']
        mapeo_lugares_comercializa = {
            'sitio': 'X39',
            'empresa': 'X40',
            'mercado': 'X41',
            'acopio': 'X43',
            'other': 'X44'
        }
        if lugares_comercializa in mapeo_lugares_comercializa:
            ws[mapeo_lugares_comercializa[lugares_comercializa]] = 'X'
            if lugares_comercializa == 'other':
                ws['K50'] = df_fila['data-start_act_economica-lugares_comercializa_other']


        frecuencia = df_fila['data-start_act_economica-frecuencia_ingresos']
        mapeo_frecuencia = {
            'diario': 'E53',
            'semanal': 'E54',
            'quincenal': 'E55',
            'mensual': 'M53',
            'semestral': 'M54',
            'other': 'M55'
        }
        if frecuencia in mapeo_frecuencia:
            ws[mapeo_frecuencia[frecuencia]] = 'X'
            if frecuencia == 'other':
                ws['K56'] = df_fila['data-start_act_economica-ingresos_other']


        if df_fila['data-start_act_economica-compra_vereda'] == 'yes':
            ws['T48'] = 'X'
        else:
            ws['W48'] = 'X'


        if df_fila['data-start_act_economica-comercializa_otra_vereda'] == 'yes':
            ws['U53'] = 'X'
            ws['U55'] = df_fila['data-start_act_economica-donde_comercializa']
        else:
            ws['Y53'] = 'X'

        estrato = int(df_fila['data-start_act_economica-estrato'])
        mapeo_estrato = {
            1: 'R59',
            2: 'T59',
            3: 'V59'
        }
        if estrato in mapeo_estrato:
            ws[mapeo_estrato[estrato]] = 'X'

        ws['Y59'] = df_fila['data-start_act_economica-servicios_publicos']

    elif df_fila['data-start_act_economica-permite_entrevista'] == 'no':
        ws['Z10'] = 'X'


def llenarFichaPredial(ws, df1_fila, df_pob_fila):

    # 1. DATOS GENERALES DE LA FICHA PREDIAL

    ws['Y1'] = df1_fila['data-info_general-num_encuesta']
    ws['C8'] = df1_fila['data-info_general-fecha']

    ws['U8'] = df1_fila['data-info_general-num_encuesta']           # Encuesta
    ws['C10'] = df1_fila['data-info_general-proyecto']              #'Proyecto'
    ws['S10'] = df1_fila['data-info_general-vereda']                #'Vereda'
    ws['C12'] = df1_fila['data-info_general-municipio']             # Municipio
    ws['L12'] = df1_fila['data-info_general-departamento']          #'Departamento'
    ws['T12'] = df1_fila['data-info_general-centro_poblado']        #'Centro Poblado'
    ws['T12'] = df1_fila['data-info_general-cuenca']                #'Cuenca'
    ws['D14'] = df1_fila['data-info_general-nombre_predio']         # 'Nombre del predio'
    
    
    mapeo_tenencia = {
        'propia': 'O14',
        'arriendo': 'R14'
    }

    ws[mapeo_tenencia.get(df1_fila['data-info_general-tenencia_pert'], '')] = 'X'
    
    mapeo_uso_para = {
        'trabajo': 'U14',
        'familiar': 'X14'
    }

    ws[mapeo_uso_para.get(df1_fila['data-info_general-tenencia_para'], '')] = 'X'

    ws['E16'] = df1_fila['data-info_general-nombre_propietario']        #'Nombre propietario'
    ws['K16'] = df1_fila['data-info_general-telefono_propietario']      #'Telefono propietario'

    
    map_vive_predio = {
        'yes': 'R16',
        'no': 'T16'
    }

    ws[map_vive_predio.get(df1_fila['data-info_general-vive_en_predio'], '')] = 'X'

    mapa_escriturada = {
        'yes': 'U18',
        'no': 'W18',
        'no_sabe': 'Y18'
    }

    ws[mapa_escriturada.get(df1_fila['data-info_general-escriturada'], '')] = 'X'

    ws['Q18'] = df1_fila['data-info_general-registro_escritura']        #'Num registro escritura'
    ws['H20'] = df1_fila['data-info_general-nombre_administrador']      #'Nombre administrador'
    ws['T20'] = df1_fila['data-info_general-telefono_administrador']    #'Telefono administrador'

    # 2. INFORMACIÓN ESPECÍFICA DEL PREDIO

    ws['F24'] = df1_fila['data-info_especifica-limite_norte']    #'Limite norte'
    ws['F26'] = df1_fila['data-info_especifica-limite_sur']     #'Limite sur'
    ws['F28'] = df1_fila['data-info_especifica-limite_este']    #'Limite este'
    ws['F30'] = df1_fila['data-info_especifica-limite_oeste']   #'Limite oeste'

    map_cuenta_con_vivienda = {
        'yes': 'K32',
        'no': 'M32'
    }

    if df1_fila['data-info_especifica-coord_captacion_este'] is not None:
        llave = 'yes'
    else:
        llave = 'no'

    ws[map_cuenta_con_vivienda.get(llave, '')] = 'X'

    ws['S32'] = df1_fila['data-info_especifica-punto_gps'] if df1_fila['data-info_especifica-punto_gps'] is not None else ''                           #'Punto GPS'
    ws['L34'] = df1_fila['data-info_especifica-coord_captacion_este'] if df1_fila['data-info_especifica-coord_captacion_este'] is not None else ''     #'Coorendenadas vivienda este'
    ws['S34'] = df1_fila['data-info_especifica-coord_captacion_norte'] if df1_fila['data-info_especifica-coord_captacion_norte'] is not None else ''   #'Coorendenadas vivienda norte'
    
    mapa_via_pavimentada = {
        'yes': 'I39',
        'no': 'K39'
    }

    ws[mapa_via_pavimentada.get(df1_fila['data-info_especifica-via_municipal'], '')] = 'X'

    ws['M39'] =  df1_fila['data-info_especifica-via_municipal_km'] if df1_fila['data-info_especifica-via_municipal_km'] is not None else '' #'Km via pavimentada'
    
    mapa_estad_via_pavimentada = {
        'B': 'R39',
        'R': 'T39',
        'M': 'W39',
    }

    ws[mapa_estad_via_pavimentada.get(df1_fila['data-info_especifica-via_municipal_estado'], '')] = 'X' if df1_fila['data-info_especifica-via_municipal_estado'] is not None else '' #'Estado via pavimentada'
  
    mapa_trocha = {
        'yes': 'I40',
        'no': 'K40'
    }

    ws[mapa_trocha.get(df1_fila['data-info_especifica-trocha'], '')] = 'X' if df1_fila['data-info_especifica-trocha'] is not None else ''

    ws['M40'] = df1_fila['data-info_especifica-trocha_km'] if df1_fila['data-info_especifica-trocha_km'] is not None else '' #'km trocha'

    mapa_estad_trocha = {
        'B': 'R40',
        'R': 'T40',
        'M': 'W40',
    }

    ws[mapa_estad_trocha.get(df1_fila['data-info_especifica-trocha_estado'], '')] = 'X' if df1_fila['data-info_especifica-trocha_estado'] is not None else '' #'Estado trocha'

    mapa_camino_herradura = {
        'yes': 'I41',
        'no': 'K41'
    }

    ws[mapa_camino_herradura.get(df1_fila['data-info_especifica-camino_herradura'], '')] = 'X' if df1_fila['data-info_especifica-camino_herradura'] is not None else '' #'km camino herradura'

    ws['M41'] = df1_fila['data-info_especifica-camino_herradura_km'] if df1_fila['data-info_especifica-camino_herradura_km'] is not None else '' #'km camino herradura'

    mapa_estad_camino_herradura = {
        'B': 'R41',
        'R': 'T41',
        'M': 'W41',
    }   

    ws[mapa_estad_camino_herradura.get(df1_fila['data-info_especifica-camino_herradura_estado'], '')] = 'X' if df1_fila['data-info_especifica-camino_herradura_estado'] is not None else '' #'Estado camino herradura'  

    mapa_via_fluvial = {
        'yes': 'I42',
        'no': 'K42'
    }

    ws[mapa_via_fluvial.get(df1_fila['data-info_especifica-via_fluvial'], '')] = 'X' if df1_fila['data-info_especifica-via_fluvial'] is not None else '' # Via fluvial

    ws['M42'] = df1_fila['data-info_especifica-via_fluvial_km'] if df1_fila['data-info_especifica-via_fluvial_km'] is not None else '' #'km via fluvial'

    mapa_estad_via_fluvial = {
        'B': 'R42',
        'R': 'T42',
        'M': 'W42',
    }

    ws[mapa_estad_via_fluvial.get(df1_fila['data-info_especifica-via_fluvial_estado'], '')] = 'X' if df1_fila['data-info_especifica-via_fluvial_estado'] is not None else '' #'Estado via fluvial'

    ws['I43'] = df1_fila['data-info_especifica-utilizable_tiempo'] if df1_fila['data-info_especifica-utilizable_tiempo'] is not None else '' # ¿Utilizable todo el tiempo?
    ws['K45'] = df1_fila['data-info_especifica-cabecera_cercana'] if df1_fila['data-info_especifica-cabecera_cercana'] is not None else '' # ¿Cuál es la cabecera más cercana?
    ws['K46'] = df1_fila['data-info_especifica-distancia_cabecera'] if df1_fila['data-info_especifica-distancia_cabecera'] is not None else '' #'Distancia (km) a la cabecera'

    # 3. SERVICIOS PUBLICOS

    map_energia = {
        'yes': 'E50',
        'no': 'G50'
    }

    ws[map_energia.get(df1_fila['data-start_servicios_publicos-energia'], '')] = 'X' if df1_fila['data-start_servicios_publicos-energia'] is not None else '' # ¿Tiene energía?

    map_contador_energia = {
        'yes': 'O50',
        'no': 'R50',
        'other': 'U50'
    }

    if df1_fila['data-start_servicios_publicos-contador_energia'] != 'other': # ¿Tiene contador de energía?
        ws[map_contador_energia.get(df1_fila['data-start_servicios_publicos-contador_energia'], '')] = 'X' if df1_fila['data-start_servicios_publicos-contador_energia'] is not None else ''
    else:
        ws['U50'] = df1_fila['data-start_servicios_publicos-contador_energia_other']


    map_cocina = {
        'gas': 'E52',
        'leña': 'G52',
        'velas': 'I52',
        'gasolina': 'M52',
        'other': 'R52'
    }

    if df1_fila['data-start_servicios_publicos-tipo_cocina'] != 'other': # ¿Tiene cocina?
        ws[map_cocina.get(df1_fila['data-start_servicios_publicos-tipo_cocina'], '')] = 'X' if df1_fila['data-start_servicios_publicos-tipo_cocina'] is not None else ''
    else:
        ws['R52'] = df1_fila['data-start_servicios_publicos-tipo_cocina_other']

    map_acueducto = {
        'yes': 'E54',
        'no': 'G54'
    }   

    ws[map_acueducto.get(df1_fila['data-start_servicios_publicos-acueducto'], '')] = 'X' if df1_fila['data-start_servicios_publicos-acueducto'] is not None else '' # ¿Tiene acueducto?

    # SUMINISTRO DE AGUA
    map_fuente_agua = {
        'pozo_aljibe': 'N54',
        'rio_quebrada': 'N56',
        'recolec_lluvia': 'V54',
        'conex_domici': 'V56',
    }

    ws[map_fuente_agua.get(df1_fila['data-start_servicios_publicos-suministro_agua'], '')] = 'X' if df1_fila['data-start_servicios_publicos-suministro_agua'] is not None else '' # ¿Cuál es la fuente de agua?

    # ALCANTARILLADO
    map_alcantarillado = {
        'yes': 'E58',
        'no': 'G58'
    }

    ws[map_alcantarillado.get(df1_fila['data-start_servicios_publicos-alcantarillado'], '')] = 'X' if df1_fila['data-start_servicios_publicos-alcantarillado'] is not None else '' # ¿Tiene alcantarillado?

    # DISPOSICION DE AGUAS RESIDUALES

    map_disposicion_aguas = {
        'inodoro': 'N58',
        'letrina': 'N60',
        'campo_abierto': 'V58',
        'pozo_septico': 'V60',
    }
    if df1_fila['data-start_servicios_publicos-disposicion_aguas'] is not None:
        for key in df1_fila['data-start_servicios_publicos-disposicion_aguas'].split(','):
            ws[map_disposicion_aguas.get(key, '')] = 'X'

    mapa_manejo_basuras = {
        'recoleccion': 'H62',
        'arroja': 'H64',
        'quema': 'N62',
        'arroja_agua': 'N64',
        'entierro': 'V62',
        'otro': 'T64'
    }

    if df1_fila['data-start_servicios_publicos-manejo_basura'] is not None:
        for key in df1_fila['data-start_servicios_publicos-manejo_basura'].split(','):
            if key != 'other':
                ws[mapa_manejo_basuras.get(key, '')] = 'X'
            else:
                ws['T64'] = df1_fila['data-start_servicios_publicos-manejo_basura_other']
        
    # TELECOMUNICACIONES
    keys_map_telecom = df1_fila['data-start_servicios_publicos-telecomunicaciones'].split(',')
    map_telecom = {
        'redes_tel': 'H66',
        'cabina_tel': 'N66',
        'internet': 'V66',
        'other': 'G68'
    }

    for key in keys_map_telecom:
        if key != 'other':
            ws[map_telecom[key]] = 'X'
        else: 
            ws[map_telecom[key]] = df1_fila['data-start_servicios_publicos-telecomunicaciones_other']

    ws['B72'] = df1_fila['data-start_servicios_publicos-observaciones_servicios_pub'] if df1_fila['data-start_servicios_publicos-observaciones_servicios_pub'] is not None else '' # Observaciones servicios públicos
    
    # 4. SERVICIOS SOCIALES

    mapa_regimen_salud_dueños = {
        'subsidiado': 'P78',
        'contributivo': 'U78',
    }

    ws[mapa_regimen_salud_dueños.get(df1_fila['data-start_servicios_sociales-regimen_salud_duenos'], '')] = 'X' if df1_fila['data-start_servicios_sociales-regimen_salud_duenos'] is not None else '' # ¿Cuál es el régimen de salud de

    mapa_regimen_salud_habitantes = {
        'subsidiado': 'P80',
        'contributivo': 'U80',
    }   

    ws[mapa_regimen_salud_habitantes.get(df1_fila['data-start_servicios_sociales-regimen_salud_habitantes'], '')] = 'X' if df1_fila['data-start_servicios_sociales-regimen_salud_habitantes'] is not None else '' # ¿Cuál es el régimen de salud de

    ws['J82'] = str(df1_fila['data-start_servicios_sociales-entidad_prestadora_duenos']) + ', ' + str(df1_fila['data-start_servicios_sociales-entidad_prestadora_habitantes']) if df1_fila['data-start_servicios_sociales-entidad_prestadora_duenos'] is not None and df1_fila['data-start_servicios_sociales-entidad_prestadora_habitantes'] is not None else '' # ¿Cuál es la entidad prestadora de salud?
    ws['J84'] = df1_fila['data-start_servicios_sociales-centro_salud'] if df1_fila['data-start_servicios_sociales-centro_salud'] is not None else 'Hospital o centro de salud más cercano'
    ws['T84'] = 'Localizado en'

       # MATERIAL DE CONSTRUCCIÓN

    ws['J86'] = 'Paredes'
    ws['O86'] = 'Techos'
    ws['T86'] = 'Pisos'

       # DISTRIBUCIÓN VIVIENDA
     
    ws['J88'] = 'Número de habitaciones'
    ws['O88'] = '¿Tiene sala?'
    ws['S88'] = '¿Tiene comedor?'
    ws['W88'] = '¿Tiene cocina?'
    ws['I90'] =  'Área total de la vivienda (m2)'
    ws['R90'] = 'Personas que habitan la vivienda'
    ws['F92'] = 'Estado de la vivienda'
    ws['B96'] = 'Observaciones sobre servicios sociales'

    # 5. CARACTERISTICAS DE LA POBLACIÓN (Habitantes de la vivienda)
    num_personas = df1_fila['data-start_carac_poblacion-num_personas']
    for i in range(num_personas):
        ws[f'B{102+i}'] = 'Nombre'
        ws[f'I{102+i}'] = 'Edad'
        # TODO LÓGICA PARA EL GÉNERO 
        if df_pob_fila[f'data-start_carac_poblacion-caracteristicas_poblacion-genero'] == 'M':
            ws[f'L{102+i}'] = 'X'
        else:
            ws[f'K{102+i}'] = 'X'
        ws[f'M{102+i}'] = 'Escolaridad'
        ws[f'P{102+i}'] = 'Relación con el encargado'
        ws[f'T{102+i}'] = 'Actividad'
    
    map_participacion_com = {
        'junta_padres': 'F114',
        'junta_accion_comunal': 'K114',
        'asociacion_empleados': 'Q114',
        'other': 'U114'
    }

    # TODO If del mapa participacion comunitaria
    
    ws['E116'] = 'Presencia institucional'

    # 6. USOS DEL SUELO
    ws['E120'] = 'Area total del predio (ha)'
    ws['K120'] = 'X' # Hectáreas por defecto, en la base de datos está el dato en fanegadas.
    ws['V120'] = 'Estrato socioeconómico del predio'

    map_uso_suelo = {
        'ganaderia': 'F122',
        'pastizales': 'F123',
        'agricultura': 'F122',
        'pancoger': 'F123',
        'other': 'F124'
    }

    # TODO If del mapa uso suelo

    
    ws['P126'] = 'Actividades complementarias'
    ws['B131'] = 'Actividades culturales'
    ws['B135'] = 'En caso de reasentamiento'
    ws['B143'] = 'Expectativas de la familia frente al proyecto'
    ws['B151'] = 'Observaciones generales'

    # TODO Toca ver como traer la firma del campo del drive.
    ws['S157'] = 'Firma de quien responde'
    ws['G159'] = 'C.C. de quien responde'

    # 7. FOTOGRAFÍA DE LA VIVIENDA

    ws['B163'] = 'Fotografía de la vivienda'

    # 8. ACTIVIDAD ECONÓMICA
    mapa_genera_actividad = {
        'yes': 'B182',
        'no': 'O182'

    }

    # TODO If del mapa genera actividad
    ws['K183'] = 'ID encuesta censo económico'

# 9. USOS Y USUARIOS (BIENES Y SERVICIOS AMBIENTALES)

    mapa_capta_aguas_sup = {
        'yes': 'I188',
        'no': 'L188'
    
    }

    # TODO If del mapa capta aguas superficiales

    ws['S188'] = 'ID Usos y Usuarios'

    mapa_capta_aguas_sub = {
        'yes': 'I189',
        'no': 'L189'
    }

    # TODO If del mapa capta aguas subterráneas
    ws['S189'] = 'ID FUNIAS'

def llenarUsosUsuarios(ws, df_fila):
    pass

def llenarInforme2(ws, df_fila):
    pass

def llenarInforme3(ws, df_fila):
    pass

def llenarInforme4(ws, df_fila):
    pass

def llenarInforme5(ws, df_fila):
    pass    

def llenarInforme6(ws, df_fila):    
    pass

def llenarInforme7(ws, df_fila):
    pass

def llenarInforme8(ws, df_fila):
    pass

def llenarInforme9(ws, df_fila):
    pass
