import pandas as pd
from openpyxl import load_workbook

xl = pd.ExcelFile(r"C:\Users\ACER\Documents\Formatos\Censo Económico Maute.xlsm") # Cambiar por la ruta en donde estén las encuestas.

direc_guardado = r"C:\Users\ACER\Documents\Formatos\Forms llenos\prueba" # Cambiar a la ruta en la que quieran que se guarden los forms.

form = r"C:\Users\Soporte\Documents\Formatos\FORMATO 6 MANUFACTURA - Aprobado.xlsx" # Cambiar por ruta en la que tengan el formato

def valorCol(base_name, index, df_fila):
    if index == 0:
        return df_fila.get(base_name, '')
    else:
        return df_fila.get(f'{base_name}.{index}', '')
    

df6 = xl.parse(sheet_name='FORMATO 6. MANUFACTURA', header=None)
df6 = df6.drop(columns=[0, 1, 2, 3])
df6_T = df6.transpose()
df6_T.columns = df6_T.iloc[0]
df6_T = df6_T.drop(df6_T.index[0])
df6_T.columns = df6_T.columns.str.strip()     
df6_T.columns = pd.io.common.dedup_names(df6_T.columns, is_potential_multiindex=False)
df_enc6 = df6_T.reset_index(drop = True)


def llenar_form_6(ws, df_fila):
    
    # A. IDENTIFICACIÓN ENTREVISTADO
    ws['AQ1'] = df_fila['Encuesta No.']

    if pd.notna(df_fila['Fecha']):
        fecha_str = str(df_fila['Fecha'])
        if '/' in fecha_str:
            ws['AO2'] = fecha_str.split('/')[0]
            ws['AR2'] = fecha_str.split('/')[1]
            ws['AU2'] = fecha_str.split('/')[2]
        elif '-' in fecha_str:
            ws['AO2'] = fecha_str.split('-')[0]
            ws['AR2'] = fecha_str.split('-')[1]
            ws['AU2'] = fecha_str.split('-')[2]
        else:
            print(f'Formato de fecha inesperado: {fecha_str}')
    else:
        print('Campo de fecha vacío')  
    
    ws['AL3'] = df_fila['Encuestador']

    ws['F7'] = df_fila['Nombre']
    ws['Z7'] = df_fila['Empresa']
    ws['AQ7'] = df_fila['Cargo']


    pertenece_asociacion = df_fila['¿Pertenece a alguna asociación?']
    if pd.notna(pertenece_asociacion):
        if pertenece_asociacion == 'Sí':
            ws['AA8'] = 'X'
            ws['AO8'] = df_fila['Otro, ¿Cuál?']
        elif pertenece_asociacion == 'No':
            ws['AC8'] = 'X'
    else:
        print("Campo vacío")

    
    # Pregunta 1: Bien final producido
    ws['A12'] = df_fila.get('Bien final producido', '')

    # Pregunta 2: ¿Con cuántos empleados cuenta la empresa?
    ws['A16'] = df_fila.get('¿Con cuántos empleados cuenta la empresa?', '')

    # Pregunta 3: La empresa cuenta con algún tipo de permiso ambiental
    permiso_ambiental = df_fila.get('La empresa cuenta con algún tipo de permiso ambiental', '')
    if permiso_ambiental == 'Sí':
        ws['Q13'] = 'x'
        ws['W13'] = df_fila.get('¿Cuál?', '')
    elif permiso_ambiental == 'No':
        ws['S13'] = 'x'
    

    # Pregunta 4: Tipo de empresa
    tipo_empresa = df_fila.get('Tipo de empresa', '')
    if tipo_empresa == 'Pública':
        ws['R17'] = 'x'
    elif tipo_empresa == 'Privada':
        ws['X17'] = 'x'
    elif tipo_empresa == 'Mixta':
        ws['AD17'] = 'x'


    # Pregunta 5: Vende principalmente en
    vende_principalmente = df_fila.get('Vende principalmente en', '')
    if vende_principalmente == 'Sitio':
        ws['AU13'] = 'x'
    elif vende_principalmente == 'Vereda':
        ws['AU14'] = 'x'
    elif vende_principalmente == 'Casco urbano':
        ws['AU15'] = 'x'
    elif vende_principalmente == 'Otros municipios y/o veredas':
        ws['AU16'] = 'x'
        ws['AN17'] = df_fila.get('Otros, ¿Cuáles?', '')

    # Pregunta 6: Procedencia de los compradores
    if pd.notna(df_fila.get('Hidrocarburos', '')):

        ws['V18'] = 'X'
        ws['AC18'] = df_fila.get('Hidrocarburos', '')
    
    elif pd.notna(df_fila.get('Otro', '')):

        ws['V19'] = 'X'
        ws['AC19'] = df_fila.get('Hidrocarburos', '')
    
    else:
        print('Campos vacíos')
    
    # Sobre la actividad, piensa:

        continuidad = df_fila['Sobre la actividad, piensa: Continuidad']
    if pd.notna(continuidad):
        if continuidad == 'Continuar con la actividad':
            ws['L21'] == 'X'
            
        elif continuidad == 'Finalizar la actividad':
            ws['N21'] == 'X'
            
    else:
        print('Campo vacío') 

    produccion = df_fila['Sobre la actividad, piensa: Producción']
    if pd.notna(produccion):
        if produccion == 'Ampliar la producción':
            ws['AB21'] == 'X'
            ws['AU21'] == 'X'               
        elif produccion == 'Permanecer con la misma producción':
            ws['AD21'] == 'X'
            ws['AS21'] == 'X'   
        else:
            ws['AD21'] == 'X'
            ws['AU21'] == 'X'
    else:
        print('Campo vacío') 

    columnas = {
        "Tipo de producto fabricado": "B",
        "Unidad de medida": "K",
        "Cantidad producida": "P",
        "Frecuencia de producción": "V",
        "Costos de producción por unidad": "AE",
        "Cantidad vendida por semana": "AM",
        "Precio de venta": "AS"
    }

    
    for i in range(3):
        fila_id = 26 + i  # Empezar desde la fila 25 y avanzar
        ws[f"{columnas['Tipo de producto fabricado']}{fila_id}"] = valorCol('Tipo de producto fabricado', i, df_fila)
        ws[f"{columnas['Unidad de medida']}{fila_id}"] = valorCol('Unidad de medida', i, df_fila)
        ws[f"{columnas['Cantidad producida']}{fila_id}"] = valorCol('Cantidad producida', i, df_fila)
        ws[f"{columnas['Frecuencia de producción']}{fila_id}"] = valorCol('Frecuencia de producción', i, df_fila)
        ws[f"{columnas['Costos de producción por unidad']}{fila_id}"] = valorCol('Costos de producción por unidad', i, df_fila)
        ws[f"{columnas['Cantidad vendida por semana']}{fila_id}"] = valorCol('Cantidad vendida por semana', i, df_fila)
        ws[f"{columnas['Precio de venta']}{fila_id}"] = valorCol('Precio de venta', i, df_fila)

    # Equipo/maquinaria
    ws['B33'] = df_fila['Equipo/maquinaria 1']
    ws['B34'] = df_fila['Equipo/maquinaria 2']
    ws['B35'] = df_fila['Equipo/maquinaria 3']

    # Precio al que lo compró
    ws['P33'] = df_fila['Precio al que lo compró']
    ws['P34'] = df_fila['Precio al que lo compró.1']
    ws['P35'] = df_fila['Precio al que lo compró.2']

    # Cantidad que posee la unidad económica
    ws['Z33'] = df_fila['Cantidad que posee la unidad económica']
    ws['Z34'] = df_fila['Cantidad que posee la unidad económica.1']
    ws['Z35'] = df_fila['Cantidad que posee la unidad económica.2']

    # Vida útil
    ws['AJ33'] = df_fila['Vida útil']
    ws['AJ34'] = df_fila['Vida útil.1']
    ws['AJ35'] = df_fila['Vida útil.2']

    # Procedencia
    ws['AR33'] = df_fila['Procedencia']
    ws['AR34'] = df_fila['Procedencia.1']
    ws['AR35'] = df_fila['Procedencia.2']
        
        # Servicios
    ws['B39'] = df_fila['Servicios']
    ws['B40'] = df_fila['Servicios.1']
    ws['B41'] = df_fila['Servicios.2']

    # Insumo/Materia prima
    ws['D39'] = df_fila['Insumo/Materia prima 1']
    ws['D40'] = df_fila['Insumo/Materia prima 2']
    ws['D41'] = df_fila['Insumo/Materia prima 3']

    # Precio compra
    ws['F39'] = df_fila['Precio compra']
    ws['F40'] = df_fila['Precio compra.1']
    ws['F41'] = df_fila['Precio compra.2']

    # Cantidad
    ws['H39'] = df_fila['Cantidad']
    ws['H40'] = df_fila['Cantidad.1']
    ws['H41'] = df_fila['Cantidad.2']

    # Frecuencia de compra
    ws['J39'] = df_fila['Frecuencia de compra']
    ws['J40'] = df_fila['Frecuencia de compra.1']
    ws['J41'] = df_fila['Frecuencia de compra.2']

    # Procedencia
    ws['L39'] = df_fila['Procedencia.3']
    ws['L40'] = df_fila['Procedencia.4']
    ws['L41'] = df_fila['Procedencia.5']

    agua_fuente = df_fila['¿De dónde se abastece del recurso hídrico?']
    if pd.notna(agua_fuente):
        if agua_fuente == 'Aljibe':
            ws['W43'] = 'X'
        elif agua_fuente == 'Acueducto veredal':
            ws['AG43'] = 'X'
        elif agua_fuente == 'Otro':
            ws['AN43'] = 'X'
            ws['AT43'] = df_fila['¿Cuál?.1']    

    ws['W44'] = df_fila['Forma de extracción']

    ws['AO44'] = df_fila['Cantidad estimada (m3)']

    energia = df_fila['¿Qué tipo de energía utiliza?']
    if energia == 'Energía Eléctrica':
        ws['AC45'] = 'X'
    elif energia == 'Energía Solar':
        ws['AL45'] = 'X'
    elif energia == 'Otro':
        ws['AT45'] = df_fila['¿Cuál?.2']

    energia_coccion = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
    if energia_coccion == 'Energía Eléctrica':
        ws['AC46'] = 'X'
    elif energia_coccion == 'Leña':
        ws['AH46'] = 'X'
    elif energia_coccion == 'Gas':
        ws['AN46'] = 'X'
    elif energia_coccion == 'Otro':       
        ws['AT46'] = df_fila['¿Cuál?.3']

    alcantarillado = df_fila['¿Cuenta con servicio de alcantarillado?']
    if pd.notna(alcantarillado):
        if alcantarillado == 'Sí':
            ws['AB47'] = 'X'
            ws['AO47'] = df_fila['¿Cuál?.4']  
        elif alcantarillado == 'No':
            ws['AD47'] = 'X'

    ws['AC49'] = df_fila['¿Cuál fue el monto total gastado en insumos del último mes??']


    servicio = df_fila['¿Demanda algún tipo de servicio de la región?']
    if pd.notna(servicio):
        if servicio == 'Seguridad':
            ws['L52'] = 'X'
        elif servicio == 'Mano de obra calificada':
            ws['L53'] = 'X'
        elif servicio == 'Mano de obra no calificada':
            ws['L54'] = 'X'
        elif servicio == 'Transporte':
            ws['L55'] = 'X'
        elif servicio == 'Alojamiento':
            ws['V52'] = 'X'
        elif servicio == 'Alimentación':
            ws['V53'] = 'X'
        elif servicio == 'Otro':
            ws['V54'] = 'X'
            ws['P55'] = df_fila['Otro, ¿Cuál?.1']
    else:
        print('Campo vacío')

    ws['AC51'] = df_fila['¿Con que frecuencia demanda servicios de la región?']
    
    for i in range(10):
        prefijo_persona = 60 + i
        ws[f'E{prefijo_persona}'] = valorCol('Cargo', i, df_fila)
        ws[f'K{prefijo_persona}'] = valorCol('Edad (años)', i, df_fila)
        ws[f'L{prefijo_persona}'] = valorCol('Duración jornada (horas)', i, df_fila)

        manoObra = valorCol('Tipo de mano de obra', i, df_fila)
        if pd.notna(manoObra):
            if manoObra == 'Familiar':
                ws[f'B{prefijo_persona}'] = 'X'
            elif manoObra == 'Contratado':
                ws[f'D{prefijo_persona}'] = 'X'

        # Genero
        genero = valorCol('Género', i,df_fila)
        if pd.notna(genero):
            if genero == 'Masculino':
                ws[f'J{prefijo_persona}'] = 'X'
            elif genero ==  'Femenino':
                ws[f'H{prefijo_persona}'] = 'X'

        # Escolaridad 
        escolaridad = valorCol('Escolaridad', i, df_fila)
        if pd.notna(escolaridad):
            if escolaridad:
                if escolaridad == 'Primaria':
                    ws[f'N{prefijo_persona}'] = 'X'
                elif escolaridad == 'Bachillerato':
                    ws[f'Q{prefijo_persona}'] = 'X'
                elif escolaridad == 'Técnico':
                    ws[f'S{prefijo_persona}'] = 'X'
                elif escolaridad == 'Pregrado':
                    ws[f'U{prefijo_persona}'] = 'X'
                elif escolaridad == 'Posgrado':
                    ws[f'W{prefijo_persona}'] = 'X'
        else:
            print(f'Campo vacio')

        # Contrato 
        contrato = valorCol('Contrato', i,df_fila)
        if contrato:
            if contrato == 'Tem.':
                ws[f'AC{prefijo_persona}'] = valorCol('Contrato', i, df_fila)
            elif contrato == 'Fij':
                ws[f'AE{prefijo_persona}'] = 'X'
        else:
            print(f'Campo vacio')

        # Pago de seguridad social 
        pago_seguridad = valorCol('Pago de seguridad social', i, df_fila)
        if pago_seguridad:
            if pago_seguridad == 'Si':
                ws[f'AG{prefijo_persona}'] = 'X'
            elif pago_seguridad == 'No':
                ws[f'AI{prefijo_persona}'] = 'X'
        else:
            print(f'Campo vacio')

        # Remuneración 
        remuneracion = valorCol('Remuneración', i, df_fila)
        if remuneracion:
            if remuneracion == 'Inferiores a $900.000':
                ws[f'AU{prefijo_persona}'] = 'X'
            elif remuneracion == '$900.000 a $1.800.000':
                ws[f'AV{prefijo_persona}'] = 'X'
            elif remuneracion == '$1.801.000 a $2.700.000':
                ws[f'AW{prefijo_persona}'] = 'X'
            elif remuneracion == 'Superiores a $2.701.000':
                ws[f'AX{prefijo_persona}'] = 'X'
        else:
            print(f'Campo vacio')

        # Información adicional
      
        ws[f'AJ{prefijo_persona}'] = valorCol('Procedencia', 6 + i, df_fila)
        ws[f'AK{prefijo_persona}'] = valorCol('Residencia', i, df_fila)
        ws[f'AN{prefijo_persona}'] = valorCol('Tiempo trabajado', i, df_fila)
        ws[f'AO{prefijo_persona}'] = valorCol('# Personas núcleo familiar', i, df_fila)
        ws[f'AQ{prefijo_persona}'] = valorCol('Personas a cargo', i, df_fila)
        ws[f'AS{prefijo_persona}'] = valorCol('Lugar de residencia familiar', i,df_fila)

   




    
