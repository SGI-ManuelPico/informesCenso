import pandas as pd
from openpyxl import load_workbook
import os, re


class InformeSegundo:
    def __init__(self):
        pass

    def lecturaArchivoSegundo(self):
        """
        Lectura del segundo archivo del censo económico.
        """
        # Lectura de rutas
        rutaArchivoInicial = os.getcwd() + "\\Censo Económico Maute.xlsm"
        archivoInicial = pd.ExcelFile(rutaArchivoInicial)
        archivoInicial = archivoInicial.parse(sheet_name="FORMATO 2. AGROPECUARIO", header=None)

        # Ajustes preliminares al archivo inicial
        archivoInicial = archivoInicial.drop(columns=[0, 1, 2, 3]).transpose()
        archivoInicial.columns = archivoInicial.iloc[0]
        archivoInicial = archivoInicial.drop(archivoInicial.index[0])
        archivoInicial.columns = archivoInicial.columns.str.strip()
        archivoInicial.columns = pd.io.common.dedup_names(archivoInicial.columns, is_potential_multiindex=False)
        archivoInicial = archivoInicial.reset_index(drop=True)

        return archivoInicial

    def valorCol(self, base_name, index, df_fila):
        if index == 0:
            return df_fila.get(base_name, '')
        else:
            return df_fila.get(f'{base_name}.{index}', '')

    def determinarMayoríaUnidad(self, unidades):
        """
        Esta función determina la unidad mayoritaria en una lista de unidades.
        """
        return max(set(unidades), key=unidades.count)

    xl = pd.ExcelFile(r"C:\Users\ACER\Documents\Formatos\Censo Económico Maute.xlsm") # Cambiar por la ruta en donde estén las encuestas.

    direc_guardado = r"C:\Users\ACER\Documents\Formatos\Forms llenos\prueba" # Cambiar a la ruta en la que quieran que se guarden los forms.

    form = r"C:\Users\Soporte\Documents\Formatos\FORMATO 2 AGROPECUARIO - Aprobado.xlsx" # Cambiar por ruta en la que tengan el formato



    def llenar_form_2(self, ws, df_fila):
        # A. IDENTIFICACIÓN ENTREVISTADO
        ws['AO1'] = df_fila['Encuesta No.']

        if pd.notna(df_fila['Fecha(DD/MM/AAAA)']):
            fecha_str = str(df_fila['Fecha(DD/MM/AAAA)'])
            if '/' in fecha_str:
                ws['AM2'] = fecha_str.split('/')[0]
                ws['AP2'] = fecha_str.split('/')[1]
                ws['AU2'] = fecha_str.split('/')[2]
            elif '-' in fecha_str:
                ws['AM2'] = fecha_str.split('-')[0]
                ws['AP2'] = fecha_str.split('-')[1]
                ws['AU2'] = fecha_str.split('-')[2]
            else:
                print(f'Formato de fecha inesperado: {fecha_str}')
        else:
            print('Campo de fecha vacío')  
        
        ws['AL3'] = df_fila['Encuestador']

        ws['E7'] = df_fila['Nombre']
        ws['AC7'] = df_fila['Empresa']
        ws['AQ7'] = df_fila['Cargo']


        pertenece_asociacion = df_fila['¿Pertenece a alguna asociación?']
        if pd.notna(pertenece_asociacion):
            if pertenece_asociacion == 'Sí':
                ws['M8'] = 'X'
                ws['W8'] = df_fila['¿Cuál?']
            elif pertenece_asociacion == 'No':
                ws['AO8'] = 'X'
        else:
            print("Campo vacío")

        # B. INFORMACIÓN E IDENTIFICACIÓN GENERAL DEL PREDIO

        # Área Total
        ws['G11'] = df_fila['Área Total (Ha)']

        # Tipo de uso
        ws['G14'] = df_fila['Cultivos (Ha)']
        ws['G15'] = df_fila['Pastos (Ha)']
        ws['G16'] = df_fila['Bosque natural (Ha)']
        ws['G17'] = df_fila['Rastrojo (Ha)']
        ws['G18'] = df_fila['Bosque plantado (Ha)']
        ws['AB14'] = df_fila['Tierras erosionadas (Ha)']
        ws['AB15'] = df_fila['Lagos y lagunas (m2)']
        ws['AB16'] = df_fila['Reservorios (m2)']
        ws['AB17'] = df_fila['Construcciones (m2)']
        ws['Y18'] = df_fila['Otros']

        # Precio de arrendamiento por Ha
        ws['AN13'] = df_fila['Precio de arrendamiento por Ha']

        # Observaciones
        ws['AM16'] = df_fila['Observaciones']

        # ACTIVIDADES AGROPECUARIAS

        # Producto
        ws['F23'] = df_fila['Producto']

        # Columna inicial para cada cultivo
        cultivo_col_start = ['V', 'AF', 'AN']

        # Listas para almacenar las unidades de cada categoría
        unidades_area_cultivada = []
        unidades_establecimiento = []
        unidades_mantenimiento = []
        unidades_cosecha = []
        unidades_precio = []
        unidades_autoconsumo = []

        for i in range(3):  # Hay tres cultivos
            cultivo_prefix = cultivo_col_start[i]

            # Área cultivada
            ws[f'{cultivo_prefix}24'] = self.valorCol('Área cultivada', i, df_fila)
            
            # Unidad de Área cultivada
            unidad_area_cultivada = self.valorCol('Unidad', 6 * i, df_fila)
            unidades_area_cultivada.append(unidad_area_cultivada)
            
            # Número de cosechas por año
            ws[f'{cultivo_prefix}25'] = self.valorCol('No. de cosechas por año', i, df_fila)
            
            # Costos de establecimiento
            ws[f'{cultivo_prefix}26'] = self.valorCol('Costos de establecimiento', i, df_fila)
            
            # Unidad de Costos de establecimiento
            unidad_establecimiento = self.valorCol('Unidad', 6 * i + 1, df_fila)
            unidades_establecimiento.append(unidad_establecimiento)
            
            # Costos de mantenimiento
            ws[f'{cultivo_prefix}27'] = self.valorCol('Costos de mantenimiento', i, df_fila)
            
            # Unidad de Costos de mantenimiento
            unidad_mantenimiento = self.valorCol('Unidad', 6 * i + 2, df_fila)
            unidades_mantenimiento.append(unidad_mantenimiento)
            
            # Costos de cosecha
            ws[f'{cultivo_prefix}28'] = self.valorCol('Costos de cosecha', i, df_fila)
            
            # Unidad de Costos de cosecha
            unidad_cosecha = self.valorCol('Unidad', 6 * i + 3, df_fila)
            unidades_cosecha.append(unidad_cosecha)
            
            # Volumen de producción
            ws[f'{cultivo_prefix}29'] = self.valorCol('Volumen de producción', i, df_fila)
            
            # Precio de venta
            ws[f'{cultivo_prefix}30'] = self.valorCol('Precio de venta', i, df_fila)
            
            # Unidad de Precio de venta
            unidad_precio = self.valorCol('Unidad', 6 * i + 4, df_fila)
            unidades_precio.append(unidad_precio)

            # Autoconsumo
            ws[f'{cultivo_prefix}31'] = self.valorCol('Autoconsumo', i, df_fila)
            
            # Unidad de Autoconsumo
            unidad_autoconsumo = self.valorCol('Unidad', 6 * i + 5, df_fila)
            unidades_autoconsumo.append(unidad_autoconsumo)
        
        # Determinar las unidades mayoritarias
        unidad_area_cultivada_mayoria = self.determinar_mayoria_unidad(unidades_area_cultivada)
        unidad_establecimiento_mayoria = self.determinar_mayoria_unidad(unidades_establecimiento)
        unidad_mantenimiento_mayoria = self.determinar_mayoria_unidad(unidades_mantenimiento)
        unidad_cosecha_mayoria = self.determinar_mayoria_unidad(unidades_cosecha)
        unidad_precio_mayoria = self.determinar_mayoria_unidad(unidades_precio)
        unidad_autoconsumo_mayoria = self.determinar_mayoria_unidad(unidades_autoconsumo)

        # Rellenar la unidad mayoritaria en el campo correspondiente
        if unidad_area_cultivada_mayoria == 'Ha':
            ws['Q24'] = 'X'
        elif unidad_area_cultivada_mayoria == 'm2':
            ws['T24'] = 'X'
        
        if unidad_establecimiento_mayoria == 'Ha':
            ws['Q26'] = 'X'
        elif unidad_establecimiento_mayoria == 'm2':
            ws['T26'] = 'X'

        if unidad_mantenimiento_mayoria == 'Ha':
            ws['Q27'] = 'X'
        elif unidad_mantenimiento_mayoria == 'm2':
            ws['T27'] = 'X'

        if unidad_cosecha_mayoria == 'Ha':
            ws['Q28'] = 'X'
        elif unidad_cosecha_mayoria == 'm2':
            ws['T28'] = 'X'

        if unidad_precio_mayoria == 'Unidad':
            ws['N30'] = 'X'
        elif unidad_precio_mayoria == 'Kg':
            ws['Q30'] = 'X'
        elif unidad_precio_mayoria == 'Ton':
            ws['T30'] = 'X'

        if unidad_autoconsumo_mayoria == 'Carga':
            ws['N31'] = 'X'
        elif unidad_autoconsumo_mayoria == 'Kg':
            ws['P31'] = 'X'
        elif unidad_autoconsumo_mayoria == '@':
            ws['R31'] = 'X'
        elif unidad_autoconsumo_mayoria == 'Ton':
            ws['U31'] = 'X' 

        ws['L32'] = df_fila['¿Destino final del producto?']
        

        continuidad = df_fila['Sobre la actividad, piensa: Continuidad']
        if pd.notna(continuidad):
            if continuidad == 'Continuar con la actividad':
                ws['G36'] = 'X'
                ws['AP36'] = 'X'
            elif continuidad == 'Finalizar la actividad':
                ws['I36'] = 'X'
                ws['AN36'] = 'X'  
        else:
            print('Campo vacío') 

        produccion = df_fila['Sobre la actividad, piensa: Producción']
        if pd.notna(produccion):
            if produccion == 'Ampliar la producción':
                ws['Q36'] = 'X'
                ws['AF36'] = 'X'               
            elif produccion == 'Permanecer con la misma producción':
                ws['S36'] = 'X'
                ws['AD36'] = 'X'   
        else:
            print('Campo vacío') 

        ws['AS36'] = df_fila['¿Por qué']



    # Llenar raza y # de cabezas
        for i in range(3):
            # Leche o Cría
            ws[f'N{41 + i}'] = self.valorCol('Raza', i, df_fila)
            ws[f'AB{41 + i}'] = self.valorCol('# de cabezas', i, df_fila)

            # Carne
            ws[f'AJ{41 + i}'] = self.valorCol('Raza', i + 3, df_fila)
            ws[f'AT{41 + i}'] = self.valorCol('# de cabezas', i + 3, df_fila)

        # Número de terneros    
        ws['Q44'] = self.valorCol('Número de reses en producción', df_fila)
        ws['AK44'] = self.valorCol('Número de reses en producción.1', df_fila)

        ws['Q45'] = self.valorCol('Número de terneros', df_fila)
        ws['AK45'] = self.valorCol('Número de terneros.1', df_fila)  

        # Número de novillos
        ws['Q46'] = self.valorCol('Número de novillos', df_fila)
        ws['AK46'] = self.valorCol('Número de novillos.1', df_fila)

        # Número de novillas
        ws['Q47'] = self.valorCol('Número de novillas', df_fila)
        ws['AK47'] = self.valorCol('Número de novillas.1', df_fila)

        # Número de reproductores
        ws['Q48'] = self.valorCol('Número de reproductores', df_fila)
        ws['AK48'] = self.valorCol('Número de reproductores.1', df_fila)

        # Número de partos al año
        ws['Q49'] = self.valorCol('Número de partos al año', df_fila)
        ws['AK49'] = self.valorCol('Número de partos al año.1', df_fila)

        # Número de vacas para ordeño
        ws['Q50'] = self.valorCol('Número de vacas para ordeño', df_fila)
        ws['AK50'] = self.valorCol('Número de vacas para ordeño.1', df_fila)

        # Tiempo de venta después destetado
        ws['Q51'] = self.valorCol('Tiempo de venta después destetado', df_fila)
        ws['AK51'] = self.valorCol('Tiempo de venta después destetado.1', df_fila)

        # Peso promedio para la venta en Kg
        ws['Q52'] = self.valorCol('Peso promedio para la venta en Kg', df_fila)
        ws['AK52'] = self.valorCol('Peso promedio para la venta en Kg.1', df_fila)

        # Precio promedio

        ws['N53'] = df_fila['Precio promedio: Litro']
        ws['X53'] = df_fila['Precio promedio: Botella']
        ws['AH53'] = df_fila['Precio promedio: Kg']
        ws['AQ53'] = df_fila['Precio promedio: Cabeza']

        # ¿Cada cuánto produce? 

        frecuencia_leche = df_fila['¿Cada cuánto produce?']
        if pd.notna(frecuencia_leche):
            if frecuencia_leche == 'Diario':
                ws['P54'] = 'X'
            elif frecuencia_leche == 'Semanal':
                ws['W54'] = 'X'
            elif frecuencia_leche == 'Otro':
                ws['AD54'] = 'X'

        # Frecuencia de producción para Carne
        frecuencia_carne = df_fila['¿Cada cuánto produce?.1']
        if pd.notna(frecuencia_carne):
            if frecuencia_carne == 'Mensual':
                ws['AJ54'] = 'X'
            elif frecuencia_carne == 'Trimestral':
                ws['AO54'] = 'X'
            elif frecuencia_carne == 'Otro':
                ws['AU54'] = 'X'

        # Litros y Botella para Leche o Cría
        promedio_produccion_litro = df_fila['Promedio de producción Litro']
        if pd.notna(promedio_produccion_litro):
            ws['M55'] = promedio_produccion_litro

        promedio_produccion_botella = df_fila['Promedio de producción Botella']
        if pd.notna(promedio_produccion_botella):
            ws['W55'] = promedio_produccion_botella

        # Kg y Cabezas para Carne
        promedio_produccion_kg = df_fila['Promedio de producción Kg']
        if pd.notna(promedio_produccion_kg):
            ws['AG55'] = promedio_produccion_kg

        promedio_produccion_cabezas = df_fila['Promedio de producción Cabezas']
        if pd.notna(promedio_produccion_cabezas):
            ws['AP55'] = promedio_produccion_cabezas   

        if (pd.notna(df_fila['¿Destino final del producto?.1'])) & (pd.notna(df_fila['¿Destino final del producto?.2'])):
            ws['L56'] = 'Leche o cría: ' + df_fila['¿Destino final del producto?.1'] + ' Carne: ' + df_fila['¿Destino final del producto?.2']

        continuidad2 = df_fila['Sobre la actividad, piensa: Continuidad.1']
        if pd.notna(continuidad2):
            if continuidad2 == 'Continuar con la actividad':
                ws['G60'] = 'X'
                ws['AP60'] = 'X'
            elif continuidad2 == 'Finalizar la actividad':
                ws['I60'] = 'X'
                ws['AN60'] = 'X'  
        else:
            print('Campo vacío') 

        produccion2 = df_fila['Sobre la actividad, piensa: Producción.1']
        if pd.notna(produccion2):
            if produccion2 == 'Ampliar la producción':
                ws['Q60'] = 'X'
                ws['AF60'] = 'X'               
            elif produccion2 == 'Permanecer con la misma producción':
                ws['S60'] = 'X'
                ws['AD60'] = 'X'   
        else:
            print('Campo vacío') 

        ws['AS77'] = df_fila['¿Por qué.1']

        #Raza

        ws[f'T64'] = self.valorCol('Raza', 6, df_fila)
        ws[f'AF64'] = self.valorCol('Raza', 7, df_fila)
        ws[f'AO64'] = self.valorCol('Raza', 8, df_fila)

        # Número de Hembras
        ws[f'T65'] = self.valorCol('# Hembras', 0, df_fila)
        ws[f'AF65'] = self.valorCol('# Hembras', 1, df_fila)
        ws[f'AO65'] = self.valorCol('# Hembras', 2, df_fila)

        # Número de machos
        ws[f'T66'] = self.valorCol('# Machos', 0, df_fila)
        ws[f'AF66'] = self.valorCol('# Machos', 1, df_fila)
        ws[f'AO66'] = self.valorCol('# Machos', 2, df_fila)

        # Tiene Marranos para la venta
        ws[f'T67'] = self.valorCol('TTiene Marranos para la venta', 0, df_fila)
        ws[f'AE67'] = self.valorCol('TTiene Marranos para la venta', 1, df_fila)
        ws[f'AM67'] = self.valorCol('TTiene Marranos para la venta', 2, df_fila)

        # Peso promedio para la venta por animal (Kg)
        ws[f'T68'] = self.valorCol('Peso promedio para la venta por animal (Kg)', 0, df_fila)
        ws[f'AF68'] = self.valorCol('Peso promedio para la venta por animal (Kg)', 1, df_fila)
        ws[f'AO68'] = self.valorCol('Peso promedio para la venta por animal (Kg)', 2, df_fila)

        # Número promedio de animales vendidos por año
        ws[f'T69'] = self.valorCol('# Promedio de animales vendidos por año', 0, df_fila) + self.valorCol('# Promedio de animales vendidos por año', 1, df_fila) + self.valorCol('# Promedio de animales vendidos por año', 2, df_fila)

        # Cantidad empleada para autoconsumo (Kg)
        ws[f'T70'] = self.valorCol('Cantidad empleada para autoconsumo', 0, df_fila)
        ws[f'AF70'] = self.valorCol('Cantidad empleada para autoconsumo', 1, df_fila)
        ws[f'AO70'] = self.valorCol('Cantidad empleada para autoconsumo', 2, df_fila)

        # Precio unitario de venta
        ws[f'T71'] = self.valorCol('Precio unitario de venta', 0, df_fila)
        ws[f'AF71'] = self.valorCol('Precio unitario de venta', 1, df_fila)
        ws[f'AO71'] = self.valorCol('Precio unitario de venta', 2, df_fila)

        # Costo aproximado de producción
        ws[f'T72'] = self.valorCol('Costo aproximado de producción', 0, df_fila)
        ws[f'AF72'] = self.valorCol('Costo aproximado de producción', 1, df_fila)
        ws[f'AO72'] = self.valorCol('Costo aproximado de producción', 2, df_fila)

        # Obtener las unidades para 'Precio unitario de venta'
        unidades_precio_venta = [
            df_fila.get('Unidad.18', ''), 
            df_fila.get('Unidad.19', ''), 
            df_fila.get('Unidad.20', '')
        ]

        # Obtener las unidades para 'Costo aproximado de producción'
        unidades_costo_produccion = [
            df_fila.get('Unidad.21', ''), 
            df_fila.get('Unidad.22', ''), 
            df_fila.get('Unidad.23', '')
        ]

        # Determinar la unidad mayoritaria para cada caso
        unidad_mayoritaria_precio_venta = self.determinar_mayoria_unidad(unidades_precio_venta)
        unidad_mayoritaria_costo_produccion = self.determinar_mayoria_unidad(unidades_costo_produccion)

        # Marcar la unidad mayoritaria en las celdas correspondientes
        # Precio unitario de venta
        ws['Q71'] = 'X' if unidad_mayoritaria_precio_venta == 'Animal' else ''
        ws['S71'] = 'X' if unidad_mayoritaria_precio_venta == 'kg' else ''

        # Costo aproximado de producción
        ws['Q72'] = 'X' if unidad_mayoritaria_costo_produccion == 'Animal' else ''
        ws['S72'] = 'X' if unidad_mayoritaria_costo_produccion == 'kg' else ''

        if pd.notna(df_fila['¿Destino final del producto?.3']):
            ws['T73'] = df_fila['¿Destino final del producto?.3']


        continuidad3 = df_fila['Sobre la actividad, piensa: Continuidad.2']
        if pd.notna(continuidad3):
            if continuidad3 == 'Continuar con la actividad':
                ws['G77'] = 'X'
                ws['AP77'] = 'X'
            elif continuidad3 == 'Finalizar la actividad':
                ws['I77'] = 'X'
                ws['AN77'] = 'X'  
        else:
            print('Campo vacío') 

        produccion3 = df_fila['Sobre la actividad, piensa: Producción.2']
        if pd.notna(produccion3):
            if produccion3 == 'Ampliar la producción':
                ws['Q77'] = 'X'
                ws['AF77'] = 'X'               
            elif produccion3 == 'Permanecer con la misma producción':
                ws['S77'] = 'X'
                ws['AD77'] = 'X'   
        else:
            print('Campo vacío') 

        ws['AS77'] = df_fila['¿Por qué.2']

        # Explotación Avícola

        tipo_explotacion = self.valorCol('Tipo de explotación', 0, df_fila)
        if pd.notna(tipo_explotacion):
            if tipo_explotacion == 'Cría':
                ws['R80'] = 'X'
            elif tipo_explotacion == 'Engorde':
                ws['AC80'] = 'X'
            elif tipo_explotacion == 'Ponedoras':
                ws['AK80'] = 'X'
            elif tipo_explotacion == 'Gallina campesina':
                ws['AU80'] = 'X'
        else:
            print('Campo vacío')

        campos = ['# Animales', 'Producción mensual (Aves)', 'Unidades vendidas al mes (Aves)', 'Valor unitario de venta', 'Costo por animal', 'Cantidad empleada para autoconsumo/Mes (Aves)']
        columnas = ['N', 'AB', 'AM']

        for i, col in enumerate(columnas):
            for j, campo in enumerate(campos):
                if campo != 'Cantidad empleada para autoconsumo/Mes (Aves)':
                    cell = f'{col}{82+j}'
                    ws[cell] = self.valorCol(campo, i, df_fila)
                    
                else:
                    cell = f'{col}{83+j}'              
                    ws[cell] = self.valorCol(campo, i, df_fila)

        # Costo por animal
        ws['O86'] = self.valorCol('Costo por animal', 0, df_fila)
        ws['AC86'] = self.valorCol('Costo por animal', 1, df_fila)
        ws['AN86'] = self.valorCol('Costo por animal', 2, df_fila)
                    

        continuidad3 = df_fila['Sobre la actividad, piensa: Continuidad.3']
        if pd.notna(continuidad3):
            if continuidad3 == 'Continuar con la actividad':
                ws['G92'] = 'X'
                ws['AP92'] = 'X'
            elif continuidad3 == 'Finalizar la actividad':
                ws['I92'] = 'X'
                ws['AN92'] = 'X'  
        else:
            print('Campo vacío') 

        produccion3 = df_fila['Sobre la actividad, piensa: Producción.3']
        if pd.notna(produccion3):
            if produccion3 == 'Ampliar la producción':
                ws['Q92'] = 'X'
                ws['AF92'] = 'X'               
            elif produccion3 == 'Permanecer con la misma producción':
                ws['S92'] = 'X'
                ws['AD92'] = 'X'   
        else:
            print('Campo vacío') 

        ws['AS92'] = df_fila['¿Por qué.3']

        # Nombre actividad 
        for i in range(1, 14):
            ws[f'F{98 + i}'] = self.valorCol('Nombre', i, df_fila)

        for i in range(13):
            valor_unidad = self.valorCol('Unidad de Medida', i, df_fila)
            fila = 99 + i
            if pd.notna(valor_unidad):
                if valor_unidad == 'Kg':
                    ws[f'M{fila}'] = 'X'
                elif valor_unidad == 'Bulto':
                    ws[f'Q{fila}'] = 'X'
                elif valor_unidad == '@':
                    ws[f'S{fila}'] = 'X'
                elif valor_unidad == 'Lt':
                    ws[f'U{fila}'] = 'X'
                elif valor_unidad == 'Gn':
                    ws[f'W{fila}'] = 'X'
            else:
                print('Campo vacío')

        for i in range(13):
            
            ws[f'X{99 + i}'] = self.valorCol('Nombre', i, df_fila)

        for i in range(13):
            valor_frec = self.valorCol('Frecuencia de compra', i, df_fila)
            fila = 99 + i
            if pd.notna(valor_frec):
                if valor_frec == 'Única':
                    ws[f'AC{fila}'] = 'X'
                elif valor_frec == 'Diario':
                    ws[f'AE{fila}'] = 'X'
                elif valor_frec == 'Semanal':
                    ws[f'AG{fila}'] = 'X'
                elif valor_frec == 'Quincenal':
                    ws[f'AJ{fila}'] = 'X'
                elif valor_frec == 'Mensual':
                    ws[f'AL{fila}'] = 'X'
                elif valor_frec == 'Trimestral':
                    ws[f'AN{fila}'] = 'X'            
            else:
                print('Campo vacío')

        for i in range(13):
            
            ws[f'AQ{99 + i}'] = self.valorCol('Precio compra/unidad', i, df_fila)      

        for i in range(13):
            
            ws[f'AT{99 + i}'] = self.valorCol('Lugar de compra', i, df_fila)   
        
        ws['L112'] = df_fila['¿Dónde guarda los productos?']

        ws['L113'] = df_fila['Procedencia de los compradores']

        agua_fuente = df_fila['¿De dónde se abastece del recurso hídrico?']
        if pd.notna(agua_fuente):
            if agua_fuente == 'Aljibe':
                ws['P114'] = 'X'
            elif agua_fuente == 'Acueducto veredal':
                ws['AC114'] = 'X'
            elif agua_fuente == 'Otro':
                ws['AJ114'] = 'X'
                ws['AP114'] = df_fila['¿Cuál?.1']    

        ws['T115'] = df_fila['Forma de extracción']

        ws['AN115'] = df_fila['Cantidad estimada (m3)']

        alcantarillado = df_fila['¿Cuenta con servicio de alcantarillado?']
        if pd.notna(alcantarillado):
            if alcantarillado == 'Sí':
                ws['U116'] = 'X'
                ws['AM116'] = df_fila['¿Cuál?.2']  
            elif alcantarillado == 'No':
                ws['AC116'] = 'X'

        energia = df_fila['¿Qué tipo de energía utiliza?']
        if energia == 'Energía Eléctrica':
            ws['T117'] = 'X'
        elif energia == 'Energía Solar':
            ws['AI117'] = 'X'
        elif energia == 'Otro':
            ws['AP117'] = df_fila['¿Cuál?.3']


        energia_coccion = df_fila['¿De dónde proviene la energía que utiliza para la cocción de alimentos?']
        if energia_coccion == 'Energía Eléctrica':
            ws['AB118'] = 'X'
        elif energia_coccion == 'Leña':
            ws['AG118'] = 'X'
        elif energia_coccion == 'Gas':
            ws['AK118'] = 'X'
        elif energia_coccion == 'Otro':       
            ws['AQ118'] = df_fila['¿Cuál?.4']

        ws['L119'] = df_fila['¿Cuál es el manejo de aguas residuales y sólidos?']

        mano_obra = df_fila['Contrata algún tipo de mano de obra']
        if mano_obra == 'Sí':
            ws['U120'] = 'X'
        elif mano_obra == 'No':
            ws['AC120'] = 'X'  

        # E. Información Laboral 

        for i in range(10):
            prefijo_persona = 124 + i
            ws[f'E{prefijo_persona}'] = self.valorCol('Cargo', i, df_fila)
            ws[f'J{prefijo_persona}'] = self.valorCol('Edad (años)', i, df_fila)
            ws[f'K{prefijo_persona}'] = self.valorCol('Duración jornada (horas)', i, df_fila)

            manoObra = self.valorCol('Tipo de mano de obra', i, df_fila)
            if pd.notna(manoObra):
                if manoObra == 'Familiar':
                    ws[f'B{prefijo_persona}'] = 'X'
                elif manoObra == 'Contratado':
                    ws[f'D{prefijo_persona}20'] = 'X'

            # Genero
            genero = self.valorCol('Género', i,df_fila)
            if pd.notna(genero):
                if genero == 'Masculino':
                    ws[f'I{prefijo_persona}20'] = 'X'
                elif genero ==  'Femenino':
                    ws[f'G{prefijo_persona}'] = 'X'

            # Escolaridad 
            escolaridad = self.valorCol('Escolaridad', i, df_fila)
            if pd.notna(escolaridad):
                if escolaridad:
                    if escolaridad == 'Primaria':
                        ws[f'M{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Bachillerato':
                        ws[f'O{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Técnico':
                        ws[f'Q{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Pregrado':
                        ws[f'S{prefijo_persona}'] = 'X'
                    elif escolaridad == 'Posgrado':
                        ws[f'U{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Contrato 
            contrato = self.valorCol('Contrato', i,df_fila)
            if contrato:
                if contrato == 'Tem.':
                    ws[f'AA{prefijo_persona}'] = self.valorCol('Contrato', i, df_fila)
                elif contrato == 'Fij':
                    ws[f'AC{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Pago de seguridad social 
            pago_seguridad = self.valorCol('Pago de seguridad social', i, df_fila)
            if pago_seguridad:
                if pago_seguridad == 'Si':
                    ws[f'AE{prefijo_persona}'] = 'X'
                elif pago_seguridad == 'No':
                    ws[f'AG{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Remuneración 
            remuneracion = self.valorCol('Remuneración', i, df_fila)
            if remuneracion:
                if remuneracion == 'Inferiores a $900.000':
                    ws[f'AS{prefijo_persona}'] = 'X'
                elif remuneracion == '$900.000 a $1.800.000':
                    ws[f'AT{prefijo_persona}'] = 'X'
                elif remuneracion == '$1.801.000 a $2.700.000':
                    ws[f'AU{prefijo_persona}'] = 'X'
                elif remuneracion == 'Superiores a $2.701.000':
                    ws[f'AV{prefijo_persona}'] = 'X'
            else:
                print(f'Campo vacio')

            # Información adicional
            
            ws[f'AH{prefijo_persona}'] = self.valorCol('Procedencia', i, df_fila)
            ws[f'AJ{prefijo_persona}'] = self.valorCol('Residencia', i, df_fila)
            ws[f'AL{prefijo_persona}'] = self.valorCol('Tiempo trabajado', i, df_fila)
            ws[f'AM{prefijo_persona}'] = self.valorCol('# Personas núcleo familiar', i, df_fila)
            ws[f'AN{prefijo_persona}'] = self.valorCol('Personas a cargo', i, df_fila)
            ws[f'AO{prefijo_persona}'] = self.valorCol('Lugar de residencia familiar', i,df_fila)

            # ¿Contrata persona por jornal? 
        contrata_persona = df_fila['¿Contrata persona por jornal?']
        if pd.notna(contrata_persona):
            if contrata_persona == 'Sí':
                ws['N80'] = 'X'
            elif contrata_persona == 'No':
                ws['P80'] = 'X'

            # Pregunta 81: Tipo de obra o labor
        ws['A138'] = df_fila.get('Tipo de obra o labor', '')

        # Pregunta 82: Frecuencia de contratación /año
        ws['G138'] = df_fila.get('Frecuencia de contratación/año', '')

        # Pregunta 83: Duración en jornales del contrato
        ws['Q138'] = df_fila.get('Duración en Jornales del contrato', '')

        # Pregunta 84: Valor del jornal
        ws['Y138'] = df_fila.get('Valor del jornal', '')

        # Pregunta 85: Cantidad de jornaleros empleados por contrato
        ws['AG138'] = df_fila.get('Cantidad de jornaleros empleados por contrato', '')

        # Pregunta 86: Residencia de los jornaleros
        ws['AN138'] = df_fila.get('Residencia de los jornaleros', '')